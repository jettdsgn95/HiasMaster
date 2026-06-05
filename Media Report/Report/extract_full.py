import sys
import io
import json
import re
from datetime import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl import load_workbook

wb = load_workbook('2026_[MEDIA - TASKS TRACKER].xlsx')
ws = wb['All Task']

def extract_computed(val):
    """Extract computed value from IFERROR formula strings"""
    if val is None:
        return None
    val = str(val)
    if 'COMPUTED_VALUE' in val or 'IFERROR' in val:
        m = re.search(r'COMPUTED_VALUE[^,]*,\"(.*?)\"\)', val)
        if m:
            return m.group(1)
        # fallback
        m = re.search(r',\"([^\"]+)\"\)$', val)
        if m:
            return m.group(1)
        return None
    return val

# Column indices (0-based):
# 0: TT (index)
# 1: Task Level 1 (project name)
# 2: Task Level 2 (subtask)
# 3: Mức độ ưu tiên (priority)
# 4: PIC
# 5: (empty)
# 6: Support
# 7: Start
# 8: End
# 9: Ngày gia hạn (extended date)
# 10: Ngày còn (days remaining)
# 11: Lý do gia hạn (extension reason)
# 12: Nội dung (content)
# 13: Trạng thái (status)
# 14: % Hoàn thành (completion %)
# 15: Huỷ/Hoãn (cancel/postpone)
# 16: State

HEADERS = {
    'tt': 0,
    'level1': 1,
    'level2': 2,
    'priority': 3,
    'pic': 4,
    'support': 6,
    'start': 7,
    'end': 8,
    'extended_date': 9,
    'days_left': 10,
    'extension_reason': 11,
    'content': 12,
    'status': 13,
    'completion': 14,
    'cancel': 15,
    'state': 16,
}

projects = []
current_project = None

for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
    if row_idx < 4:  # skip header rows
        continue
    
    cells = []
    for cell in row:
        if isinstance(cell, datetime):
            cells.append(cell.strftime('%Y-%m-%d'))
        elif cell is None:
            cells.append(None)
        else:
            val = str(cell)
            if 'COMPUTED_VALUE' in val or 'IFERROR' in val:
                val = extract_computed(val)
            cells.append(val)
    
    # Check if this is a project row (Level 1) or subtask row (Level 2)
    tt = cells[HEADERS['tt']] if HEADERS['tt'] < len(cells) else None
    level1 = cells[HEADERS['level1']] if HEADERS['level1'] < len(cells) else None
    level2 = cells[HEADERS['level2']] if HEADERS['level2'] < len(cells) else None
    pic = cells[HEADERS['pic']] if HEADERS['pic'] < len(cells) else None
    start = cells[HEADERS['start']] if HEADERS['start'] < len(cells) else None
    end = cells[HEADERS['end']] if HEADERS['end'] < len(cells) else None
    extended = cells[HEADERS['extended_date']] if HEADERS['extended_date'] < len(cells) else None
    completion = cells[HEADERS['completion']] if HEADERS['completion'] < len(cells) else None
    state = cells[HEADERS['state']] if HEADERS['state'] < len(cells) else None
    cancel = cells[HEADERS['cancel']] if HEADERS['cancel'] < len(cells) else None
    content = cells[HEADERS['content']] if HEADERS['content'] < len(cells) else None
    priority = cells[HEADERS['priority']] if HEADERS['priority'] < len(cells) else None
    
    # Skip completely empty rows
    if all(c is None for c in cells):
        continue
    
    # Project row: has level1 but no level2, TT is a number
    if level1 and not level2:
        current_project = {
            'id': tt,
            'name': level1,
            'tasks': []
        }
        projects.append(current_project)
    # Subtask row: has level2
    elif level2 and current_project is not None:
        task = {
            'tt': tt,
            'name': level2,
            'pic': pic,
            'start': start,
            'end': end,
            'extended': extended,
            'completion': completion,
            'state': state,
            'cancel': cancel,
            'content': content,
            'priority': priority,
        }
        current_project['tasks'].append(task)

# Dump to JSON
output = json.dumps(projects, ensure_ascii=False, indent=2)
with open('tasks_data.json', 'w', encoding='utf-8') as f:
    f.write(output)

print(f"Extracted {len(projects)} projects")
total_tasks = sum(len(p['tasks']) for p in projects)
print(f"Total tasks: {total_tasks}")

# Stats
states = {}
for p in projects:
    for t in p['tasks']:
        s = t.get('state') or 'Unknown'
        states[s] = states.get(s, 0) + 1
print("States:", states)

# PICs
pics = {}
for p in projects:
    for t in p['tasks']:
        pic = t.get('pic') or 'Unknown'
        for name in pic.split(','):
            name = name.strip()
            if name:
                pics[name] = pics.get(name, 0) + 1
print("PICs:", pics)
