import sys
import io
import json
import re
from datetime import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl import load_workbook

wb = load_workbook('2026_[MEDIA - TASKS TRACKER].xlsx')
ws = wb['All Task']

def extract_computed(val):
    if val is None:
        return None
    val = str(val)
    if 'COMPUTED_VALUE' in val or 'IFERROR' in val:
        m = re.search(r'COMPUTED_VALUE[^,]*,\"(.*?)\"\)', val)
        if m:
            return m.group(1)
        m = re.search(r',\"([^\"]+)\"\)$', val)
        if m:
            return m.group(1)
        return None
    return val

HEADERS = {
    'tt': 0, 'level1': 1, 'level2': 2, 'priority': 3,
    'pic': 4, 'support': 6, 'start': 7, 'end': 8,
    'extended_date': 9, 'days_left': 10, 'extension_reason': 11,
    'content': 12, 'status': 13, 'completion': 14,
    'cancel': 15, 'state': 16,
}

projects = []
current_project = None

for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
    if row_idx < 4:
        continue
    cells = []
    for cell in row:
        if isinstance(cell, datetime):
            cells.append(cell.strftime('%Y-%m-%d'))
        elif cell is None:
            cells.append(None)
        else:
            val = str(cell)
            if 'COMPUTED_VALUE' in val or 'IFERROR' in val:
                val = extract_computed(val)
            cells.append(val)

    if all(c is None for c in cells):
        continue

    def get(key):
        idx = HEADERS.get(key, -1)
        return cells[idx] if 0 <= idx < len(cells) else None

    level1 = get('level1')
    level2 = get('level2')
    tt = get('tt')

    if level1 and not level2:
        current_project = {'id': tt, 'name': level1, 'tasks': []}
        projects.append(current_project)
    elif level2 and current_project is not None:
        try:
            comp_val = float(get('completion')) if get('completion') else 0
        except:
            comp_val = 0
        task = {
            'tt': tt, 'name': level2,
            'pic': get('pic'), 'start': get('start'), 'end': get('end'),
            'extended': get('extended_date'), 'completion': comp_val,
            'state': get('state'), 'cancel': get('cancel'),
            'content': get('content'), 'priority': get('priority'),
        }
        current_project['tasks'].append(task)

# ---- Stats ----
total_tasks = sum(len(p['tasks']) for p in projects)
states = {'Hoàn thành': 0, 'Đang làm': 0, 'Hoãn': 0, 'Quá hạn': 0, 'Khác': 0}
for p in projects:
    for t in p['tasks']:
        s = t.get('state') or 'Khác'
        if s in states:
            states[s] += 1
        else:
            states['Khác'] += 1

pics = {}
for p in projects:
    for t in p['tasks']:
        pic = t.get('pic') or ''
        for name in pic.split(','):
            name = name.strip()
            if name:
                pics[name] = pics.get(name, 0) + 1

# JSON encode for JS
projects_json = json.dumps(projects, ensure_ascii=False)

# ---- Generate HTML ----
html = '''<!DOCTYPE html>
<html lang="vi">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>BÁO CÁO CÔNG VIỆC 2026 – TEAM MEDIA</title>
<meta name="description" content="Báo cáo tổng hợp công việc Team Media – Phòng Marketing 2026. Theo dõi tiến độ, trạng thái và phân công nhiệm vụ.">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap" rel="stylesheet">
<style>
:root {
  --bg-primary: #0a0e1a;
  --bg-secondary: #0f1628;
  --bg-card: #141d35;
  --bg-card-hover: #1a2540;
  --accent-purple: #7c3aed;
  --accent-purple-light: #a855f7;
  --accent-blue: #3b82f6;
  --accent-cyan: #06b6d4;
  --accent-pink: #ec4899;
  --accent-green: #10b981;
  --accent-orange: #f59e0b;
  --accent-red: #ef4444;
  --text-primary: #f1f5f9;
  --text-secondary: #94a3b8;
  --text-muted: #475569;
  --border: rgba(148, 163, 184, 0.1);
  --border-bright: rgba(124, 58, 237, 0.4);
  --shadow-glow: 0 0 40px rgba(124, 58, 237, 0.15);
  --radius: 16px;
  --radius-sm: 8px;
}

* { margin: 0; padding: 0; box-sizing: border-box; }

body {
  font-family: 'Inter', sans-serif;
  background: var(--bg-primary);
  color: var(--text-primary);
  min-height: 100vh;
  overflow-x: hidden;
}

/* ---- BACKGROUND ---- */
body::before {
  content: '';
  position: fixed; inset: 0; z-index: 0;
  background:
    radial-gradient(ellipse 80% 60% at 20% -10%, rgba(124,58,237,0.15) 0%, transparent 60%),
    radial-gradient(ellipse 60% 40% at 80% 100%, rgba(6,182,212,0.1) 0%, transparent 60%),
    radial-gradient(ellipse 40% 30% at 50% 50%, rgba(59,130,246,0.05) 0%, transparent 60%);
  pointer-events: none;
}

/* ---- LAYOUT ---- */
.container { position: relative; z-index: 1; max-width: 1400px; margin: 0 auto; padding: 0 24px; }

/* ---- HEADER ---- */
header {
  padding: 40px 0 32px;
  border-bottom: 1px solid var(--border);
  margin-bottom: 40px;
}

.header-content {
  display: flex;
  align-items: center;
  justify-content: space-between;
  flex-wrap: wrap;
  gap: 20px;
}

.header-left {}

.badge-team {
  display: inline-flex; align-items: center; gap: 8px;
  background: linear-gradient(135deg, rgba(124,58,237,0.2), rgba(59,130,246,0.2));
  border: 1px solid var(--border-bright);
  border-radius: 50px; padding: 6px 16px;
  font-size: 12px; font-weight: 600; letter-spacing: 1.5px;
  color: var(--accent-purple-light);
  text-transform: uppercase;
  margin-bottom: 12px;
}

.badge-team::before {
  content: ''; width: 6px; height: 6px;
  border-radius: 50%; background: var(--accent-purple-light);
  animation: pulse 2s ease-in-out infinite;
}

@keyframes pulse {
  0%, 100% { opacity: 1; transform: scale(1); }
  50% { opacity: 0.5; transform: scale(0.8); }
}

h1 {
  font-size: clamp(28px, 4vw, 48px);
  font-weight: 900;
  line-height: 1.1;
  background: linear-gradient(135deg, #f1f5f9 0%, var(--accent-purple-light) 50%, var(--accent-cyan) 100%);
  -webkit-background-clip: text;
  -webkit-text-fill-color: transparent;
  background-clip: text;
  margin-bottom: 8px;
}

.header-subtitle {
  color: var(--text-secondary);
  font-size: 15px;
  font-weight: 400;
}

.header-meta {
  display: flex; flex-direction: column; align-items: flex-end; gap: 8px;
}

.report-date {
  font-size: 13px; color: var(--text-secondary);
  background: var(--bg-card); border: 1px solid var(--border);
  padding: 8px 16px; border-radius: 50px;
}

/* ---- STATS GRID ---- */
.stats-grid {
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
  gap: 16px;
  margin-bottom: 40px;
}

.stat-card {
  background: var(--bg-card);
  border: 1px solid var(--border);
  border-radius: var(--radius);
  padding: 24px;
  position: relative;
  overflow: hidden;
  transition: transform 0.2s, border-color 0.2s, box-shadow 0.2s;
  cursor: default;
}

.stat-card::before {
  content: '';
  position: absolute; top: 0; left: 0; right: 0; height: 3px;
  border-radius: var(--radius) var(--radius) 0 0;
}

.stat-card:hover {
  transform: translateY(-4px);
  box-shadow: var(--shadow-glow);
  border-color: var(--border-bright);
}

.stat-card.purple::before { background: linear-gradient(90deg, var(--accent-purple), var(--accent-purple-light)); }
.stat-card.green::before  { background: linear-gradient(90deg, #059669, var(--accent-green)); }
.stat-card.blue::before   { background: linear-gradient(90deg, var(--accent-blue), var(--accent-cyan)); }
.stat-card.orange::before { background: linear-gradient(90deg, #d97706, var(--accent-orange)); }
.stat-card.red::before    { background: linear-gradient(90deg, #dc2626, var(--accent-red)); }
.stat-card.pink::before   { background: linear-gradient(90deg, #db2777, var(--accent-pink)); }

.stat-icon {
  font-size: 28px; margin-bottom: 12px;
}

.stat-value {
  font-size: 40px; font-weight: 800; line-height: 1;
  margin-bottom: 6px;
}

.stat-card.purple .stat-value { color: var(--accent-purple-light); }
.stat-card.green .stat-value  { color: var(--accent-green); }
.stat-card.blue .stat-value   { color: var(--accent-cyan); }
.stat-card.orange .stat-value { color: var(--accent-orange); }
.stat-card.red .stat-value    { color: var(--accent-red); }
.stat-card.pink .stat-value   { color: var(--accent-pink); }

.stat-label {
  font-size: 13px; color: var(--text-secondary); font-weight: 500;
}

/* ---- CHARTS ROW ---- */
.charts-row {
  display: grid;
  grid-template-columns: 1fr 1fr;
  gap: 20px;
  margin-bottom: 40px;
}

@media (max-width: 768px) { .charts-row { grid-template-columns: 1fr; } }

.chart-card {
  background: var(--bg-card);
  border: 1px solid var(--border);
  border-radius: var(--radius);
  padding: 28px;
}

.chart-card h3 {
  font-size: 15px; font-weight: 700; color: var(--text-primary);
  margin-bottom: 20px;
  display: flex; align-items: center; gap: 8px;
}

.chart-card h3::before {
  content: '';
  display: inline-block; width: 4px; height: 16px;
  border-radius: 2px;
  background: linear-gradient(180deg, var(--accent-purple), var(--accent-cyan));
}

/* Donut chart */
.donut-wrap {
  display: flex; align-items: center; gap: 24px; flex-wrap: wrap;
}

.donut-svg { flex-shrink: 0; }

.donut-legend {
  display: flex; flex-direction: column; gap: 10px; flex: 1; min-width: 150px;
}

.legend-item {
  display: flex; align-items: center; gap: 10px;
  font-size: 13px;
}

.legend-dot {
  width: 10px; height: 10px; border-radius: 50%; flex-shrink: 0;
}

.legend-label { color: var(--text-secondary); flex: 1; }
.legend-value { font-weight: 700; font-size: 14px; }

/* Bar chart */
.bar-list {
  display: flex; flex-direction: column; gap: 14px;
}

.bar-item {}

.bar-header {
  display: flex; justify-content: space-between; align-items: center;
  margin-bottom: 6px;
}

.bar-name { font-size: 13px; font-weight: 600; color: var(--text-primary); }
.bar-count { font-size: 13px; color: var(--text-secondary); }

.bar-track {
  height: 8px; background: rgba(255,255,255,0.05); border-radius: 50px; overflow: hidden;
}

.bar-fill {
  height: 100%; border-radius: 50px;
  transition: width 1.2s cubic-bezier(0.16, 1, 0.3, 1);
}

/* ---- CONTROLS ---- */
.controls {
  display: flex; gap: 12px; flex-wrap: wrap; align-items: center;
  margin-bottom: 24px;
}

.search-box {
  flex: 1; min-width: 260px;
  position: relative;
}

.search-box input {
  width: 100%;
  background: var(--bg-card);
  border: 1px solid var(--border);
  border-radius: 50px;
  padding: 12px 20px 12px 44px;
  font-size: 14px; font-family: 'Inter', sans-serif;
  color: var(--text-primary);
  outline: none;
  transition: border-color 0.2s, box-shadow 0.2s;
}

.search-box input:focus {
  border-color: var(--accent-purple);
  box-shadow: 0 0 0 3px rgba(124,58,237,0.2);
}

.search-icon {
  position: absolute; left: 16px; top: 50%; transform: translateY(-50%);
  color: var(--text-muted); font-size: 16px; pointer-events: none;
}

.filter-group {
  display: flex; gap: 8px; flex-wrap: wrap;
}

.filter-btn {
  padding: 10px 18px;
  border-radius: 50px;
  border: 1px solid var(--border);
  background: var(--bg-card);
  color: var(--text-secondary);
  font-size: 13px; font-weight: 500;
  cursor: pointer; font-family: 'Inter', sans-serif;
  transition: all 0.2s;
  white-space: nowrap;
}

.filter-btn:hover { border-color: var(--accent-purple); color: var(--text-primary); }

.filter-btn.active {
  background: linear-gradient(135deg, var(--accent-purple), var(--accent-blue));
  border-color: transparent;
  color: #fff;
  box-shadow: 0 4px 15px rgba(124,58,237,0.4);
}

/* ---- PROJECT SECTION ---- */
.results-summary {
  font-size: 13px; color: var(--text-secondary); margin-bottom: 20px;
}

.results-summary strong { color: var(--text-primary); }

.projects-list { display: flex; flex-direction: column; gap: 16px; }

.project-card {
  background: var(--bg-card);
  border: 1px solid var(--border);
  border-radius: var(--radius);
  overflow: hidden;
  transition: border-color 0.2s;
}

.project-card:hover { border-color: rgba(124,58,237,0.3); }

.project-header {
  display: flex; align-items: center; justify-content: space-between;
  padding: 20px 24px; cursor: pointer;
  gap: 16px;
  user-select: none;
}

.project-header:hover { background: var(--bg-card-hover); }

.project-title-group {
  display: flex; align-items: center; gap: 16px; flex: 1; min-width: 0;
}

.project-num {
  font-size: 11px; font-weight: 700; color: var(--accent-purple-light);
  background: rgba(124,58,237,0.15); border: 1px solid rgba(124,58,237,0.3);
  padding: 4px 10px; border-radius: 50px; flex-shrink: 0;
  letter-spacing: 0.5px;
}

.project-name {
  font-size: 15px; font-weight: 700; color: var(--text-primary);
  overflow: hidden; text-overflow: ellipsis; white-space: nowrap;
}

.project-meta {
  display: flex; align-items: center; gap: 12px; flex-shrink: 0;
}

.task-count-badge {
  font-size: 12px; color: var(--text-secondary);
  background: rgba(255,255,255,0.05);
  padding: 4px 12px; border-radius: 50px;
  white-space: nowrap;
}

.project-progress-wrap {
  display: flex; align-items: center; gap: 10px; min-width: 120px;
}

.mini-progress {
  flex: 1; height: 6px; background: rgba(255,255,255,0.08); border-radius: 50px; overflow: hidden;
}

.mini-progress-fill {
  height: 100%; border-radius: 50px;
  background: linear-gradient(90deg, var(--accent-purple), var(--accent-cyan));
}

.progress-pct {
  font-size: 12px; font-weight: 700; color: var(--accent-cyan); min-width: 36px; text-align: right;
}

.chevron {
  color: var(--text-muted); font-size: 18px; transition: transform 0.3s;
  flex-shrink: 0;
}

.project-card.open .chevron { transform: rotate(180deg); }

/* ---- TASK TABLE ---- */
.task-table-wrap {
  max-height: 0; overflow: hidden; transition: max-height 0.4s ease;
  border-top: 1px solid transparent;
}

.project-card.open .task-table-wrap {
  max-height: 2000px;
  border-top-color: var(--border);
}

.task-table {
  width: 100%; border-collapse: collapse; font-size: 13px;
}

.task-table th {
  padding: 12px 16px;
  text-align: left;
  font-size: 11px; font-weight: 700; letter-spacing: 1px;
  text-transform: uppercase;
  color: var(--text-muted);
  background: rgba(0,0,0,0.2);
  white-space: nowrap;
}

.task-table td {
  padding: 12px 16px;
  border-top: 1px solid var(--border);
  vertical-align: middle;
}

.task-table tr:hover td { background: rgba(255,255,255,0.02); }

.task-name { font-weight: 500; color: var(--text-primary); }
.task-pic  { color: var(--accent-cyan); font-weight: 600; }

.task-date {
  color: var(--text-secondary); white-space: nowrap; font-size: 12px;
}

.task-date.extended { color: var(--accent-orange); }

/* Status badges */
.status-badge {
  display: inline-flex; align-items: center; gap: 5px;
  padding: 3px 10px; border-radius: 50px;
  font-size: 11px; font-weight: 700; white-space: nowrap;
}

.status-badge.done     { background: rgba(16,185,129,0.15); color: var(--accent-green); border: 1px solid rgba(16,185,129,0.3); }
.status-badge.doing    { background: rgba(59,130,246,0.15); color: var(--accent-blue);  border: 1px solid rgba(59,130,246,0.3); }
.status-badge.late     { background: rgba(239,68,68,0.15);  color: var(--accent-red);   border: 1px solid rgba(239,68,68,0.3); }
.status-badge.paused   { background: rgba(245,158,11,0.15); color: var(--accent-orange);border: 1px solid rgba(245,158,11,0.3); }
.status-badge.canceled { background: rgba(148,163,184,0.1); color: var(--text-secondary);border: 1px solid var(--border); }

.task-completion-bar {
  width: 60px; height: 5px; background: rgba(255,255,255,0.08); border-radius: 50px; overflow: hidden;
  display: inline-block; vertical-align: middle; margin-right: 6px;
}

.task-completion-fill {
  height: 100%; border-radius: 50px;
  background: linear-gradient(90deg, var(--accent-purple), var(--accent-cyan));
}

/* ---- PAGINATION ---- */
.pagination {
  display: flex; justify-content: center; align-items: center;
  gap: 8px; margin-top: 32px; flex-wrap: wrap;
}

.page-btn {
  width: 38px; height: 38px; border-radius: 8px;
  border: 1px solid var(--border);
  background: var(--bg-card);
  color: var(--text-secondary);
  font-size: 14px; cursor: pointer;
  display: flex; align-items: center; justify-content: center;
  font-family: 'Inter', sans-serif;
  transition: all 0.2s;
}

.page-btn:hover { border-color: var(--accent-purple); color: var(--text-primary); }
.page-btn.active {
  background: linear-gradient(135deg, var(--accent-purple), var(--accent-blue));
  border-color: transparent; color: #fff;
  box-shadow: 0 4px 12px rgba(124,58,237,0.4);
}

.page-btn.wide { width: auto; padding: 0 14px; }

/* ---- FOOTER ---- */
footer {
  margin-top: 60px; padding: 32px 0; border-top: 1px solid var(--border);
  text-align: center; color: var(--text-muted); font-size: 13px;
}

/* ---- RESPONSIVE ---- */
@media (max-width: 768px) {
  .project-progress-wrap { display: none; }
  .task-table th:nth-child(n+4), .task-table td:nth-child(n+4) { display: none; }
}

/* ---- ANIMATIONS ---- */
@keyframes fadeInUp {
  from { opacity: 0; transform: translateY(20px); }
  to   { opacity: 1; transform: translateY(0); }
}

.stat-card { animation: fadeInUp 0.5s ease both; }
.stat-card:nth-child(1) { animation-delay: 0.05s; }
.stat-card:nth-child(2) { animation-delay: 0.1s; }
.stat-card:nth-child(3) { animation-delay: 0.15s; }
.stat-card:nth-child(4) { animation-delay: 0.2s; }
.stat-card:nth-child(5) { animation-delay: 0.25s; }
.stat-card:nth-child(6) { animation-delay: 0.3s; }

/* ---- NO RESULTS ---- */
.no-results {
  text-align: center; padding: 60px 20px;
  color: var(--text-muted); font-size: 15px;
}
.no-results .icon { font-size: 48px; margin-bottom: 12px; }

/* Scrollbar */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: var(--bg-primary); }
::-webkit-scrollbar-thumb { background: rgba(124,58,237,0.4); border-radius: 3px; }

</style>
</head>
<body>

<div class="container">

  <!-- ======== HEADER ======== -->
  <header>
    <div class="header-content">
      <div class="header-left">
        <div class="badge-team">Team Media &nbsp;·&nbsp; Phòng Marketing</div>
        <h1>BÁO CÁO CÔNG VIỆC 2026</h1>
        <p class="header-subtitle">Tổng hợp tiến độ & phân công nhiệm vụ toàn năm</p>
      </div>
      <div class="header-meta">
        <div class="report-date" id="reportDate"></div>
      </div>
    </div>
  </header>

  <!-- ======== STATS ======== -->
  <div class="stats-grid" id="statsGrid"></div>

  <!-- ======== CHARTS ======== -->
  <div class="charts-row">
    <div class="chart-card">
      <h3>Trạng thái công việc</h3>
      <div class="donut-wrap">
        <svg class="donut-svg" id="donutSvg" width="140" height="140" viewBox="0 0 140 140"></svg>
        <div class="donut-legend" id="donutLegend"></div>
      </div>
    </div>
    <div class="chart-card">
      <h3>Phân công theo nhân sự</h3>
      <div class="bar-list" id="picBars"></div>
    </div>
  </div>

  <!-- ======== CONTROLS ======== -->
  <div class="controls">
    <div class="search-box">
      <span class="search-icon">🔍</span>
      <input type="text" id="searchInput" placeholder="Tìm dự án, công việc, nhân sự..." autocomplete="off">
    </div>
    <div class="filter-group" id="filterGroup">
      <button class="filter-btn active" data-filter="all" id="btn-all">Tất cả</button>
      <button class="filter-btn" data-filter="Hoàn thành" id="btn-done">✅ Hoàn thành</button>
      <button class="filter-btn" data-filter="Đang làm" id="btn-doing">🔵 Đang làm</button>
      <button class="filter-btn" data-filter="Hoãn" id="btn-paused">⏸ Hoãn</button>
    </div>
  </div>

  <div class="results-summary" id="resultsSummary"></div>

  <!-- ======== PROJECTS LIST ======== -->
  <div class="projects-list" id="projectsList"></div>

  <!-- ======== PAGINATION ======== -->
  <div class="pagination" id="pagination"></div>

  <!-- ======== FOOTER ======== -->
  <footer>
    <p>📊 Báo cáo được tạo tự động từ dữ liệu Task Tracker · Team Media · HIAS Master</p>
  </footer>

</div>

<script>
// ===================== DATA =====================
const ALL_PROJECTS = ''' + projects_json + ''';

// ===================== UTILS =====================
function stateClass(state, cancel) {
  if (cancel) return 'canceled';
  if (!state) return 'doing';
  if (state.includes('Hoàn thành')) return 'done';
  if (state.includes('Hoãn')) return 'paused';
  if (state.includes('Quá hạn')) return 'late';
  if (state.includes('Đang làm')) return 'doing';
  return 'doing';
}

function stateLabel(state, cancel) {
  if (cancel) return '🚫 ' + cancel;
  if (!state) return '🔵 Đang làm';
  if (state.includes('Hoàn thành')) return '✅ Hoàn thành';
  if (state.includes('Hoãn')) return '⏸ Hoãn';
  if (state.includes('Quá hạn')) return '🔴 Quá hạn';
  if (state.includes('Đang làm')) return '🔵 Đang làm';
  return state;
}

function projectCompletion(project) {
  const tasks = project.tasks;
  if (!tasks.length) return 0;
  const done = tasks.filter(t => stateClass(t.state, t.cancel) === 'done').length;
  return Math.round(done / tasks.length * 100);
}

function projectMatchesFilter(project, filter) {
  if (filter === 'all') return true;
  return project.tasks.some(t => (t.state || '').includes(filter));
}

function projectMatchesSearch(project, query) {
  if (!query) return true;
  query = query.toLowerCase();
  if (project.name.toLowerCase().includes(query)) return true;
  return project.tasks.some(t =>
    (t.name || '').toLowerCase().includes(query) ||
    (t.pic  || '').toLowerCase().includes(query)
  );
}

// ===================== STATS =====================
function renderStats() {
  const totalProjects = ALL_PROJECTS.length;
  const totalTasks = ALL_PROJECTS.reduce((s, p) => s + p.tasks.length, 0);
  let done = 0, doing = 0, paused = 0, late = 0;
  ALL_PROJECTS.forEach(p => p.tasks.forEach(t => {
    const sc = stateClass(t.state, t.cancel);
    if (sc === 'done') done++;
    else if (sc === 'doing') doing++;
    else if (sc === 'paused') paused++;
    else if (sc === 'late') late++;
    else if (sc === 'canceled') paused++;
  }));
  const pct = Math.round(done / totalTasks * 100);

  const cards = [
    { icon: '📁', value: totalProjects, label: 'Dự án', cls: 'purple' },
    { icon: '📋', value: totalTasks,    label: 'Công việc', cls: 'blue' },
    { icon: '✅', value: done,          label: 'Hoàn thành', cls: 'green' },
    { icon: '🔵', value: doing,         label: 'Đang làm', cls: 'blue' },
    { icon: '⏸',  value: paused,        label: 'Hoãn/Huỷ', cls: 'orange' },
    { icon: '📈', value: pct + '%',     label: 'Tỉ lệ hoàn thành', cls: 'pink' },
  ];

  document.getElementById('statsGrid').innerHTML = cards.map(c => `
    <div class="stat-card ${c.cls}">
      <div class="stat-icon">${c.icon}</div>
      <div class="stat-value">${c.value}</div>
      <div class="stat-label">${c.label}</div>
    </div>
  `).join('');
}

// ===================== DONUT =====================
function renderDonut() {
  let done = 0, doing = 0, paused = 0, late = 0;
  ALL_PROJECTS.forEach(p => p.tasks.forEach(t => {
    const sc = stateClass(t.state, t.cancel);
    if (sc === 'done') done++;
    else if (sc === 'doing') doing++;
    else if (sc === 'paused' || sc === 'canceled') paused++;
    else if (sc === 'late') late++;
  }));
  const total = done + doing + paused + late;
  const segments = [
    { label: 'Hoàn thành', value: done,   color: '#10b981' },
    { label: 'Đang làm',   value: doing,  color: '#3b82f6' },
    { label: 'Hoãn/Huỷ',  value: paused, color: '#f59e0b' },
    { label: 'Quá hạn',   value: late,   color: '#ef4444' },
  ].filter(s => s.value > 0);

  const cx = 70, cy = 70, r = 50, strokeW = 22;
  const circumference = 2 * Math.PI * r;
  let offset = 0;
  let pathsHtml = '';

  segments.forEach(seg => {
    const pct = seg.value / total;
    const dash = circumference * pct;
    const gap  = circumference - dash;
    pathsHtml += `<circle
      cx="${cx}" cy="${cy}" r="${r}"
      fill="none"
      stroke="${seg.color}"
      stroke-width="${strokeW}"
      stroke-dasharray="${dash.toFixed(2)} ${gap.toFixed(2)}"
      stroke-dashoffset="${(-offset).toFixed(2)}"
      transform="rotate(-90 ${cx} ${cy})"
      style="transition: stroke-dasharray 1s ease"
    />`;
    offset += dash;
  });

  pathsHtml += `<text x="${cx}" y="${cy - 6}" text-anchor="middle" fill="#f1f5f9" font-size="20" font-weight="800" font-family="Inter">${Math.round(done/total*100)}%</text>`;
  pathsHtml += `<text x="${cx}" y="${cy + 14}" text-anchor="middle" fill="#94a3b8" font-size="10" font-family="Inter">hoàn thành</text>`;

  document.getElementById('donutSvg').innerHTML = pathsHtml;

  document.getElementById('donutLegend').innerHTML = segments.map(s => `
    <div class="legend-item">
      <span class="legend-dot" style="background:${s.color}"></span>
      <span class="legend-label">${s.label}</span>
      <span class="legend-value" style="color:${s.color}">${s.value}</span>
    </div>
  `).join('');
}

// ===================== PIC BARS =====================
function renderPicBars() {
  const pics = {};
  ALL_PROJECTS.forEach(p => p.tasks.forEach(t => {
    const pic = t.pic || '';
    pic.split(',').forEach(name => {
      name = name.trim();
      if (name) pics[name] = (pics[name] || 0) + 1;
    });
  }));
  const sorted = Object.entries(pics).sort((a, b) => b[1] - a[1]).slice(0, 8);
  const max = sorted[0][1];
  const colors = ['#7c3aed', '#3b82f6', '#06b6d4', '#10b981', '#f59e0b', '#ec4899', '#a855f7', '#14b8a6'];

  document.getElementById('picBars').innerHTML = sorted.map(([name, count], i) => `
    <div class="bar-item">
      <div class="bar-header">
        <span class="bar-name">👤 ${name}</span>
        <span class="bar-count">${count} task</span>
      </div>
      <div class="bar-track">
        <div class="bar-fill" style="width:${Math.round(count/max*100)}%; background:${colors[i]}"></div>
      </div>
    </div>
  `).join('');
}

// ===================== PROJECTS RENDER =====================
const PAGE_SIZE = 20;
let currentFilter = 'all';
let currentSearch = '';
let currentPage  = 1;
let filteredProjects = [];

function getFilteredProjects() {
  return ALL_PROJECTS.filter(p =>
    projectMatchesFilter(p, currentFilter) &&
    projectMatchesSearch(p, currentSearch)
  );
}

function renderProjectCard(project, idx) {
  const pct = projectCompletion(project);
  const tasksDone = project.tasks.filter(t => stateClass(t.state, t.cancel) === 'done').length;
  
  return `
  <div class="project-card" id="proj-${idx}" data-idx="${idx}">
    <div class="project-header" onclick="toggleProject(${idx})">
      <div class="project-title-group">
        <span class="project-num">#${project.id || (idx+1)}</span>
        <span class="project-name">${project.name}</span>
      </div>
      <div class="project-meta">
        <span class="task-count-badge">${tasksDone}/${project.tasks.length} việc</span>
        <div class="project-progress-wrap">
          <div class="mini-progress">
            <div class="mini-progress-fill" style="width:${pct}%"></div>
          </div>
          <span class="progress-pct">${pct}%</span>
        </div>
        <span class="chevron">▼</span>
      </div>
    </div>
    <div class="task-table-wrap">
      <table class="task-table">
        <thead>
          <tr>
            <th>Công việc</th>
            <th>PIC</th>
            <th>Start</th>
            <th>End</th>
            <th>Trạng thái</th>
            <th>%</th>
          </tr>
        </thead>
        <tbody>
          ${project.tasks.map(t => `
            <tr>
              <td class="task-name">${t.name || ''}</td>
              <td class="task-pic">${t.pic || '–'}</td>
              <td class="task-date">${t.start || '–'}</td>
              <td class="task-date${t.extended ? ' extended' : ''}">${t.extended || t.end || '–'}${t.extended ? ' ⤴' : ''}</td>
              <td><span class="status-badge ${stateClass(t.state, t.cancel)}">${stateLabel(t.state, t.cancel)}</span></td>
              <td>
                <div class="task-completion-bar"><div class="task-completion-fill" style="width:${Math.round((t.completion||0)*100)}%"></div></div>
                <span style="font-size:11px;color:var(--text-secondary)">${Math.round((t.completion||0)*100)}%</span>
              </td>
            </tr>
          `).join('')}
        </tbody>
      </table>
    </div>
  </div>`;
}

function renderProjects() {
  filteredProjects = getFilteredProjects();
  const total = filteredProjects.length;
  const totalTasks = filteredProjects.reduce((s, p) => s + p.tasks.length, 0);

  document.getElementById('resultsSummary').innerHTML =
    `Hiển thị <strong>${total}</strong> dự án · <strong>${totalTasks}</strong> công việc`;

  const start = (currentPage - 1) * PAGE_SIZE;
  const pageData = filteredProjects.slice(start, start + PAGE_SIZE);

  if (pageData.length === 0) {
    document.getElementById('projectsList').innerHTML =
      `<div class="no-results"><div class="icon">🔍</div><p>Không tìm thấy kết quả phù hợp</p></div>`;
  } else {
    document.getElementById('projectsList').innerHTML =
      pageData.map((p, i) => renderProjectCard(p, start + i)).join('');
  }

  renderPagination(total);
}

function renderPagination(total) {
  const pages = Math.ceil(total / PAGE_SIZE);
  if (pages <= 1) { document.getElementById('pagination').innerHTML = ''; return; }

  let html = '';
  if (currentPage > 1) html += `<button class="page-btn wide" onclick="goPage(${currentPage-1})">← Trước</button>`;

  const range = [];
  for (let p = 1; p <= pages; p++) {
    if (p === 1 || p === pages || (p >= currentPage - 2 && p <= currentPage + 2)) range.push(p);
    else if (range[range.length-1] !== '…') range.push('…');
  }

  range.forEach(p => {
    if (p === '…') html += `<button class="page-btn" disabled>…</button>`;
    else html += `<button class="page-btn ${p === currentPage ? 'active' : ''}" onclick="goPage(${p})">${p}</button>`;
  });

  if (currentPage < pages) html += `<button class="page-btn wide" onclick="goPage(${currentPage+1})">Sau →</button>`;

  document.getElementById('pagination').innerHTML = html;
}

function goPage(p) {
  currentPage = p;
  renderProjects();
  window.scrollTo({ top: 0, behavior: 'smooth' });
}

function toggleProject(idx) {
  const card = document.getElementById('proj-' + idx);
  card.classList.toggle('open');
}

// ===================== FILTERS =====================
document.addEventListener('DOMContentLoaded', () => {
  // Date
  const now = new Date();
  document.getElementById('reportDate').textContent =
    '📅 Cập nhật: ' + now.toLocaleDateString('vi-VN', { weekday:'long', year:'numeric', month:'long', day:'numeric' });

  renderStats();
  renderDonut();
  renderPicBars();
  renderProjects();

  // Filter buttons
  document.querySelectorAll('.filter-btn').forEach(btn => {
    btn.addEventListener('click', () => {
      document.querySelectorAll('.filter-btn').forEach(b => b.classList.remove('active'));
      btn.classList.add('active');
      currentFilter = btn.dataset.filter;
      currentPage = 1;
      renderProjects();
    });
  });

  // Search
  let searchTimer;
  document.getElementById('searchInput').addEventListener('input', e => {
    clearTimeout(searchTimer);
    searchTimer = setTimeout(() => {
      currentSearch = e.target.value.trim();
      currentPage = 1;
      renderProjects();
    }, 250);
  });
});
</script>

</body>
</html>
'''

with open('2026_MEDIA_TASKS_REPORT.html', 'w', encoding='utf-8') as f:
    f.write(html)

print("Done! File saved: 2026_MEDIA_TASKS_REPORT.html")
