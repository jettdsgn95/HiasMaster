import sys, io, json, re
from datetime import datetime, date
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook

# ── Excel extraction ──────────────────────────────────────
wb = load_workbook('2026_[MEDIA - TASKS TRACKER].xlsx')
ws = wb['All Task']

def ext(val):
    if val is None: return None
    val = str(val)
    if 'COMPUTED_VALUE' in val or 'IFERROR' in val:
        m = re.search(r'COMPUTED_VALUE[^,]*,\"(.*?)\"\)', val)
        if m: return m.group(1)
        m = re.search(r',\"([^\"]+)\"\)$', val)
        if m: return m.group(1)
        return None
    return val

HDR = {'tt':0,'level1':1,'level2':2,'pic':4,'start':7,'end':8,'extended_date':9,'completion':14,'state':16,'cancel':15}
D_FROM, D_TO = date(2026,5,1), date(2026,5,31)

def pd(s):
    try: return datetime.strptime(s,'%Y-%m-%d').date() if s else None
    except: return None

def in_may(t):
    s, e = pd(t.get('start')), pd(t.get('extended') or t.get('end'))
    if not s and not e: return False
    if not e: return D_FROM <= s <= D_TO
    if not s: return D_FROM <= e <= D_TO
    return s <= D_TO and e >= D_FROM

raw_projs, cur = [], None
for ri, row in enumerate(ws.iter_rows(values_only=True)):
    if ri < 4: continue
    cells = []
    for c in row:
        if isinstance(c, datetime): cells.append(c.strftime('%Y-%m-%d'))
        elif c is None: cells.append(None)
        else:
            v = str(c)
            cells.append(ext(v) if ('COMPUTED_VALUE' in v or 'IFERROR' in v) else v)
    if all(x is None for x in cells): continue
    def g(k): i=HDR.get(k,-1); return cells[i] if 0<=i<len(cells) else None
    l1, l2 = g('level1'), g('level2')
    if l1 and not l2:
        cur = {'id':g('tt'),'name':l1,'tasks':[]}; raw_projs.append(cur)
    elif l2 and cur:
        try: cv = float(g('completion')) if g('completion') else 0
        except: cv = 0
        cur['tasks'].append({'tt':g('tt'),'name':l2,'pic':g('pic'),'start':g('start'),
            'end':g('end'),'extended':g('extended_date'),'completion':cv,'state':g('state'),'cancel':g('cancel')})

projects = [{'id':p['id'],'name':p['name'],'tasks':[t for t in p['tasks'] if in_may(t)]}
            for p in raw_projs if any(in_may(t) for t in p['tasks'])]

# ── Shooting schedule ─────────────────────────────────────
shooting = [
    {"stt":1, "ten":"Hội nghị Đối tác 2026","yc":"Event (Ms.Quyên)","dd":"Thiên Hộ Plaza","thu":"Chủ Nhật","ngay":"03/05","gio":"TBU","pt":"Du, Vinh","tt":"Đã bàn giao"},
    {"stt":2, "ten":"Họp Phụ huynh KV Sóc Trăng","yc":"Event (Ms.Quyên)","dd":"Sóc Trăng","thu":"Thứ 7","ngay":"09/05","gio":"TBU","pt":"","tt":"Đã bàn giao"},
    {"stt":3, "ten":"Họp Phụ huynh KV Vĩnh Long","yc":"Event (Ms.Quyên)","dd":"TTC Palace","thu":"Chủ Nhật","ngay":"10/05","gio":"TBU","pt":"Vinh, Du","tt":"Đã bàn giao"},
    {"stt":4, "ten":"Khai trương Gò Công Đông","yc":"Event (Ms.Quyên)","dd":"CB Gò Công Đông","thu":"TBU","ngay":"TBU","gio":"TBU","pt":"","tt":""},
    {"stt":5, "ten":"Hội thảo 'Chương trình làm quen tiếng Anh'","yc":"MKT (Mr.Phong)","dd":"TBU","thu":"Thứ 4","ngay":"06/05","gio":"Sáng - TBU","pt":"Vinh","tt":"Đã bàn giao"},
    {"stt":6, "ten":"SaiGon x Emasi Tour","yc":"SBU WR (Ms.Thu)","dd":"Emasi – SG","thu":"Thứ 5","ngay":"14/05","gio":"TBU","pt":"Vinh","tt":""},
    {"stt":7, "ten":"Store – Set up đèn quay phim","yc":"SBU Store (Ms.Trang)","dd":"Studio","thu":"Thứ 6","ngay":"15/05","gio":"Chiều 4:00","pt":"Du","tt":"Đã bàn giao"},
    {"stt":8, "ten":"CB Mỹ Tho – Ngoại khóa tại trường","yc":"CB Mỹ Tho (Mr.Hoàng)","dd":"Trường TH Thiên Hộ Dương CS1","thu":"Thứ 2","ngay":"18/05","gio":"Sáng 7:00","pt":"Du","tt":"Đã bàn giao"},
    {"stt":9, "ten":"CB Mỹ Tho – RCV tại trường","yc":"CB Mỹ Tho (Mr.Hoàng)","dd":"Trường THPT Trần Hưng Đạo","thu":"Thứ 5","ngay":"21/05","gio":"Sáng 7:00","pt":"Vinh","tt":"Đã bàn giao"},
    {"stt":10,"ten":"Chụp ảnh lớp học CB Mekong","yc":"Media (Mr.Hòa)","dd":"CB Mekong","thu":"Thứ 7","ngay":"23/05","gio":"Sáng 9:15","pt":"Vinh","tt":"Đã bàn giao"},
    {"stt":11,"ten":"Chụp ảnh lớp học CB Mỹ Tho","yc":"Media (Mr.Hòa)","dd":"CB Mỹ Tho","thu":"Thứ 7","ngay":"23/05","gio":"Chiều 3:00","pt":"Du","tt":"Đã bàn giao"},
    {"stt":12,"ten":"Quay source giới thiệu Futura","yc":"SBU FTR (Mr.Duy)","dd":"CB Mekong","thu":"Thứ 6","ngay":"29/05","gio":"Chiều 2:00","pt":"Du","tt":""},
    {"stt":13,"ten":"Hội thi Hùng biện tiếng Anh x Đồng Tháp","yc":"SBU FTR (Mr.Duy)","dd":"CB Mekong","thu":"Thứ 7","ngay":"30/05","gio":"TBU","pt":"Vinh, Du","tt":""},
    {"stt":14,"ten":"Store – Set up quay/chụp SP","yc":"SBU Store (Ms.Trang)","dd":"Studio","thu":"Thứ 3","ngay":"26/05","gio":"Chiều 3:30","pt":"","tt":""},
]

# ── Video Checklist ───────────────────────────────────────
videos = [
    {"stt":36,"da":"CB Adventures","vid":"Giới thiệu Chương trình hè","kb":"Kịch bản video – CT hè","ord":"05/05","quay":"","dl":"16/05","pic":"Du","pct":100,"tt":"Hoàn thành","done":"16/05"},
    {"stt":37,"da":"CB Wonders","vid":"Giới thiệu Chương trình học","kb":"Kịch bản video – CT hè","ord":"05/05","quay":"","dl":"16/05","pic":"Du","pct":100,"tt":"Hoàn thành","done":"16/05"},
    {"stt":38,"da":"CB Mỹ Tho","vid":"Animation Ads PreS","kb":"Video ads mỹ tho","ord":"05/05","quay":"","dl":"11/05","pic":"Vinh","pct":100,"tt":"Hoàn thành","done":"11/05"},
    {"stt":39,"da":"CB Mỹ Tho","vid":"Animation Ads PreS – Revised 1","kb":"Video ads mỹ tho","ord":"11/05","quay":"","dl":"12/05","pic":"Vinh","pct":100,"tt":"Hoàn thành","done":"12/05"},
    {"stt":40,"da":"CB Mỹ Tho","vid":"Animation Ads YLE","kb":"Video ads mỹ tho","ord":"05/05","quay":"","dl":"09/05","pic":"Vinh","pct":100,"tt":"Hoàn thành","done":"09/05"},
    {"stt":41,"da":"CB WorldReady","vid":"Tour Sing","kb":"CB Worldready – SG Tour","ord":"05/05","quay":"","dl":"08/05","pic":"Du","pct":100,"tt":"Hoàn thành","done":"08/05"},
    {"stt":42,"da":"CB WorldReady","vid":"Tour Emasi","kb":"CB Worldready – Emasi Tour","ord":"05/05","quay":"","dl":"09/05","pic":"","pct":100,"tt":"Hoàn thành","done":"08/05"},
    {"stt":43,"da":"CB Adventures","vid":"Animation (hè)","kb":"KB Video animation hè","ord":"05/05","quay":"","dl":"","pic":"Vinh","pct":100,"tt":"Hoàn thành","done":""},
    {"stt":44,"da":"Hội thảo","vid":"Bộ quà tổng Hộp bút – Revised","kb":"Cắt video phỏng vấn","ord":"29/05","quay":"","dl":"29/05","pic":"Vinh","pct":100,"tt":"Hoàn thành","done":"29/05"},
    {"stt":45,"da":"Hội thảo","vid":"Cut raw","kb":"Cắt video phỏng vấn","ord":"05/05","quay":"","dl":"13/05","pic":"Du","pct":100,"tt":"Hoàn thành","done":"12/05"},
    {"stt":46,"da":"Hội thảo","vid":"Short PV","kb":"Cắt video phỏng vấn","ord":"05/05","quay":"","dl":"15/05","pic":"Du","pct":100,"tt":"Hoàn thành","done":"13/05"},
    {"stt":47,"da":"CX – Khi Cici | CB talk","vid":"Ep1 – Char","kb":"","ord":"","quay":"","dl":"","pic":"Du","pct":0,"tt":"","done":""},
    {"stt":48,"da":"CX – Khi Cici | CB talk","vid":"Ep1 – Storyboard","kb":"","ord":"","quay":"","dl":"","pic":"Du","pct":0,"tt":"","done":""},
    {"stt":49,"da":"CX – Khi Cici | CB talk","vid":"Ep1 – Demo","kb":"","ord":"","quay":"","dl":"","pic":"Du","pct":0,"tt":"","done":""},
    {"stt":50,"da":"CB NEWS","vid":"#143","kb":"CB News 143","ord":"12/05","quay":"","dl":"16/05","pic":"Vinh","pct":100,"tt":"Hoàn thành","done":"16/05"},
    {"stt":51,"da":"CB NEWS","vid":"Voice thông báo CB Summer","kb":"","ord":"","quay":"","dl":"","pic":"Du","pct":100,"tt":"Hoàn thành","done":"16/05"},
    {"stt":52,"da":"Họp Phụ Huynh 2026","vid":"Recap sự kiện họp PH TTC","kb":"","ord":"","quay":"","dl":"","pic":"Du","pct":100,"tt":"Hoàn thành","done":"15/05"},
    {"stt":53,"da":"CB WorldReady","vid":"Voice thông báo Tour Sing","kb":"CB Worldready – Phát thanh","ord":"19/05","quay":"19/05","dl":"","pic":"Du","pct":100,"tt":"Hoàn thành","done":"19/05"},
    {"stt":54,"da":"CB WorldReady","vid":"Voice thông báo Tour Emasi","kb":"CB Worldready – Phát thanh","ord":"19/05","quay":"19/05","dl":"","pic":"","pct":0,"tt":"","done":""},
    {"stt":55,"da":"CB Chợ Gao","vid":"Ads PreS","kb":"Đổi địa chỉ 2 CS","ord":"19/05","quay":"","dl":"20/05","pic":"Vinh","pct":100,"tt":"Hoàn thành","done":"20/05"},
    {"stt":56,"da":"CB Chợ Gao","vid":"Ads YLE","kb":"Đổi địa chỉ 2 CS","ord":"19/05","quay":"","dl":"20/05","pic":"","pct":100,"tt":"Hoàn thành","done":"20/05"},
    {"stt":57,"da":"Về Xanh yêu Thương","vid":"Voice thông báo","kb":"Phát thanh 2026","ord":"19/05","quay":"","dl":"19/05","pic":"Du","pct":100,"tt":"Hoàn thành","done":"19/05"},
    {"stt":58,"da":"CB Thừa Đức","vid":"Ads YLE","kb":"CB Khai trường TH Thừa Đức","ord":"19/05","quay":"","dl":"31/05","pic":"Vinh","pct":100,"tt":"Hoàn thành","done":"21/05"},
    {"stt":59,"da":"Series Phụ huynh nói gì?","vid":"Phụ huynh: Diễm Thúy","kb":"Series PH CB nói gì?","ord":"21/05","quay":"","dl":"26/05","pic":"Du","pct":100,"tt":"Hoàn thành","done":"01/06"},
    {"stt":60,"da":"Series Phụ huynh nói gì?","vid":"Phụ huynh: Khánh Linh","kb":"Series PH CB nói gì?","ord":"21/05","quay":"","dl":"26/05","pic":"Du","pct":100,"tt":"","done":""},
    {"stt":61,"da":"Series HV nói gì?","vid":"Hải Phong","kb":"","ord":"21/05","quay":"","dl":"26/05","pic":"Du","pct":100,"tt":"","done":""},
    {"stt":62,"da":"Series HV nói gì?","vid":"Khánh Vy","kb":"Series HV CB nói gì?","ord":"21/05","quay":"","dl":"28/05","pic":"Du","pct":100,"tt":"Hoàn thành","done":""},
    {"stt":63,"da":"Series HV nói gì?","vid":"Nhú Uyên","kb":"Series HV CB nói gì?","ord":"21/05","quay":"","dl":"28/05","pic":"Du","pct":100,"tt":"Hoàn thành","done":""},
    {"stt":64,"da":"Series HV nói gì?","vid":"Bảo Lam","kb":"Series HV CB nói gì?","ord":"21/05","quay":"","dl":"28/05","pic":"Du","pct":100,"tt":"Hoàn thành","done":""},
    {"stt":65,"da":"Series HV nói gì?","vid":"Ngọc Hân","kb":"Series HV CB nói gì?","ord":"21/05","quay":"","dl":"28/05","pic":"Du","pct":100,"tt":"Hoàn thành","done":""},
    {"stt":66,"da":"Series HV nói gì?","vid":"Hữu Vĩnh","kb":"Series HV CB nói gì?","ord":"21/05","quay":"","dl":"28/05","pic":"Du","pct":100,"tt":"Hoàn thành","done":""},
    {"stt":67,"da":"Series HV nói gì?","vid":"Ngọc Thiện","kb":"Series HV CB nói gì?","ord":"21/05","quay":"","dl":"28/05","pic":"Du","pct":100,"tt":"Hoàn thành","done":""},
    {"stt":68,"da":"CB Gò Công","vid":"Ads PreS","kb":"Đổi địa chỉ Gò Công","ord":"26/05","quay":"","dl":"27/05","pic":"Vinh","pct":100,"tt":"Hoàn thành","done":"27/05"},
    {"stt":69,"da":"CB Gò Công","vid":"Ads YLE","kb":"Đổi địa chỉ Gò Công","ord":"26/05","quay":"","dl":"27/05","pic":"Du","pct":100,"tt":"Hoàn thành","done":"27/05"},
    {"stt":70,"da":"Hệ thống","vid":"Ads PreS","kb":"Hotline + web hệ thống","ord":"26/05","quay":"","dl":"26/05","pic":"Vinh","pct":100,"tt":"Hoàn thành","done":"27/05"},
    {"stt":71,"da":"Hệ thống","vid":"Ads YLE","kb":"Hotline + web hệ thống","ord":"26/05","quay":"","dl":"26/05","pic":"Vinh","pct":100,"tt":"","done":""},
    {"stt":72,"da":"Hội thi HBTA tỉnh Đồng Tháp","vid":"Recap","kb":"Recap","ord":"26/05","quay":"","dl":"05/06","pic":"Du","pct":10,"tt":"Đang làm","done":""},
    {"stt":73,"da":"CB Futura Ads","vid":"Futura Q2","kb":"Ads Futura Quý 2","ord":"26/05","quay":"29/05","dl":"","pic":"Du","pct":100,"tt":"Hoàn thành","done":"01/06"},
]

# ── Social Post data (from screenshots – Theo dõi Social Post T5/2026)
# sl_anh: nếu None → sl_anh = sl_bai
social = [
    {"stt":1,  "nd":"CB SUMMER OF WONDERS 2026 – HÈ VUI KHÁM PHÁ, MỞ LỐI TƯƠNG LAI",              "loai":"Chiến Dịch","sl_bai":1,"sl_anh":None,"ngay":"04/05/2026","cnt":"Tú",    "media":"Quyên","link":"CB Wonders",          "tt":"Hoàn thành"},
    {"stt":2,  "nd":"6 TUẦN HÈ CB SUMMER OF WONDERS – CON KHÁM PHÁ ĐIỀU GÌ?",                     "loai":"Chiến Dịch","sl_bai":1,"sl_anh":None,"ngay":"06/05/2026","cnt":"Vinh",  "media":"Quyên","link":"6 Tuần hè, Con K...","tt":"Hoàn thành"},
    {"stt":3,  "nd":"CON HỌC GÌ TẠI SUMMER OF WONDER THEO ĐỘ TUỔI",                               "loai":"Chiến Dịch","sl_bai":1,"sl_anh":None,"ngay":"08/05/2026","cnt":"Tú",    "media":"Quyên","link":"summer of wonder",     "tt":"Hoàn thành"},
    {"stt":4,  "nd":"CB SUMMER ADVENTURES – BREAK THE LIMITS",                                     "loai":"Chiến Dịch","sl_bai":1,"sl_anh":None,"ngay":"09/05/2026","cnt":"Tú",    "media":"Quyên","link":"CB SUMMER ADVE...",    "tt":"Hoàn thành"},
    {"stt":5,  "nd":"Cover CB Summer Adventure – Break the Limits (2 KT: Fanpage + Website)",      "loai":"Chiến Dịch","sl_bai":1,"sl_anh":None,"ngay":"05/05/2026","cnt":"Tú",    "media":"Quyên","link":"Cover adventure",       "tt":"Hoàn thành"},
    {"stt":6,  "nd":"BÀ MẸ CÓ CON CHUẨN BỊ VÀO LỚP 1 ĐỪNG BỎ QUA!",                             "loai":"Ads",       "sl_bai":1,"sl_anh":None,"ngay":"05/05/2026","cnt":"Bảo",   "media":"",      "link":"Lịch KG CB Tiên Tr...","tt":"Hoàn thành"},
    {"stt":7,  "nd":"CON HỌC TIẾNG ANH NHƯNG VẪN KHÔNG NÓI ĐƯỢC?",                                "loai":"Ads",       "sl_bai":1,"sl_anh":None,"ngay":"05/05/2026","cnt":"Bảo",   "media":"",      "link":"Lịch KG CB Tiên Tr...","tt":"Hoàn thành"},
    {"stt":8,  "nd":"Bản tin CB T4",                                                               "loai":"",          "sl_bai":1,"sl_anh":None,"ngay":"11/05/2026","cnt":"Bảo",   "media":"",      "link":"bản tin",              "tt":""},
    {"stt":9,  "nd":"CB Summer of Wonder 2026",                                                    "loai":"Web",       "sl_bai":2,"sl_anh":None,"ngay":"07/05/2026","cnt":"Bảo",   "media":"Quyên","link":"web",                  "tt":"Hoàn thành"},
    {"stt":10, "nd":"CB Summer Adventures 2026 – Học Hè Bứt Phá Cho Con",                         "loai":"Web",       "sl_bai":1,"sl_anh":None,"ngay":"08/05/2026","cnt":"Vinh",  "media":"",      "link":"CB Summer Adven...",   "tt":"Hoàn thành"},
    {"stt":11, "nd":"CON HỌC ĐÚNG LỘ TRÌNH, KHÔNG LẠC MẠN",                                      "loai":"Ads",       "sl_bai":1,"sl_anh":None,"ngay":"15/05/2026","cnt":"Bảo",   "media":"Quyên","link":"Học đúng lộ trình",   "tt":"Hoàn thành"},
    {"stt":12, "nd":"CHÀO ĐÓN 3 CHI NHÁNH MỚI",                                                   "loai":"Post Đơn",  "sl_bai":1,"sl_anh":None,"ngay":"12/05/2026","cnt":"Thanh", "media":"Quyên","link":"CHÀO ĐÓN 3 CHI...",   "tt":"Hoàn thành"},
    {"stt":13, "nd":"LÀM SAO ĐỂ GIÚP CON VƯỢT QUA NỖI SỢ \"NÓI TIẾNG ANH\"?",                    "loai":"Post Đơn",  "sl_bai":1,"sl_anh":None,"ngay":"12/05/2026","cnt":"Thanh", "media":"Quyên","link":"LÀM SAO ĐỂ GIÚP...", "tt":"Hoàn thành"},
    {"stt":14, "nd":"Chào đón 3 chi nhánh mới",                                                   "loai":"Post Đơn",  "sl_bai":1,"sl_anh":None,"ngay":"12/05/2026","cnt":"Thanh", "media":"Quyên","link":"CHÀO ĐÓN 3 CHI...",   "tt":"Hoàn thành"},
    {"stt":15, "nd":"KHỞI ĐỘNG NHẸ NHÀNG – VỮNG VÀNG HỌC TẬP",                                   "loai":"Chiến Dịch","sl_bai":1,"sl_anh":1,   "ngay":"18/05/2026","cnt":"Tú",    "media":"Quyên","link":"KHỞI ĐỘNG NHẸ...",    "tt":"Hoàn thành"},
    {"stt":16, "nd":"CB SUMMER ADVENTURE (tùy chỉnh mẫu các chi nhánh)",                          "loai":"Ads",       "sl_bai":4,"sl_anh":None,"ngay":"14/05/2026","cnt":"Bảo",   "media":"Quyên","link":"ads",                  "tt":"Hoàn thành"},
    {"stt":17, "nd":"KHAI TRƯƠNG VUI HẾT NẤC – ĐÈN CB TÂN ĐỒNG NHẬN QUÀ NGAY",                  "loai":"Chiến Dịch","sl_bai":1,"sl_anh":1,   "ngay":"18/05/2026","cnt":"Bảo",   "media":"Quyên","link":"CB Centres khai tr...","tt":"Hoàn thành"},
    {"stt":18, "nd":"KHAI TRƯƠNG CB TÂN ĐỒNG – NHẬN ƯU ĐÃI CỰC HOT",                             "loai":"Chiến Dịch","sl_bai":1,"sl_anh":None,"ngay":"16/05/2026","cnt":"Bảo",   "media":"",      "link":"KHAI TRƯƠNG CB...",    "tt":"Hoàn thành"},
    {"stt":19, "nd":"MUỐN CON TIẾU HỌC GIỎI TIẾNG ANH – ĐỪNG BỎ QUA CB TÂN ĐỒNG",               "loai":"Chiến Dịch","sl_bai":1,"sl_anh":None,"ngay":"18/05/2026","cnt":"Bảo",   "media":"",      "link":"4.đb",                 "tt":"Hoàn thành"},
    {"stt":20, "nd":"Cover CB Summer Adventure – Break the Limits – MCN (bộ tour trải nghiệm)",   "loai":"Chiến Dịch","sl_bai":1,"sl_anh":None,"ngay":"11/05/2026","cnt":"Tú",    "media":"Quyên","link":"CB CENTRES_Cov...",   "tt":"Hoàn thành"},
    {"stt":21, "nd":"Chào đón 3 chi nhánh mới ver.2",                                             "loai":"Post Đơn",  "sl_bai":1,"sl_anh":None,"ngay":"13/05/2026","cnt":"Thanh", "media":"",      "link":"Post khai trương",     "tt":"Hoàn thành"},
    {"stt":22, "nd":"HINT về xanh yêu thương",                                                    "loai":"Chiến Dịch","sl_bai":1,"sl_anh":None,"ngay":"14/05/2026","cnt":"Thanh", "media":"Hòa",   "link":"Post_vẽ-xanh.jpg",     "tt":"Hoàn thành"},
    {"stt":23, "nd":"KHÓA TIẾNG ANH THỪA ĐỨC – NHẬN ƯU ĐÃI ĐẾN 10%",                            "loai":"Ads",       "sl_bai":1,"sl_anh":None,"ngay":"14/05/2026","cnt":"Bảo",   "media":"Vinh",  "link":"KHAI TRƯƠNG CB...",    "tt":"Hoàn thành"},
    {"stt":24, "nd":"CHÍNH THỨC KHAI TRƯƠNG CB THỪA ĐỨC",                                        "loai":"Ads",       "sl_bai":1,"sl_anh":None,"ngay":"18/05/2026","cnt":"Bảo",   "media":"Vinh",  "link":"CHÍNH THỨC KHA...",    "tt":"Hoàn thành"},
    {"stt":25, "nd":"LỘ TRÌNH HỌC TIẾNG ANH HÈ CÁ NHÂN HÓA CHO CON",                            "loai":"Ads",       "sl_bai":1,"sl_anh":None,"ngay":"18/05/2026","cnt":"Bảo",   "media":"Vinh",  "link":"LỊCH KG CB TIẾN T...", "tt":"Hoàn thành"},
    {"stt":26, "nd":"CON HỌC TIẾNG ANH HÈ TẠI CB – BA MẸ AN TÂM TRỌN VẸN",                      "loai":"Ads",       "sl_bai":1,"sl_anh":None,"ngay":"19/05/2026","cnt":"Bảo",   "media":"Vinh",  "link":"CON HỌC TIẾNG...",    "tt":"Hoàn thành"},
    {"stt":27, "nd":"CB Summer of Wonders – Khóa Học Tiếng Anh Hè Giúp Con Tự Tin Hội Nhập",     "loai":"Web",       "sl_bai":3,"sl_anh":None,"ngay":"21/05/2026","cnt":"Bảo",   "media":"Quyên","link":"web adventure",        "tt":"Hoàn thành"},
    {"stt":28, "nd":"Phiếu đăng ký học trải nghiệm CB Tân Đồng",                                 "loai":"Form cover","sl_bai":1,"sl_anh":None,"ngay":"14/05/2026","cnt":"Trang", "media":"Hòa",   "link":"Cover GG Form",        "tt":"Hoàn thành"},
    {"stt":29, "nd":"Trải nghiệm hè bung nở – ngóng trăm quà tổng",                              "loai":"Chiến Dịch","sl_bai":1,"sl_anh":4,   "ngay":"20/05/2026","cnt":"Tú",    "media":"Vinh",  "link":"SẴN SÀNG TRẢI...",     "tt":"Hoàn thành"},
    {"stt":30, "nd":"Cover Summer of Wonders (resize từ banner Wonders) – KT website",            "loai":"Web",       "sl_bai":1,"sl_anh":4,   "ngay":"16/05/2026","cnt":"Tú",    "media":"Vinh",  "link":"Cover CB Wo...",        "tt":"Hoàn thành"},
    {"stt":31, "nd":"Happy Children's Day 1/6",                                                   "loai":"Chiến Dịch","sl_bai":1,"sl_anh":4,   "ngay":"",          "cnt":"Thanh", "media":"Quyên","link":"Childrenday",           "tt":"Hoàn thành"},
    {"stt":32, "nd":"LÊN THCS, THPT – TIẾNG ANH KHÔNG THỂ HỌC \"ĐỐI PHÓ\"",                     "loai":"Ads",       "sl_bai":1,"sl_anh":1,   "ngay":"",          "cnt":"Bảo",   "media":"Vinh",  "link":"GIÚP VIÊN T...",        "tt":"Hoàn thành"},
    {"stt":33, "nd":"Về xanh – Hướng dẫn tham gia",                                              "loai":"Chiến Dịch","sl_bai":1,"sl_anh":8,   "ngay":"22/05/2026","cnt":"Thanh", "media":"Hòa",   "link":"Về Xanh_HDTG",          "tt":"Hoàn thành"},
    {"stt":34, "nd":"Giày khen trên tay – quá xịn ao ngay",                                      "loai":"Chiến Dịch","sl_bai":1,"sl_anh":4,   "ngay":"23/05/2026","cnt":"Tú",    "media":"Vinh",  "link":"KHOE GIÀY KHEN...",    "tt":"Hoàn thành"},
    {"stt":35, "nd":"Ads YLE – Kiên Giang (tùy chỉnh mẫu)",                                      "loai":"Ads",       "sl_bai":1,"sl_anh":1,   "ngay":"22/05/2026","cnt":"Bảo",   "media":"Quyên","link":"CB Centres kiên Gi...","tt":"Hoàn thành"},
    {"stt":36, "nd":"Ba mẹ Tân Đồng",                                                            "loai":"Ads",       "sl_bai":1,"sl_anh":None,"ngay":"22/05/2026","cnt":"Bảo",   "media":"Quyên","link":"ĐỪNG BỎ LỠ \"Đ...\"",  "tt":"Hoàn thành"},
    {"stt":37, "nd":"Ads Ba Trị – Giồng Trôm – Gò Công (tùy chỉnh mẫu có sẵn)",                 "loai":"Ads",       "sl_bai":1,"sl_anh":2,   "ngay":"22/05/2026","cnt":"Bảo",   "media":"Vinh",  "link":"ĐỪNG BỎ LỠ \"Đ...\"",  "tt":"Hoàn thành"},
    {"stt":38, "nd":"Ads Kiên Giang – Cái Bè – Long Trung – Cai Lậy (tùy chỉnh mẫu có sẵn)",   "loai":"Ads",       "sl_bai":1,"sl_anh":4,   "ngay":"22/05/2026","cnt":"Bảo",   "media":"Vinh",  "link":"ĐỪNG BỎ LỠ \"Đ...\"",  "tt":"Hoàn thành"},
    {"stt":39, "nd":"BÍ QUYẾT GIÚP TRẺ YÊU TIẾNG ANH",                                          "loai":"Post Đơn",  "sl_bai":1,"sl_anh":None,"ngay":"25/05/2026","cnt":"Thanh", "media":"Hòa",   "link":"Phương pháp TPR",       "tt":"Hoàn thành"},
    {"stt":40, "nd":"Tổng kết năm học",                                                           "loai":"Post Đơn",  "sl_bai":1,"sl_anh":5,   "ngay":"",          "cnt":"Thanh", "media":"Quyên","link":"Tổng kết năm học",      "tt":"Hoàn thành"},
    {"stt":41, "nd":"HÀNG TRĂM BA MẸ CHỌN KHÓA HỌC TIẾNG ANH TẠI CB MỎ CÀY NAM",               "loai":"Ads",       "sl_bai":1,"sl_anh":None,"ngay":"25/05/2026","cnt":"Bảo",   "media":"",      "link":"Lịch Khai giảng C...",  "tt":"Gấp"},
    {"stt":42, "nd":"LỊCH KHAI GIẢNG LỚP TIẾNG ANH – CB MỎ CÀY NAM",                            "loai":"Ads",       "sl_bai":1,"sl_anh":None,"ngay":"25/05/2026","cnt":"Bảo",   "media":"",      "link":"Lịch Khai giảng C...",  "tt":"Gấp"},
    {"stt":43, "nd":"Vì sao ba mẹ cho con học tiếng Anh tại CB Bình Đại? (theo mẫu sẵn)",        "loai":"Ads",       "sl_bai":1,"sl_anh":1,   "ngay":"",          "cnt":"Bảo",   "media":"",      "link":"Lịch Khai giảng C...",  "tt":"Hoàn thành"},
    {"stt":44, "nd":"CB Huỳnh Tấn Phát giúp HV tiếp thu kiến thức nhanh chóng – theo mẫu",      "loai":"Ads",       "sl_bai":1,"sl_anh":4,   "ngay":"25/05/2026","cnt":"Bảo",   "media":"Vinh",  "link":"Lịch Khai giảng M...",  "tt":"Hoàn thành"},
    {"stt":45, "nd":"ĐẦU TƯ TIẾNG ANH NGAY TỪ SỚM – MÓN QUÀ GIÁ TRỊ BA MẸ TRAO CON",          "loai":"Ads",       "sl_bai":1,"sl_anh":None,"ngay":"25/05/2026","cnt":"Bảo",   "media":"Vinh",  "link":"TIẾNG ANH...",          "tt":"Hoàn thành"},
    {"stt":46, "nd":"LỊCH KHAI GIẢNG LỚP TIẾNG ANH – CB MỎ CÀY NAM (ver 2)",                    "loai":"Ads",       "sl_bai":1,"sl_anh":None,"ngay":"27/05/2026","cnt":"Bảo",   "media":"",      "link":"lịch khai giảng MC...", "tt":"Hoàn thành"},
    {"stt":47, "nd":"BÀ MẸ ĐI CÀY NHỚ BỔN PHẬN \"ĐỐI VÀNG\" học tiếng Anh cho con",             "loai":"Ads",       "sl_bai":1,"sl_anh":None,"ngay":"27/05/2026","cnt":"Bảo",   "media":"",      "link":"MCN-01.jpg",            "tt":"Hoàn thành"},
    {"stt":48, "nd":"Thông báo cuộc thi Speak to Shine",                                         "loai":"Chiến Dịch","sl_bai":1,"sl_anh":None,"ngay":"25/05/2026","cnt":"Trang", "media":"Quyên","link":"Speak to shine",         "tt":"Hoàn thành"},
    {"stt":49, "nd":"Thông báo cuộc thi Speak to Shine (Cover)",                                 "loai":"Cover",     "sl_bai":1,"sl_anh":None,"ngay":"25/05/2026","cnt":"Trang", "media":"Quyên","link":"COVER PAGE SPE...",      "tt":"Hoàn thành"},
    {"stt":50, "nd":"Thông báo cuộc thi Speak to Shine (Web)",                                   "loai":"Web",       "sl_bai":1,"sl_anh":None,"ngay":"26/05/2026","cnt":"Trang", "media":"Quyên","link":"Web SPEAKTOSHI...",      "tt":"Hoàn thành"},
    {"stt":51, "nd":"CÙNG CÚC KHÁM PHÁ THÊM 6 TỪ VỰNG",                                         "loai":"Post Đơn",  "sl_bai":1,"sl_anh":7,   "ngay":"",          "cnt":"Thanh", "media":"Vinh",  "link":"6 TỪ VỰNG...",          "tt":"Hoàn thành"},
    {"stt":52, "nd":"Gợi ý chủ đề về xanh yêu thương",                                          "loai":"Chiến Dịch","sl_bai":1,"sl_anh":6,   "ngay":"28/05/2026","cnt":"Thanh", "media":"Hòa",   "link":"CB CENTRES-KHỔ...",    "tt":"Hoàn thành"},
    {"stt":53, "nd":"Phát động cuộc thi Speak to Shine",                                         "loai":"Chiến Dịch","sl_bai":1,"sl_anh":None,"ngay":"28/05/2026","cnt":"Trang", "media":"Quyên","link":"CB Trang",               "tt":"Hoàn thành"},
    {"stt":54, "nd":"3 lỗi phổ biến khiến bài thi Speak to Shine không hợp lệ",                 "loai":"Chiến Dịch","sl_bai":1,"sl_anh":4,   "ngay":"29/05/2026","cnt":"Trang", "media":"Quyên","link":"3 lỗi phổ biến",         "tt":"Hoàn thành"},
    {"stt":55, "nd":"Những điểm khác biệt của Speak to Shine kỳ 6",                              "loai":"Chiến Dịch","sl_bai":1,"sl_anh":None,"ngay":"30/05/2026","cnt":"Trang", "media":"Quyên","link":"Những điểm khác biệt", "tt":"Hoàn thành"},
    {"stt":56, "nd":"Khám phá từng hạng mục giải thưởng Speak to Shine",                         "loai":"Chiến Dịch","sl_bai":1,"sl_anh":7,   "ngay":"01/06/2026","cnt":"Trang", "media":"Quyên","link":"Điểm khác biệt",         "tt":"Hoàn thành"},
    {"stt":57, "nd":"Intro + Outro Speak to Shine",                                               "loai":"Chiến Dịch","sl_bai":1,"sl_anh":None,"ngay":"01/06/2026","cnt":"Trang", "media":"Quyên","link":"Intro + Outro Spe...",   "tt":"Hoàn thành"},
    {"stt":58, "nd":"5 tuần hè tại CB Centres có gì khác biệt?",                                 "loai":"Chiến Dịch","sl_bai":1,"sl_anh":5,   "ngay":"28/05/2026","cnt":"Tú",    "media":"Vinh",  "link":"5 TUẦN HÈ TẠI C...",    "tt":"Hoàn thành"},
    {"stt":59, "nd":"LỘ TRÌNH HỌC TIẾNG ANH CÁ NHÂN HÓA CHO TRẺ 7-11 TUỔI",                    "loai":"Ads",       "sl_bai":1,"sl_anh":None,"ngay":"29/05/2026","cnt":"Bảo",   "media":"Quyên","link":"CB CENTRES_LỘ...",      "tt":"Hoàn thành"},
    {"stt":60, "nd":"EXPLORE VIETNAM THIS SUMMER WITH CICI",                                      "loai":"Post Đơn",  "sl_bai":1,"sl_anh":None,"ngay":"29/05/2026","cnt":"Thanh", "media":"Quyên","link":"Cùng Cici vi vu Vie...","tt":"Hoàn thành"},
    {"stt":61, "nd":"MUỐN CON TIẾU HỌC GIỎI TIẾNG ANH – ĐỪNG BỎ QUA CB GÒ CÔNG (MCN)",         "loai":"Ads",       "sl_bai":1,"sl_anh":None,"ngay":"28/05/2026","cnt":"Bảo",   "media":"Vinh",  "link":"Gò Công.jpg",            "tt":"Hoàn thành"},
    {"stt":62, "nd":"Hàng trăm ba mẹ chọn lớp tiếng Anh tại CB Gò Công – tùy chỉnh mẫu MCN",   "loai":"Ads",       "sl_bai":1,"sl_anh":None,"ngay":"28/05/2026","cnt":"Bảo",   "media":"Vinh",  "link":"",                       "tt":"Hoàn thành"},
    {"stt":63, "nd":"1 DAY LEFT – READY FOR CHILDREN'S DAY?",                                    "loai":"Chiến Dịch","sl_bai":1,"sl_anh":None,"ngay":"",          "cnt":"Thanh", "media":"",      "link":"1 DAY LEFT – REA...",    "tt":"Hoàn thành"},
]

# Apply rule: sl_anh = sl_bai if sl_anh is None
for s in social:
    if s['sl_anh'] is None:
        s['sl_anh'] = s['sl_bai']
    s['sl_anh_display'] = s['sl_anh']

# ── Stats ─────────────────────────────────────────────────
ttasks     = sum(len(p['tasks']) for p in projects)
tdone      = sum(1 for p in projects for t in p['tasks'] if 'Hoàn thành' in (t.get('state') or ''))
sdone_s    = sum(1 for s in shooting if s['tt'] == 'Đã bàn giao')
vdone      = sum(1 for v in videos if v['tt'] == 'Hoàn thành')
vdoing     = sum(1 for v in videos if v['tt'] == 'Đang làm')
sp_done    = sum(1 for s in social if s['tt'] == 'Hoàn thành')
sp_urgent  = sum(1 for s in social if s['tt'] == 'Gấp')
total_bai  = sum(s['sl_bai'] for s in social)
total_anh  = sum(s['sl_anh'] for s in social)

pj  = json.dumps(projects, ensure_ascii=False)
sj  = json.dumps(shooting, ensure_ascii=False)
vj  = json.dumps(videos,   ensure_ascii=False)
spj = json.dumps(social,   ensure_ascii=False)
now = datetime.now().strftime('%H:%M – %d/%m/%Y')

HTML = f'''<!DOCTYPE html>
<html lang="vi">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Báo Cáo Tháng 5/2026 – Team Media CB</title>
<meta name="description" content="Báo cáo tổng hợp tháng 5/2026: Task thiết kế, Lịch quay chụp, Video Checklist và Social Post – Team Media CB.">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap" rel="stylesheet">
<style>
/* ═══ BRAND TOKENS #BA110F · #191970 ═══ */
:root {{
  --red:#BA110F; --red-dk:#8C0D0B; --red-md:#D4140F; --red-lt:#E63836;
  --red-pale:#FDE8E8; --red-tint:#FFF4F4;
  --navy:#191970; --navy-dk:#0F0F50; --navy-md:#222290; --navy-lt:#3333BB;
  --navy-pale:#EAEAF8; --navy-tint:#F2F2FC;
  --bg:#F5F6FA; --surf:#FFFFFF; --surf2:#F8F9FD;
  --border:#E2E5F0; --divider:rgba(25,25,112,.07);
  --t1:#0D0D30; --t2:#424275; --t3:#8888BB;
  --green:#0A7A52; --amber:#B07600; --orange:#C05000;
  --violet:#6B21A8; --teal:#0E7490;
  --sh-xs:0 1px 3px rgba(25,25,112,.06);
  --sh-sm:0 2px 8px rgba(25,25,112,.08);
  --sh-md:0 4px 20px rgba(25,25,112,.12);
  --sh-red:0 4px 16px rgba(186,17,15,.18);
  --sh-nav:0 4px 16px rgba(25,25,112,.22);
  --r0:6px; --r1:10px; --r2:16px; --r3:24px;
}}
*,*::before,*::after{{margin:0;padding:0;box-sizing:border-box;}}
html{{scroll-behavior:smooth;}}
body{{font-family:'Inter',system-ui,sans-serif;background:var(--bg);color:var(--t1);font-size:14px;line-height:1.55;overflow-x:hidden;}}
::-webkit-scrollbar{{width:5px;height:5px;}}
::-webkit-scrollbar-track{{background:var(--bg);}}
::-webkit-scrollbar-thumb{{background:rgba(25,25,112,.25);border-radius:3px;}}
::-webkit-scrollbar-thumb:hover{{background:var(--navy);}}
.page{{max-width:1440px;margin:0 auto;}}

/* ═══ HERO ═══ */
.hero{{background:linear-gradient(135deg,var(--navy-dk) 0%,var(--navy) 55%,var(--navy-md) 100%);padding:44px 40px 38px;position:relative;overflow:hidden;}}
.hero::before{{content:'';position:absolute;top:-80px;right:-80px;width:340px;height:340px;border-radius:50%;background:radial-gradient(circle,rgba(186,17,15,.3) 0%,transparent 70%);pointer-events:none;}}
.hero::after{{content:'';position:absolute;bottom:-60px;left:20%;width:220px;height:220px;border-radius:50%;background:radial-gradient(circle,rgba(255,255,255,.05) 0%,transparent 70%);pointer-events:none;}}
.hero-inner{{position:relative;z-index:1;display:flex;align-items:flex-end;justify-content:space-between;flex-wrap:wrap;gap:24px;}}
.hero-left{{flex:1;min-width:280px;}}
.hero-kicker{{display:inline-flex;align-items:center;gap:8px;background:rgba(255,255,255,.1);border:1px solid rgba(255,255,255,.18);border-radius:50px;padding:5px 16px;margin-bottom:14px;font-size:11px;font-weight:700;letter-spacing:1.8px;text-transform:uppercase;color:rgba(255,255,255,.9);}}
.hero-dot{{width:6px;height:6px;border-radius:50%;background:#FF6B69;box-shadow:0 0 8px #FF6B69;animation:blink 2s ease-in-out infinite;}}
@keyframes blink{{0%,100%{{opacity:1;transform:scale(1)}}50%{{opacity:.25;transform:scale(.6)}}}}
.hero-title{{font-size:clamp(26px,3.5vw,48px);font-weight:900;line-height:1.08;margin-bottom:10px;color:#fff;letter-spacing:-.5px;}}
.hero-title .hl{{color:#FF8A88;}}
.hero-sub{{display:flex;align-items:center;gap:10px;flex-wrap:wrap;font-size:13px;color:rgba(255,255,255,.6);}}
.hero-sub .sep{{color:rgba(255,255,255,.25);}}
.hero-right{{display:flex;flex-direction:column;align-items:flex-end;gap:10px;flex-shrink:0;}}
.period-card{{background:rgba(255,255,255,.1);border:1px solid rgba(255,255,255,.18);border-radius:var(--r1);padding:14px 20px;display:flex;align-items:center;gap:12px;}}
.pc-icon{{font-size:24px;}}
.pc-lbl{{font-size:10px;font-weight:700;letter-spacing:1.2px;text-transform:uppercase;color:rgba(255,255,255,.45);}}
.pc-val{{font-size:14px;font-weight:800;color:#fff;margin-top:2px;}}
.hero-time{{font-size:11.5px;color:rgba(255,255,255,.4);}}
.accent-stripe{{height:4px;background:linear-gradient(90deg,var(--red) 0%,var(--red-lt) 30%,var(--navy-lt) 70%,var(--navy) 100%);}}

/* ═══ KPI ROW (4 cards) ═══ */
.kpi-row{{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;padding:24px 40px 20px;border-bottom:1px solid var(--divider);}}
@media(max-width:900px){{.kpi-row{{grid-template-columns:repeat(2,1fr);padding:16px;}}}}
@media(max-width:540px){{.kpi-row{{grid-template-columns:1fr;}}}}
.kpi-card{{background:var(--surf);border:1px solid var(--border);border-radius:var(--r2);padding:18px;display:flex;align-items:center;gap:14px;box-shadow:var(--sh-sm);position:relative;overflow:hidden;transition:transform .2s,box-shadow .2s;}}
.kpi-card:hover{{transform:translateY(-4px);box-shadow:var(--sh-md);}}
.kpi-card::before{{content:'';position:absolute;top:0;left:0;bottom:0;width:4px;border-radius:var(--r2) 0 0 var(--r2);}}
.k-task::before{{background:linear-gradient(180deg,var(--navy-dk),var(--navy-lt));}}
.k-shoot::before{{background:linear-gradient(180deg,var(--red-dk),var(--red-lt));}}
.k-video::before{{background:linear-gradient(180deg,#6B21A8,#A855F7);}}
.k-social::before{{background:linear-gradient(180deg,#0E7490,#22D3EE);}}
.kpi-icon{{font-size:34px;flex-shrink:0;}}
.kpi-info{{flex:1;}}
.kpi-num{{font-size:34px;font-weight:900;line-height:1;margin-bottom:4px;}}
.k-task .kpi-num{{color:var(--navy);}} .k-shoot .kpi-num{{color:var(--red);}}
.k-video .kpi-num{{color:#7B21A8;}} .k-social .kpi-num{{color:#0E7490;}}
.kpi-lbl{{font-size:12.5px;font-weight:600;color:var(--t2);}}
.kpi-chips{{display:flex;gap:5px;margin-top:8px;flex-wrap:wrap;}}
.chip{{font-size:10.5px;font-weight:700;padding:2px 9px;border-radius:50px;white-space:nowrap;}}
.chip-g{{background:rgba(10,122,82,.1);color:var(--green);border:1px solid rgba(10,122,82,.2);}}
.chip-n{{background:var(--navy-pale);color:var(--navy);border:1px solid rgba(25,25,112,.18);}}
.chip-a{{background:rgba(176,118,0,.1);color:var(--amber);border:1px solid rgba(176,118,0,.18);}}
.chip-r{{background:var(--red-pale);color:var(--red-dk);border:1px solid rgba(186,17,15,.18);}}
.chip-t{{background:rgba(14,116,144,.1);color:#0E7490;border:1px solid rgba(14,116,144,.2);}}

/* ═══ TABS ═══ */
.tabs-wrap{{padding:20px 40px 0;position:sticky;top:0;z-index:200;background:rgba(245,246,250,.93);backdrop-filter:blur(14px);border-bottom:1px solid var(--border);}}
.tab-nav{{display:flex;gap:0;background:var(--surf);border:1px solid var(--border);border-radius:var(--r2);padding:5px;box-shadow:var(--sh-sm);width:fit-content;flex-wrap:wrap;}}
.tab-btn{{display:flex;align-items:center;gap:8px;padding:9px 18px;border-radius:var(--r1);background:transparent;border:none;color:var(--t2);font-size:13px;font-weight:600;cursor:pointer;font-family:inherit;transition:all .2s;white-space:nowrap;}}
.tab-btn:hover{{color:var(--navy);background:var(--navy-tint);}}
.tab-btn.active{{background:linear-gradient(135deg,var(--navy),var(--navy-lt));color:#fff;box-shadow:var(--sh-nav);}}
.tab-btn.active.t-red{{background:linear-gradient(135deg,var(--red-dk),var(--red-md));box-shadow:var(--sh-red);}}
.tab-btn.active.t-purple{{background:linear-gradient(135deg,#6B21A8,#9333EA);box-shadow:0 4px 16px rgba(107,33,168,.25);}}
.tab-btn.active.t-teal{{background:linear-gradient(135deg,#0C4A6E,#0E7490);box-shadow:0 4px 16px rgba(14,116,144,.3);}}
.tbadge{{font-size:10px;font-weight:800;padding:2px 7px;border-radius:50px;background:var(--navy-pale);color:var(--navy);min-width:20px;text-align:center;}}
.tab-btn.active .tbadge{{background:rgba(255,255,255,.25);color:#fff;}}
.tab-panel{{display:none;}}
.tab-panel.active{{display:block;animation:fadeUp .28s ease;}}
@keyframes fadeUp{{from{{opacity:0;transform:translateY(6px)}}to{{opacity:1;transform:translateY(0)}}}}

/* ═══ CONTENT + SCARD GRID ═══ */
.content{{padding:24px 40px 40px;}}
.scard-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(145px,1fr));gap:12px;margin-bottom:20px;}}
.scard{{background:var(--surf);border:1px solid var(--border);border-radius:var(--r1);padding:15px;box-shadow:var(--sh-xs);position:relative;overflow:hidden;transition:transform .18s,box-shadow .18s;animation:slideUp .38s ease both;}}
.scard:hover{{transform:translateY(-3px);box-shadow:var(--sh-sm);}}
@keyframes slideUp{{from{{opacity:0;transform:translateY(12px)}}to{{opacity:1;transform:translateY(0)}}}}
.scard:nth-child(1){{animation-delay:.04s}}.scard:nth-child(2){{animation-delay:.08s}}
.scard:nth-child(3){{animation-delay:.12s}}.scard:nth-child(4){{animation-delay:.16s}}
.scard:nth-child(5){{animation-delay:.20s}}.scard:nth-child(6){{animation-delay:.24s}}
.scard:nth-child(7){{animation-delay:.28s}}
.scard::before{{content:'';position:absolute;top:0;left:0;right:0;height:3px;border-radius:var(--r1) var(--r1) 0 0;}}
.scard.cn::before{{background:linear-gradient(90deg,var(--navy-dk),var(--navy-lt));}}
.scard.cr::before{{background:linear-gradient(90deg,var(--red-dk),var(--red-lt));}}
.scard.cg::before{{background:linear-gradient(90deg,#0A5C3E,#0A7A52);}}
.scard.cb::before{{background:linear-gradient(90deg,#1D4ED8,#60A5FA);}}
.scard.ca::before{{background:linear-gradient(90deg,#92400E,#D97706);}}
.scard.cp::before{{background:linear-gradient(90deg,#6B21A8,#A855F7);}}
.scard.ct::before{{background:linear-gradient(90deg,#0C4A6E,#0E7490);}}
.scard-icon{{font-size:20px;margin-bottom:8px;}}
.scard-val{{font-size:28px;font-weight:900;line-height:1;margin-bottom:3px;}}
.scard-lbl{{font-size:11px;color:var(--t2);font-weight:500;}}
.scard.cn .scard-val{{color:var(--navy);}} .scard.cr .scard-val{{color:var(--red);}}
.scard.cg .scard-val{{color:var(--green);}} .scard.cb .scard-val{{color:#1D4ED8;}}
.scard.ca .scard-val{{color:var(--amber);}} .scard.cp .scard-val{{color:#7B21A8;}}
.scard.ct .scard-val{{color:#0E7490;}}

/* ═══ CONTROLS ═══ */
.controls{{display:flex;gap:10px;flex-wrap:wrap;align-items:center;margin-bottom:14px;}}
.s-wrap{{flex:1;min-width:200px;position:relative;}}
.s-wrap input{{width:100%;background:var(--surf);border:1px solid var(--border);border-radius:50px;padding:9px 18px 9px 40px;font-size:13px;font-family:inherit;color:var(--t1);outline:none;box-shadow:var(--sh-xs);transition:border-color .2s,box-shadow .2s;}}
.s-wrap input::placeholder{{color:var(--t3);}}
.s-wrap input:focus{{border-color:var(--navy);box-shadow:0 0 0 3px rgba(25,25,112,.1);}}
.s-ico{{position:absolute;left:13px;top:50%;transform:translateY(-50%);color:var(--t3);font-size:13px;}}
.f-group{{display:flex;gap:5px;flex-wrap:wrap;}}
.fbtn{{padding:7px 14px;border-radius:50px;border:1px solid var(--border);background:var(--surf);color:var(--t2);font-size:12px;font-weight:600;cursor:pointer;font-family:inherit;box-shadow:var(--sh-xs);transition:all .18s;white-space:nowrap;}}
.fbtn:hover{{border-color:var(--navy);color:var(--navy);background:var(--navy-tint);}}
.fbtn.active{{background:linear-gradient(135deg,var(--navy),var(--navy-lt));border-color:transparent;color:#fff;box-shadow:var(--sh-nav);}}

/* ═══ RESULTS BAR ═══ */
.rbar{{display:flex;align-items:center;justify-content:space-between;margin-bottom:12px;flex-wrap:wrap;gap:8px;}}
.rinfo{{font-size:12px;color:var(--t2);}} .rinfo strong{{color:var(--t1);}}
.expand-btn{{padding:5px 12px;border-radius:var(--r0);border:1px solid var(--border);background:var(--surf);color:var(--t2);font-size:11.5px;font-weight:600;cursor:pointer;font-family:inherit;box-shadow:var(--sh-xs);transition:all .18s;}}
.expand-btn:hover{{border-color:var(--navy);color:var(--navy);background:var(--navy-tint);}}

/* ═══ PROJECT ACCORDION ═══ */
.proj-list{{display:flex;flex-direction:column;gap:9px;}}
.proj-card{{background:var(--surf);border:1px solid var(--border);border-radius:var(--r1);overflow:hidden;box-shadow:var(--sh-xs);transition:border-color .18s,box-shadow .18s;}}
.proj-card:hover{{border-color:rgba(25,25,112,.28);box-shadow:var(--sh-sm);}}
.proj-head{{display:flex;align-items:center;justify-content:space-between;padding:13px 18px;cursor:pointer;gap:12px;user-select:none;transition:background .14s;}}
.proj-head:hover{{background:var(--navy-tint);}}
.ph-left{{display:flex;align-items:center;gap:12px;flex:1;min-width:0;}}
.ph-num{{font-size:9.5px;font-weight:800;letter-spacing:.5px;color:var(--navy);background:var(--navy-pale);border:1px solid rgba(25,25,112,.18);padding:3px 9px;border-radius:50px;flex-shrink:0;}}
.ph-name{{font-size:13.5px;font-weight:700;color:var(--t1);overflow:hidden;text-overflow:ellipsis;white-space:nowrap;}}
.ph-right{{display:flex;align-items:center;gap:10px;flex-shrink:0;}}
.tc-badge{{font-size:11px;color:var(--t2);background:var(--surf2);border:1px solid var(--border);padding:3px 10px;border-radius:50px;white-space:nowrap;}}
.ph-prog{{display:flex;align-items:center;gap:7px;min-width:105px;}}
.ph-bar{{flex:1;height:5px;background:var(--border);border-radius:50px;overflow:hidden;}}
.ph-fill{{height:100%;border-radius:50px;background:linear-gradient(90deg,var(--navy),var(--navy-lt));}}
.ph-pct{{font-size:11px;font-weight:700;color:var(--navy);min-width:32px;text-align:right;}}
.chevron{{color:var(--t3);font-size:14px;transition:transform .25s;flex-shrink:0;}}
.proj-card.open .chevron{{transform:rotate(180deg);}}
.task-wrap{{max-height:0;overflow:hidden;transition:max-height .35s ease;border-top:1px solid transparent;}}
.proj-card.open .task-wrap{{max-height:4000px;border-top-color:var(--divider);}}
.tt{{width:100%;border-collapse:collapse;font-size:12.5px;}}
.tt th{{padding:8px 14px;text-align:left;font-size:10px;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:var(--t3);background:var(--surf2);white-space:nowrap;border-bottom:1px solid var(--border);}}
.tt td{{padding:9px 14px;border-top:1px solid var(--divider);vertical-align:middle;}}
.tt tr:hover td{{background:var(--navy-tint);}}
.t-name{{font-weight:600;color:var(--t1);}} .t-pic{{color:var(--navy);font-weight:700;font-size:12px;}}
.t-date{{color:var(--t2);white-space:nowrap;font-size:11.5px;}} .t-date.ext{{color:var(--amber);font-weight:600;}}
.sb{{display:inline-flex;align-items:center;gap:4px;padding:2px 9px;border-radius:50px;font-size:10.5px;font-weight:700;white-space:nowrap;}}
.sb.done{{background:rgba(10,122,82,.1);color:#0A5C3E;border:1px solid rgba(10,122,82,.18);}}
.sb.doing{{background:var(--navy-pale);color:var(--navy);border:1px solid rgba(25,25,112,.18);}}
.sb.late{{background:var(--red-pale);color:var(--red-dk);border:1px solid rgba(186,17,15,.18);}}
.sb.paused{{background:rgba(176,118,0,.1);color:var(--amber);border:1px solid rgba(176,118,0,.18);}}
.sb.canceled{{background:var(--surf2);color:var(--t3);border:1px solid var(--border);}}
.c-wrap{{display:inline-flex;align-items:center;gap:5px;}}
.c-bar{{width:50px;height:4px;background:var(--border);border-radius:50px;overflow:hidden;}}
.c-fill{{height:100%;border-radius:50px;background:linear-gradient(90deg,var(--navy),var(--navy-lt));}}
.c-txt{{font-size:11px;color:var(--t2);}}
.pagination{{display:flex;justify-content:center;align-items:center;gap:5px;margin-top:20px;flex-wrap:wrap;}}
.pg{{min-width:34px;height:34px;padding:0 10px;border-radius:var(--r0);border:1px solid var(--border);background:var(--surf);color:var(--t2);font-size:13px;cursor:pointer;display:flex;align-items:center;justify-content:center;font-family:inherit;box-shadow:var(--sh-xs);transition:all .18s;}}
.pg:hover{{border-color:var(--navy);color:var(--navy);background:var(--navy-tint);}}
.pg.active{{background:linear-gradient(135deg,var(--navy),var(--navy-lt));border-color:transparent;color:#fff;box-shadow:var(--sh-nav);}}
.pg:disabled{{opacity:.35;cursor:default;}}

/* ═══ SHOOTING SCHEDULE ═══ */
.shoot-note{{display:flex;align-items:flex-start;gap:10px;background:rgba(186,17,15,.05);border:1px solid rgba(186,17,15,.18);border-left:3px solid var(--red);border-radius:var(--r0);padding:12px 16px;font-size:12.5px;color:var(--t2);margin-bottom:18px;}}
.shoot-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(295px,1fr));gap:13px;}}
.scard-item{{background:var(--surf);border:1px solid var(--border);border-radius:var(--r1);overflow:hidden;box-shadow:var(--sh-xs);transition:transform .2s,box-shadow .2s,border-color .2s;animation:slideUp .38s ease both;}}
.scard-item:hover{{transform:translateY(-4px);box-shadow:var(--sh-md);border-color:rgba(186,17,15,.2);}}
.sc-head{{display:flex;align-items:center;gap:12px;padding:13px 16px;background:linear-gradient(135deg,var(--navy-tint),var(--red-tint));border-bottom:1px solid var(--border);}}
.sc-num{{width:30px;height:30px;border-radius:var(--r0);background:linear-gradient(135deg,var(--navy),var(--navy-lt));display:flex;align-items:center;justify-content:center;font-size:12px;font-weight:800;color:#fff;flex-shrink:0;box-shadow:var(--sh-nav);}}
.sc-title{{font-size:12.5px;font-weight:700;color:var(--t1);flex:1;line-height:1.3;}}
.sc-status{{flex-shrink:0;}}
.sc-body{{padding:12px 16px;display:flex;flex-direction:column;gap:9px;}}
.sc-row{{display:flex;align-items:flex-start;gap:9px;font-size:12.5px;}}
.sc-icon{{font-size:13px;flex-shrink:0;margin-top:1px;}}
.sc-lbl{{font-size:10px;font-weight:700;letter-spacing:.7px;text-transform:uppercase;color:var(--t3);min-width:55px;}}
.sc-val{{color:var(--t2);flex:1;}}
.dt-chip{{display:inline-flex;align-items:center;gap:6px;background:var(--navy-pale);border:1px solid rgba(25,25,112,.18);padding:3px 11px;border-radius:50px;font-size:11.5px;font-weight:700;color:var(--navy);}}
.pic-chips{{display:flex;flex-wrap:wrap;gap:6px;}}
.pic-chip{{font-size:11px;font-weight:700;padding:3px 10px;border-radius:50px;background:var(--red-pale);border:1px solid rgba(186,17,15,.18);color:var(--red-dk);}}
.st-del{{background:rgba(10,122,82,.1);color:#0A5C3E;border:1px solid rgba(10,122,82,.18);}}
.st-pnd{{background:rgba(176,118,0,.1);color:var(--amber);border:1px solid rgba(176,118,0,.18);}}
.st-tbu{{background:var(--surf2);color:var(--t3);border:1px solid var(--border);}}
.st-del,.st-pnd,.st-tbu{{display:inline-flex;align-items:center;gap:4px;padding:3px 10px;border-radius:50px;font-size:10.5px;font-weight:700;white-space:nowrap;}}

/* ═══ VIDEO + SOCIAL TABLE (shared) ═══ */
.vc-wrap{{border:1px solid var(--border);border-radius:var(--r1);overflow:hidden;overflow-x:auto;box-shadow:var(--sh-sm);}}
.vc-table{{width:100%;border-collapse:collapse;font-size:12.5px;min-width:880px;}}
.vc-table thead tr{{background:linear-gradient(135deg,var(--navy-dk),var(--navy));}}
.vc-table th{{padding:11px 14px;text-align:left;font-size:10px;font-weight:800;letter-spacing:1px;text-transform:uppercase;color:rgba(255,255,255,.7);white-space:nowrap;}}
.vc-table td{{padding:9px 14px;border-top:1px solid var(--divider);vertical-align:middle;background:var(--surf);}}
.vc-table tr:hover td{{background:var(--navy-tint);}}
.vc-stt{{color:var(--t3);font-size:11px;text-align:center;}}
.vc-da{{font-weight:700;color:var(--navy);max-width:220px;}}
.vc-vid{{font-weight:500;color:var(--t1);}}
.vc-pic{{color:var(--red);font-weight:700;}}
.vc-date{{color:var(--t2);font-size:11.5px;white-space:nowrap;}}
.vc-dl{{color:var(--red-dk);font-size:11.5px;white-space:nowrap;font-weight:600;}}
.vc-done{{color:var(--green);font-size:11.5px;white-space:nowrap;}}
.prog-line{{display:flex;align-items:center;gap:6px;min-width:85px;}}
.prog-line .pb{{flex:1;height:5px;background:var(--border);border-radius:50px;overflow:hidden;}}
.prog-line .pf{{height:100%;border-radius:50px;}}
.prog-line .pt{{font-size:11px;color:var(--t2);min-width:28px;text-align:right;font-weight:600;}}
.vs-done{{background:rgba(10,122,82,.1);color:#0A5C3E;border:1px solid rgba(10,122,82,.18);}}
.vs-doing{{background:var(--navy-pale);color:var(--navy);border:1px solid rgba(25,25,112,.18);}}
.vs-pend{{background:var(--surf2);color:var(--t3);border:1px solid var(--border);}}
.vs-urgent{{background:var(--red-pale);color:var(--red-dk);border:1px solid rgba(186,17,15,.18);}}
.vs-done,.vs-doing,.vs-pend,.vs-urgent{{display:inline-flex;align-items:center;gap:4px;padding:2px 9px;border-radius:50px;font-size:10.5px;font-weight:700;white-space:nowrap;}}

/* ═══ SOCIAL POST – Loại badges ═══ */
.loai-cd{{background:rgba(186,17,15,.1);color:var(--red-dk);border:1px solid rgba(186,17,15,.2);}}
.loai-ads{{background:var(--navy-pale);color:var(--navy);border:1px solid rgba(25,25,112,.2);}}
.loai-web{{background:rgba(14,116,144,.1);color:#0C4A6E;border:1px solid rgba(14,116,144,.2);}}
.loai-post{{background:rgba(10,122,82,.1);color:#0A5C3E;border:1px solid rgba(10,122,82,.2);}}
.loai-form{{background:rgba(107,33,168,.1);color:var(--violet);border:1px solid rgba(107,33,168,.2);}}
.loai-cover{{background:rgba(176,118,0,.1);color:var(--amber);border:1px solid rgba(176,118,0,.2);}}
.loai-other{{background:var(--surf2);color:var(--t3);border:1px solid var(--border);}}
.loai-cd,.loai-ads,.loai-web,.loai-post,.loai-form,.loai-cover,.loai-other{{
  display:inline-flex;align-items:center;padding:2px 9px;border-radius:50px;
  font-size:10.5px;font-weight:700;white-space:nowrap;
}}

/* SL ảnh counter */
.anh-badge{{
  display:inline-flex;align-items:center;gap:3px;
  padding:2px 8px;border-radius:50px;
  font-size:11px;font-weight:700;
  background:rgba(14,116,144,.1);color:#0C4A6E;
  border:1px solid rgba(14,116,144,.2);
  white-space:nowrap;
}}
.anh-badge.inferred{{
  background:rgba(176,118,0,.08);color:var(--amber);
  border-color:rgba(176,118,0,.18);
  font-style:italic;
}}

/* ═══ FOOTER ═══ */
footer{{background:linear-gradient(135deg,var(--navy-dk),var(--navy));padding:24px 40px;text-align:center;color:rgba(255,255,255,.5);font-size:12.5px;margin-top:44px;}}
footer strong{{color:rgba(255,255,255,.8);}}
.no-res{{text-align:center;padding:52px 20px;color:var(--t3);font-size:15px;}}
.no-res .nr-icon{{font-size:48px;display:block;margin-bottom:12px;}}

@media(max-width:900px){{
  .ph-prog{{display:none;}}
  .tt th:nth-child(n+4),.tt td:nth-child(n+4){{display:none;}}
  .tab-btn .tb-txt{{display:none;}}
  .hero{{padding:30px 20px;}}
  .kpi-row,.content,.tabs-wrap{{padding-left:20px;padding-right:20px;}}
}}
@media(max-width:640px){{
  .hero-right{{display:none;}}
  .shoot-grid{{grid-template-columns:1fr;}}
}}
</style>
</head>
<body>
<div class="page">

<!-- ═ HERO ═ -->
<div class="hero">
  <div class="hero-inner">
    <div class="hero-left">
      <div class="hero-kicker"><span class="hero-dot"></span>Tháng 5 · 2026 · Team Media</div>
      <h1 class="hero-title">BÁO CÁO CÔNG VIỆC<br><span class="hl">THÁNG 5 / 2026</span></h1>
      <div class="hero-sub">
        <span>📂 Team Media</span><span class="sep">·</span>
        <span>🏢 Phòng Marketing</span><span class="sep">·</span>
        <span>📍 HIAS Master · CB</span>
      </div>
    </div>
    <div class="hero-right">
      <div class="period-card">
        <div class="pc-icon">📅</div>
        <div><div class="pc-lbl">Kỳ báo cáo</div><div class="pc-val">01/05/2026 → 31/05/2026</div></div>
      </div>
      <div class="hero-time">🕐 Xuất: {now}</div>
    </div>
  </div>
</div>
<div class="accent-stripe"></div>

<!-- ═ KPI ROW ═ -->
<div class="kpi-row">
  <div class="kpi-card k-task">
    <div class="kpi-icon">🎨</div>
    <div class="kpi-info">
      <div class="kpi-num">{ttasks}</div>
      <div class="kpi-lbl">Task thiết kế</div>
      <div class="kpi-chips">
        <span class="chip chip-g">✅ {tdone} HT</span>
        <span class="chip chip-n">🔵 {ttasks-tdone} Đang</span>
      </div>
    </div>
  </div>
  <div class="kpi-card k-shoot">
    <div class="kpi-icon">📷</div>
    <div class="kpi-info">
      <div class="kpi-num">{len(shooting)}</div>
      <div class="kpi-lbl">Lịch quay chụp</div>
      <div class="kpi-chips">
        <span class="chip chip-g">✅ {sdone_s} Giao</span>
        <span class="chip chip-a">⏳ {len(shooting)-sdone_s} Chờ</span>
      </div>
    </div>
  </div>
  <div class="kpi-card k-video">
    <div class="kpi-icon">🎬</div>
    <div class="kpi-info">
      <div class="kpi-num">{len(videos)}</div>
      <div class="kpi-lbl">Video Checklist</div>
      <div class="kpi-chips">
        <span class="chip chip-g">✅ {vdone} HT</span>
        <span class="chip chip-n">🔵 {vdoing} Đang</span>
      </div>
    </div>
  </div>
  <div class="kpi-card k-social">
    <div class="kpi-icon">📱</div>
    <div class="kpi-info">
      <div class="kpi-num">{len(social)}</div>
      <div class="kpi-lbl">Social Posts</div>
      <div class="kpi-chips">
        <span class="chip chip-g">✅ {sp_done} HT</span>
        <span class="chip chip-r">⚡ {sp_urgent} Gấp</span>
        <span class="chip chip-t">🖼 {total_anh} ảnh</span>
      </div>
    </div>
  </div>
</div>

<!-- ═ TABS ═ -->
<div class="tabs-wrap">
  <div class="tab-nav">
    <button class="tab-btn active" id="tab-tasks" onclick="sw('tasks')">
      🎨 <span class="tb-txt">Công việc chính</span> <span class="tbadge">{ttasks}</span>
    </button>
    <button class="tab-btn t-red" id="tab-shoot" onclick="sw('shoot')">
      📷 <span class="tb-txt">Lịch quay chụp</span> <span class="tbadge">{len(shooting)}</span>
    </button>
    <button class="tab-btn t-purple" id="tab-video" onclick="sw('video')">
      🎬 <span class="tb-txt">Video Checklist</span> <span class="tbadge">{len(videos)}</span>
    </button>
    <button class="tab-btn t-teal" id="tab-social" onclick="sw('social')">
      📱 <span class="tb-txt">Social Posts</span> <span class="tbadge">{len(social)}</span>
    </button>
  </div>
</div>

<!-- ═ PANEL: TASKS ═ -->
<div class="tab-panel active content" id="panel-tasks">
  <div class="scard-grid" id="tStats"></div>
  <div class="controls">
    <div class="s-wrap"><span class="s-ico">🔍</span><input type="text" id="tSearch" placeholder="Tìm dự án, công việc, nhân sự..."></div>
    <div class="f-group" id="tFilters">
      <button class="fbtn active" data-f="all">Tất cả</button>
      <button class="fbtn" data-f="Hoàn thành">✅ Hoàn thành</button>
      <button class="fbtn" data-f="Đang làm">🔵 Đang làm</button>
      <button class="fbtn" data-f="Hoãn">⏸ Hoãn</button>
    </div>
  </div>
  <div class="rbar">
    <div class="rinfo" id="tInfo"></div>
    <button class="expand-btn" id="expandBtn" onclick="toggleAll()">⤢ Mở rộng tất cả</button>
  </div>
  <div class="proj-list" id="projList"></div>
  <div class="pagination" id="tPager"></div>
</div>

<!-- ═ PANEL: SHOOT ═ -->
<div class="tab-panel content" id="panel-shoot">
  <div class="scard-grid" id="sStats"></div>
  <div class="shoot-note">
    <span style="font-size:16px">⚠️</span>
    <div><strong style="color:var(--red)">Lưu ý:</strong> Source hình ảnh bàn giao trong <strong>1 ngày</strong> kể từ khi kết thúc sự kiện.</div>
  </div>
  <div class="controls">
    <div class="s-wrap"><span class="s-ico">🔍</span><input type="text" id="sSearch" placeholder="Tìm sự kiện, địa điểm, phụ trách..."></div>
    <div class="f-group" id="sFilters">
      <button class="fbtn active" data-f="all">Tất cả</button>
      <button class="fbtn" data-f="del">✅ Đã bàn giao</button>
      <button class="fbtn" data-f="pnd">⏳ Chưa / TBU</button>
    </div>
  </div>
  <div class="rbar"><div class="rinfo" id="sInfo"></div></div>
  <div class="shoot-grid" id="shootGrid"></div>
</div>

<!-- ═ PANEL: VIDEO ═ -->
<div class="tab-panel content" id="panel-video">
  <div class="scard-grid" id="vStats"></div>
  <div class="controls">
    <div class="s-wrap"><span class="s-ico">🔍</span><input type="text" id="vSearch" placeholder="Tìm dự án, video, kịch bản, PIC..."></div>
    <div class="f-group" id="vFilters">
      <button class="fbtn active" data-f="all">Tất cả</button>
      <button class="fbtn" data-f="Hoàn thành">✅ Hoàn thành</button>
      <button class="fbtn" data-f="Đang làm">🔵 Đang làm</button>
      <button class="fbtn" data-f="pend">⏳ Chờ xử lý</button>
    </div>
  </div>
  <div class="rbar"><div class="rinfo" id="vInfo"></div></div>
  <div class="vc-wrap">
    <table class="vc-table">
      <thead><tr>
        <th style="text-align:center">#</th>
        <th>Dự án / CN</th><th>Video</th><th>Kịch bản</th>
        <th>Nhận order</th><th>Ngày quay</th><th>Deadline</th>
        <th>PIC</th><th>Tiến độ</th><th>Tình trạng</th><th>Hoàn thành</th>
      </tr></thead>
      <tbody id="vBody"></tbody>
    </table>
  </div>
</div>

<!-- ═ PANEL: SOCIAL ═ -->
<div class="tab-panel content" id="panel-social">
  <div class="scard-grid" id="spStats"></div>

  <div class="shoot-note" style="background:rgba(14,116,144,.05);border-color:rgba(14,116,144,.2);border-left-color:#0E7490;">
    <span style="font-size:16px">ℹ️</span>
    <div>
      <strong style="color:#0C4A6E">Lưu ý SL ảnh:</strong>
      Nếu cột <em>SL Ảnh</em> để trống thì <strong>số lượng ảnh = số bài post</strong>.
      Số in nghiêng màu cam = giá trị được tự suy (inferred).
    </div>
  </div>

  <div class="controls">
    <div class="s-wrap"><span class="s-ico">🔍</span><input type="text" id="spSearch" placeholder="Tìm nội dung post, loại, content, media..."></div>
    <div class="f-group" id="spFilters">
      <button class="fbtn active" data-f="all">Tất cả</button>
      <button class="fbtn" data-f="Chiến Dịch">🔴 Chiến Dịch</button>
      <button class="fbtn" data-f="Ads">🔵 Ads</button>
      <button class="fbtn" data-f="Web">🌐 Web</button>
      <button class="fbtn" data-f="Post Đơn">📝 Post Đơn</button>
      <button class="fbtn" data-f="other">📦 Khác</button>
    </div>
    <div class="f-group" id="spTTFilters">
      <button class="fbtn active" data-tf="all">Mọi TT</button>
      <button class="fbtn" data-tf="Hoàn thành">✅ HT</button>
      <button class="fbtn" data-tf="Gấp">⚡ Gấp</button>
      <button class="fbtn" data-tf="pending">⏳ Chờ</button>
    </div>
  </div>

  <div class="rbar"><div class="rinfo" id="spInfo"></div></div>

  <div class="vc-wrap">
    <table class="vc-table" style="min-width:1000px">
      <thead><tr>
        <th style="text-align:center">#</th>
        <th>Nội dung (Post)</th>
        <th>Loại</th>
        <th style="text-align:center">SL Bài</th>
        <th style="text-align:center">SL Ảnh</th>
        <th>Ngày đăng</th>
        <th>Content</th>
        <th>Media</th>
        <th>Link sản phẩm</th>
        <th>Tình trạng</th>
      </tr></thead>
      <tbody id="spBody"></tbody>
    </table>
  </div>
</div>

<!-- ═ FOOTER ═ -->
<footer>
  <p>📊 Báo cáo tháng 5/2026 &nbsp;·&nbsp; <strong>01/05 – 31/05/2026</strong> &nbsp;·&nbsp; Team Media · CB Media Production Team · HIAS Master</p>
  <p style="margin-top:6px">Xuất lúc: {now}</p>
</footer>
</div>

<script>
const PROJ  = {pj};
const SHOOT = {sj};
const VIDS  = {vj};
const SOCIAL = {spj};

// ── TABS ──
function sw(id) {{
  ['tasks','shoot','video','social'].forEach(t=>{{
    document.getElementById('tab-'+t).classList.remove('active');
    document.getElementById('panel-'+t).classList.remove('active');
  }});
  document.getElementById('tab-'+id).classList.add('active');
  document.getElementById('panel-'+id).classList.add('active');
}}

// ── HELPERS ──
function sc(state,cancel){{
  if(cancel)return'canceled';if(!state)return'doing';
  if(state.includes('Hoàn thành'))return'done';
  if(state.includes('Hoãn'))return'paused';
  if(state.includes('Quá hạn'))return'late';
  return'doing';
}}
function sl(state,cancel){{
  if(cancel)return'🚫 '+cancel;if(!state)return'🔵 Đang làm';
  if(state.includes('Hoàn thành'))return'✅ Hoàn thành';
  if(state.includes('Hoãn'))return'⏸ Hoãn';
  if(state.includes('Quá hạn'))return'🔴 Quá hạn';
  return'🔵 Đang làm';
}}
function pct(p){{
  if(!p.tasks.length)return 0;
  return Math.round(p.tasks.filter(t=>sc(t.state,t.cancel)==='done').length/p.tasks.length*100);
}}
function fd(d){{
  if(!d)return'–';const pts=d.split('-');
  return pts.length===3?pts[2]+'/'+pts[1]:d;
}}

// ══ TASKS ══
const PAGE=20;let tFil='all',tSrch='',tPage=1,tFiltd=[],allOpen=false;
function renderTaskStats(){{
  let tot=0,done=0,doing=0,paused=0;
  PROJ.forEach(p=>p.tasks.forEach(t=>{{
    tot++;const s=sc(t.state,t.cancel);
    if(s==='done')done++;else if(s==='doing')doing++;else if(s==='paused'||s==='canceled')paused++;
  }}));
  const p=tot?Math.round(done/tot*100):0;
  document.getElementById('tStats').innerHTML=[
    {{v:PROJ.length,l:'Dự án',c:'cn',i:'📁'}},{{v:tot,l:'Công việc',c:'cb',i:'📋'}},
    {{v:done,l:'Hoàn thành',c:'cg',i:'✅'}},{{v:doing,l:'Đang thực hiện',c:'cn',i:'🔵'}},
    {{v:paused,l:'Hoãn / Huỷ',c:'ca',i:'⏸'}},{{v:p+'%',l:'Tỉ lệ HT',c:'cp',i:'📈'}},
  ].map(s=>`<div class="scard ${{s.c}}"><div class="scard-icon">${{s.i}}</div><div class="scard-val">${{s.v}}</div><div class="scard-lbl">${{s.l}}</div></div>`).join('');
}}
function getFiltd(){{
  return PROJ.filter(p=>{{
    const mF=tFil==='all'||p.tasks.some(t=>(t.state||'').includes(tFil));
    const q=tSrch.toLowerCase();
    const mS=!q||p.name.toLowerCase().includes(q)||p.tasks.some(t=>(t.name||'').toLowerCase().includes(q)||(t.pic||'').toLowerCase().includes(q));
    return mF&&mS;
  }});
}}
function renderProjs(){{
  tFiltd=getFiltd();
  const totT=tFiltd.reduce((s,p)=>s+p.tasks.length,0);
  document.getElementById('tInfo').innerHTML=`Hiển thị <strong>${{tFiltd.length}}</strong> dự án · <strong>${{totT}}</strong> công việc`;
  const start=(tPage-1)*PAGE,page=tFiltd.slice(start,start+PAGE);
  document.getElementById('projList').innerHTML=page.length?page.map((p,i)=>renderPC(p,start+i)).join(''):'<div class="no-res"><span class="nr-icon">🔍</span><p>Không tìm thấy kết quả</p></div>';
  renderTPager(tFiltd.length);
}}
function renderPC(p,idx){{
  const pct2=pct(p),done=p.tasks.filter(t=>sc(t.state,t.cancel)==='done').length;
  return`<div class="proj-card" id="pc-${{idx}}">
    <div class="proj-head" onclick="toggleP(${{idx}})">
      <div class="ph-left"><span class="ph-num">#${{p.id||idx+1}}</span><span class="ph-name" title="${{p.name}}">${{p.name}}</span></div>
      <div class="ph-right">
        <span class="tc-badge">${{done}}/${{p.tasks.length}} việc</span>
        <div class="ph-prog"><div class="ph-bar"><div class="ph-fill" style="width:${{pct2}}%"></div></div><span class="ph-pct">${{pct2}}%</span></div>
        <span class="chevron">▾</span>
      </div>
    </div>
    <div class="task-wrap">
      <table class="tt">
        <thead><tr><th>Công việc</th><th>PIC</th><th>Bắt đầu</th><th>Kết thúc</th><th>Trạng thái</th><th>% HT</th></tr></thead>
        <tbody>${{p.tasks.map(t=>`<tr>
          <td class="t-name">${{t.name||''}}</td><td class="t-pic">${{t.pic||'–'}}</td>
          <td class="t-date">${{fd(t.start)}}</td>
          <td class="t-date${{t.extended?' ext':''}}">${{fd(t.extended||t.end)}}${{t.extended?' ⤴':''}}</td>
          <td><span class="sb ${{sc(t.state,t.cancel)}}">${{sl(t.state,t.cancel)}}</span></td>
          <td><div class="c-wrap"><div class="c-bar"><div class="c-fill" style="width:${{Math.round((t.completion||0)*100)}}%"></div></div><span class="c-txt">${{Math.round((t.completion||0)*100)}}%</span></div></td>
        </tr>`).join('')}}</tbody>
      </table>
    </div>
  </div>`;
}}
function toggleP(idx){{document.getElementById('pc-'+idx).classList.toggle('open');}}
function toggleAll(){{
  allOpen=!allOpen;
  document.querySelectorAll('#projList .proj-card').forEach(c=>c.classList.toggle('open',allOpen));
  document.getElementById('expandBtn').textContent=allOpen?'⤡ Thu gọn tất cả':'⤢ Mở rộng tất cả';
}}
function renderTPager(tot){{
  const pages=Math.ceil(tot/PAGE);
  if(pages<=1){{document.getElementById('tPager').innerHTML='';return;}}
  let h='';
  if(tPage>1)h+=`<button class="pg" onclick="goTP(${{tPage-1}})">←</button>`;
  for(let p=1;p<=pages;p++){{
    if(p===1||p===pages||Math.abs(p-tPage)<=1)h+=`<button class="pg${{p===tPage?' active':''}}" onclick="goTP(${{p}})">${{p}}</button>`;
    else if(h.slice(-10)!=='>…</button>')h+=`<button class="pg" disabled>…</button>`;
  }}
  if(tPage<pages)h+=`<button class="pg" onclick="goTP(${{tPage+1}})">→</button>`;
  document.getElementById('tPager').innerHTML=h;
}}
function goTP(p){{tPage=p;renderProjs();window.scrollTo({{top:0,behavior:'smooth'}});}}

// ══ SHOOT ══
let sFil='all',sSrch='';
function renderShootStats(){{
  const tot=SHOOT.length,del=SHOOT.filter(s=>s.tt==='Đã bàn giao').length;
  document.getElementById('sStats').innerHTML=[
    {{v:tot,l:'Lịch quay chụp',c:'cr',i:'📷'}},
    {{v:del,l:'Đã bàn giao',c:'cg',i:'✅'}},
    {{v:tot-del,l:'Chờ / TBU',c:'ca',i:'⏳'}},
  ].map(s=>`<div class="scard ${{s.c}}"><div class="scard-icon">${{s.i}}</div><div class="scard-val">${{s.v}}</div><div class="scard-lbl">${{s.l}}</div></div>`).join('');
}}
function getSFiltd(){{
  return SHOOT.filter(s=>{{
    const mF=sFil==='all'||(sFil==='del'&&s.tt==='Đã bàn giao')||(sFil==='pnd'&&s.tt!=='Đã bàn giao');
    const q=sSrch.toLowerCase();
    return mF&&(!q||s.ten.toLowerCase().includes(q)||s.dd.toLowerCase().includes(q)||s.pt.toLowerCase().includes(q)||s.yc.toLowerCase().includes(q));
  }});
}}
function renderShoot(){{
  const list=getSFiltd();
  document.getElementById('sInfo').innerHTML=`Hiển thị <strong>${{list.length}}</strong> / ${{SHOOT.length}} lịch`;
  if(!list.length){{document.getElementById('shootGrid').innerHTML='<div class="no-res" style="grid-column:1/-1"><span class="nr-icon">📷</span><p>Không tìm thấy</p></div>';return;}}
  document.getElementById('shootGrid').innerHTML=list.map((s,i)=>{{
    const stB=s.tt==='Đã bàn giao'?'<span class="st-del">✅ Đã bàn giao</span>':s.ngay==='TBU'?'<span class="st-tbu">📌 TBU</span>':'<span class="st-pnd">⏳ Chờ</span>';
    const picH=s.pt?s.pt.split(',').map(n=>`<span class="pic-chip">👤 ${{n.trim()}}</span>`).join(''):'<span style="color:var(--t3);font-size:12px">Chưa phân công</span>';
    return`<div class="scard-item" style="animation-delay:${{i*.04}}s">
      <div class="sc-head"><div class="sc-num">${{s.stt}}</div><div class="sc-title">${{s.ten}}</div><div class="sc-status">${{stB}}</div></div>
      <div class="sc-body">
        <div class="sc-row"><span class="sc-icon">📌</span><div><div class="sc-lbl">Yêu cầu</div><div class="sc-val">${{s.yc}}</div></div></div>
        <div class="sc-row"><span class="sc-icon">📍</span><div><div class="sc-lbl">Địa điểm</div><div class="sc-val">${{s.dd}}</div></div></div>
        <div class="sc-row"><span class="sc-icon">🗓</span><div><div class="sc-lbl">Thời gian</div><span class="dt-chip">${{s.ngay}}/2026 · ${{s.gio}} (${{s.thu}})</span></div></div>
        <div class="sc-row"><span class="sc-icon">👤</span><div><div class="sc-lbl">Phụ trách</div><div class="pic-chips">${{picH}}</div></div></div>
      </div>
    </div>`;
  }}).join('');
}}

// ══ VIDEO ══
let vFil='all',vSrch='';
function renderVideoStats(){{
  const tot=VIDS.length,done=VIDS.filter(v=>v.tt==='Hoàn thành').length,doing=VIDS.filter(v=>v.tt==='Đang làm').length;
  document.getElementById('vStats').innerHTML=[
    {{v:tot,l:'Tổng video',c:'cp',i:'🎬'}},{{v:done,l:'Hoàn thành',c:'cg',i:'✅'}},
    {{v:doing,l:'Đang xử lý',c:'cn',i:'🔵'}},{{v:tot-done-doing,l:'Chờ xử lý',c:'ca',i:'⏳'}},
    {{v:tot?Math.round(done/tot*100)+'%':'0%',l:'Tỉ lệ HT',c:'cr',i:'📈'}},
  ].map(s=>`<div class="scard ${{s.c}}"><div class="scard-icon">${{s.i}}</div><div class="scard-val">${{s.v}}</div><div class="scard-lbl">${{s.l}}</div></div>`).join('');
}}
function getVFiltd(){{
  return VIDS.filter(v=>{{
    const mF=vFil==='all'||(vFil==='Hoàn thành'&&v.tt==='Hoàn thành')||(vFil==='Đang làm'&&v.tt==='Đang làm')||(vFil==='pend'&&v.tt!=='Hoàn thành'&&v.tt!=='Đang làm');
    const q=vSrch.toLowerCase();
    return mF&&(!q||v.da.toLowerCase().includes(q)||v.vid.toLowerCase().includes(q)||(v.kb||'').toLowerCase().includes(q)||(v.pic||'').toLowerCase().includes(q));
  }});
}}
function renderVideos(){{
  const list=getVFiltd();
  document.getElementById('vInfo').innerHTML=`Hiển thị <strong>${{list.length}}</strong> / ${{VIDS.length}} video`;
  const pc=p=>p===100?'linear-gradient(90deg,#0A7A52,#10B981)':p>0?'linear-gradient(90deg,var(--navy),var(--navy-lt))':'var(--border)';
  document.getElementById('vBody').innerHTML=list.map(v=>{{
    const cls=v.tt==='Hoàn thành'?'vs-done':v.tt==='Đang làm'?'vs-doing':'vs-pend';
    const lbl=v.tt==='Hoàn thành'?'✅ Hoàn thành':v.tt==='Đang làm'?'🔵 Đang làm':v.tt||'⏳ Chờ xử lý';
    return`<tr>
      <td class="vc-stt">${{v.stt}}</td><td class="vc-da">${{v.da}}</td><td class="vc-vid">${{v.vid}}</td>
      <td style="color:var(--t2);font-size:11.5px">${{v.kb||'–'}}</td>
      <td class="vc-date">${{v.ord||'–'}}</td><td class="vc-date">${{v.quay||'–'}}</td>
      <td class="vc-dl">${{v.dl||'–'}}</td><td class="vc-pic">${{v.pic||'–'}}</td>
      <td><div class="prog-line"><div class="pb"><div class="pf" style="width:${{v.pct}}%;background:${{pc(v.pct)}}"></div></div><span class="pt">${{v.pct}}%</span></div></td>
      <td><span class="${{cls}}">${{lbl}}</span></td>
      <td class="vc-done">${{v.done||'–'}}</td>
    </tr>`;
  }}).join('');
}}

// ══ SOCIAL POST ══
let spFil='all', spTTFil='all', spSrch='';

function renderSocialStats(){{
  const tot=SOCIAL.length;
  const done=SOCIAL.filter(s=>s.tt==='Hoàn thành').length;
  const urgent=SOCIAL.filter(s=>s.tt==='Gấp').length;
  const pend=SOCIAL.filter(s=>s.tt!=='Hoàn thành'&&s.tt!=='Gấp').length;
  const totalBai=SOCIAL.reduce((a,s)=>a+s.sl_bai,0);
  const totalAnh=SOCIAL.reduce((a,s)=>a+s.sl_anh,0);
  // by type
  const cd=SOCIAL.filter(s=>s.loai==='Chiến Dịch').length;
  const ads=SOCIAL.filter(s=>s.loai==='Ads').length;

  document.getElementById('spStats').innerHTML=[
    {{v:tot,   l:'Tổng bài post', c:'ct',i:'📱'}},
    {{v:done,  l:'Hoàn thành',    c:'cg',i:'✅'}},
    {{v:urgent,l:'Gấp',           c:'cr',i:'⚡'}},
    {{v:pend,  l:'Chờ xử lý',     c:'ca',i:'⏳'}},
    {{v:totalBai,l:'Tổng SL bài', c:'cn',i:'📝'}},
    {{v:totalAnh,l:'Tổng SL ảnh', c:'cb',i:'🖼'}},
    {{v:cd,    l:'Chiến Dịch',    c:'cr',i:'🔴'}},
    {{v:ads,   l:'Ads',           c:'cn',i:'🔵'}},
  ].map(s=>`<div class="scard ${{s.c}}"><div class="scard-icon">${{s.i}}</div><div class="scard-val">${{s.v}}</div><div class="scard-lbl">${{s.l}}</div></div>`).join('');
}}

function loaiBadge(loai){{
  if(!loai)return'<span class="loai-other">—</span>';
  if(loai==='Chiến Dịch')return`<span class="loai-cd">🔴 ${{loai}}</span>`;
  if(loai==='Ads')return`<span class="loai-ads">🔵 Ads</span>`;
  if(loai==='Web')return`<span class="loai-web">🌐 Web</span>`;
  if(loai==='Post Đơn')return`<span class="loai-post">📝 Post Đơn</span>`;
  if(loai==='Form cover')return`<span class="loai-form">📋 Form cover</span>`;
  if(loai==='Cover')return`<span class="loai-cover">🖼 Cover</span>`;
  return`<span class="loai-other">${{loai}}</span>`;
}}

function getSPFiltd(){{
  return SOCIAL.filter(s=>{{
    const mF=spFil==='all'
      ||(spFil==='other'&&s.loai!=='Chiến Dịch'&&s.loai!=='Ads'&&s.loai!=='Web'&&s.loai!=='Post Đơn')
      ||(s.loai===spFil);
    const mTT=spTTFil==='all'
      ||(spTTFil==='Hoàn thành'&&s.tt==='Hoàn thành')
      ||(spTTFil==='Gấp'&&s.tt==='Gấp')
      ||(spTTFil==='pending'&&s.tt!=='Hoàn thành'&&s.tt!=='Gấp');
    const q=spSrch.toLowerCase();
    const mS=!q||s.nd.toLowerCase().includes(q)||(s.loai||'').toLowerCase().includes(q)||
             (s.cnt||'').toLowerCase().includes(q)||(s.media||'').toLowerCase().includes(q);
    return mF&&mTT&&mS;
  }});
}}

function renderSocial(){{
  const list=getSPFiltd();
  const totAnh=list.reduce((a,s)=>a+s.sl_anh,0);
  document.getElementById('spInfo').innerHTML=`Hiển thị <strong>${{list.length}}</strong> / ${{SOCIAL.length}} bài · <strong>🖼 ${{totAnh}}</strong> ảnh thiết kế`;

  document.getElementById('spBody').innerHTML=list.map(s=>{{
    // tt badge
    const ttCls=s.tt==='Hoàn thành'?'vs-done':s.tt==='Gấp'?'vs-urgent':s.tt?'vs-doing':'vs-pend';
    const ttLbl=s.tt==='Hoàn thành'?'✅ Hoàn thành':s.tt==='Gấp'?'⚡ Gấp':s.tt||'⏳ Chờ';
    // anh badge – inferred if originally null (we set sl_anh_display same but track via comparison)
    const isInferred=(s.sl_anh===s.sl_bai);
    const anhClass=isInferred?'anh-badge inferred':'anh-badge';
    const anhTitle=isInferred?('title="Inferred từ SL Bài ('+s.sl_bai+')"') : '';
    return`<tr>
      <td class="vc-stt">${{s.stt}}</td>
      <td class="vc-da" style="max-width:260px;white-space:normal;line-height:1.4;">${{s.nd}}</td>
      <td>${{loaiBadge(s.loai)}}</td>
      <td style="text-align:center;font-weight:700;color:var(--t1)">${{s.sl_bai}}</td>
      <td style="text-align:center"><span class="${{anhClass}}" ${{anhTitle}}>🖼 ${{s.sl_anh}}</span></td>
      <td class="vc-date">${{s.ngay||'–'}}</td>
      <td style="color:var(--navy);font-weight:600">${{s.cnt||'–'}}</td>
      <td class="vc-pic">${{s.media||'–'}}</td>
      <td style="color:var(--t2);font-size:11.5px;max-width:130px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;" title="${{s.link}}">${{s.link||'–'}}</td>
      <td><span class="${{ttCls}}">${{ttLbl}}</span></td>
    </tr>`;
  }}).join('');
}}

// ══ INIT ══
document.addEventListener('DOMContentLoaded',()=>{{
  renderTaskStats();renderShootStats();renderVideoStats();renderSocialStats();
  renderProjs();renderShoot();renderVideos();renderSocial();

  // task
  document.querySelectorAll('#tFilters .fbtn').forEach(b=>b.addEventListener('click',()=>{{
    document.querySelectorAll('#tFilters .fbtn').forEach(x=>x.classList.remove('active'));
    b.classList.add('active');tFil=b.dataset.f;tPage=1;renderProjs();
  }}));
  let tt;document.getElementById('tSearch').addEventListener('input',e=>{{
    clearTimeout(tt);tt=setTimeout(()=>{{tSrch=e.target.value.trim();tPage=1;renderProjs();}},220);
  }});
  // shoot
  document.querySelectorAll('#sFilters .fbtn').forEach(b=>b.addEventListener('click',()=>{{
    document.querySelectorAll('#sFilters .fbtn').forEach(x=>x.classList.remove('active'));
    b.classList.add('active');sFil=b.dataset.f;renderShoot();
  }}));
  let st;document.getElementById('sSearch').addEventListener('input',e=>{{
    clearTimeout(st);st=setTimeout(()=>{{sSrch=e.target.value.trim();renderShoot();}},220);
  }});
  // video
  document.querySelectorAll('#vFilters .fbtn').forEach(b=>b.addEventListener('click',()=>{{
    document.querySelectorAll('#vFilters .fbtn').forEach(x=>x.classList.remove('active'));
    b.classList.add('active');vFil=b.dataset.f;renderVideos();
  }}));
  let vt;document.getElementById('vSearch').addEventListener('input',e=>{{
    clearTimeout(vt);vt=setTimeout(()=>{{vSrch=e.target.value.trim();renderVideos();}},220);
  }});
  // social – type filter
  document.querySelectorAll('#spFilters .fbtn').forEach(b=>b.addEventListener('click',()=>{{
    document.querySelectorAll('#spFilters .fbtn').forEach(x=>x.classList.remove('active'));
    b.classList.add('active');spFil=b.dataset.f;renderSocial();
  }}));
  // social – tt filter
  document.querySelectorAll('#spTTFilters .fbtn').forEach(b=>b.addEventListener('click',()=>{{
    document.querySelectorAll('#spTTFilters .fbtn').forEach(x=>x.classList.remove('active'));
    b.classList.add('active');spTTFil=b.dataset.tf;renderSocial();
  }}));
  let spt;document.getElementById('spSearch').addEventListener('input',e=>{{
    clearTimeout(spt);spt=setTimeout(()=>{{spSrch=e.target.value.trim();renderSocial();}},220);
  }});
}});
</script>
</body>
</html>'''

with open('2026_T5_MEDIA_FULL.html', 'w', encoding='utf-8') as f:
    f.write(HTML)

print(f"Tasks: {ttasks} | Done: {tdone}")
print(f"Shooting: {len(shooting)} | Delivered: {sdone_s}")
print(f"Videos: {len(videos)} | Done: {vdone}")
print(f"Social Posts: {len(social)} | Done: {sp_done} | Urgent: {sp_urgent}")
print(f"Tổng SL bài: {total_bai} | Tổng SL ảnh: {total_anh}")
print("Saved: 2026_T5_MEDIA_FULL.html")
