import sys
import io
import json
import re
from datetime import datetime, date

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl import load_workbook

# ============================================================
# PART 1: Extract Task Tracker from Excel (same as before)
# ============================================================
wb = load_workbook('2026_[MEDIA - TASKS TRACKER].xlsx')
ws = wb['All Task']

def extract_computed(val):
    if val is None:
        return None
    val = str(val)
    if 'COMPUTED_VALUE' in val or 'IFERROR' in val:
        m = re.search(r'COMPUTED_VALUE[^,]*,\"(.*?)\"\)', val)
        if m:
            return m.group(1)
        m = re.search(r',\"([^\"]+)\"\)$', val)
        if m:
            return m.group(1)
        return None
    return val

HEADERS = {
    'tt': 0, 'level1': 1, 'level2': 2, 'priority': 3,
    'pic': 4, 'support': 6, 'start': 7, 'end': 8,
    'extended_date': 9, 'days_left': 10, 'extension_reason': 11,
    'content': 12, 'status': 13, 'completion': 14,
    'cancel': 15, 'state': 16,
}

FILTER_FROM = date(2026, 5, 1)
FILTER_TO   = date(2026, 5, 31)

def parse_date(s):
    if not s:
        return None
    try:
        return datetime.strptime(s, '%Y-%m-%d').date()
    except:
        return None

def task_in_may2026(task):
    start = parse_date(task.get('start'))
    end_raw = task.get('extended') or task.get('end')
    end   = parse_date(end_raw)
    if start is None and end is None:
        return False
    if end is None:
        return FILTER_FROM <= start <= FILTER_TO
    if start is None:
        return FILTER_FROM <= end <= FILTER_TO
    return start <= FILTER_TO and end >= FILTER_FROM

all_projects_raw = []
current_project = None

for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
    if row_idx < 4:
        continue
    cells = []
    for cell in row:
        if isinstance(cell, datetime):
            cells.append(cell.strftime('%Y-%m-%d'))
        elif cell is None:
            cells.append(None)
        else:
            val = str(cell)
            if 'COMPUTED_VALUE' in val or 'IFERROR' in val:
                val = extract_computed(val)
            cells.append(val)

    if all(c is None for c in cells):
        continue

    def get(key):
        idx = HEADERS.get(key, -1)
        return cells[idx] if 0 <= idx < len(cells) else None

    level1 = get('level1')
    level2 = get('level2')
    tt = get('tt')

    if level1 and not level2:
        current_project = {'id': tt, 'name': level1, 'tasks': []}
        all_projects_raw.append(current_project)
    elif level2 and current_project is not None:
        try:
            comp_val = float(get('completion')) if get('completion') else 0
        except:
            comp_val = 0
        task = {
            'tt': tt, 'name': level2,
            'pic': get('pic'), 'start': get('start'), 'end': get('end'),
            'extended': get('extended_date'), 'completion': comp_val,
            'state': get('state'), 'cancel': get('cancel'),
            'content': get('content'), 'priority': get('priority'),
        }
        current_project['tasks'].append(task)

projects = []
for proj in all_projects_raw:
    filtered_tasks = [t for t in proj['tasks'] if task_in_may2026(t)]
    if filtered_tasks:
        projects.append({'id': proj['id'], 'name': proj['name'], 'tasks': filtered_tasks})

# ============================================================
# PART 2: Lịch quay chụp (hardcoded from image)
# ============================================================
shooting_schedule = [
    {"stt":1,  "chuong_trinh":"Hội nghị Đối tác 2026",                    "yeu_cau":"Event (Ms.Quyên)",  "dia_diem":"Thiên Hộ Plaza",                  "thu":"Chủ Nhật","ngay":"03/05","gio":"TBU",         "phu_trach":"Du, Vinh","tinh_trang":"Đã bàn giao"},
    {"stt":2,  "chuong_trinh":"Họp Phụ huynh Khu vực Sóc Trăng",          "yeu_cau":"Event (Ms.Quyên)",  "dia_diem":"Sóc Trăng",                       "thu":"Thứ 7",    "ngay":"09/05","gio":"TBU",         "phu_trach":"",        "tinh_trang":"Đã bàn giao"},
    {"stt":3,  "chuong_trinh":"Họp Phụ huynh Khu vực Vĩnh Long",          "yeu_cau":"Event (Ms.Quyên)",  "dia_diem":"TTC Palace",                      "thu":"Chủ Nhật","ngay":"10/05","gio":"TBU",         "phu_trach":"Vinh, Du","tinh_trang":"Đã bàn giao"},
    {"stt":4,  "chuong_trinh":"Khai trương Gò Công Đông",                  "yeu_cau":"Event (Ms.Quyên)",  "dia_diem":"CB Gò Công Đông",                 "thu":"TBU",      "ngay":"TBU",  "gio":"TBU",         "phu_trach":"",        "tinh_trang":""},
    {"stt":5,  "chuong_trinh":"Hội thảo 'Chương trình làm quen tiếng Anh'","yeu_cau":"MKT (Mr.Phong)",    "dia_diem":"TBU",                             "thu":"Thứ 4",    "ngay":"06/05","gio":"Sáng - TBU",  "phu_trach":"Vinh",    "tinh_trang":"Đã bàn giao"},
    {"stt":6,  "chuong_trinh":"SaiGon x Emasi Tour",                       "yeu_cau":"SBU WR (Ms.Thu)",   "dia_diem":"Emasi - SG",                      "thu":"Thứ 5",    "ngay":"14/05","gio":"TBU",         "phu_trach":"Vinh",    "tinh_trang":""},
    {"stt":7,  "chuong_trinh":"Store – Set up đèn quay phim",              "yeu_cau":"SBU Store (Ms.Trang)","dia_diem":"Studio",                         "thu":"Thứ 6",    "ngay":"15/05","gio":"Chiều 4:00",  "phu_trach":"Du",      "tinh_trang":"Đã bàn giao"},
    {"stt":8,  "chuong_trinh":"CB Mỹ Tho – Ngoại khóa tại trường",        "yeu_cau":"CB Mỹ Tho (Mr.Hoàng)","dia_diem":"Trường TH Thiên Hộ Dương CS1", "thu":"Thứ 2",    "ngay":"18/05","gio":"Sáng 7:00",   "phu_trach":"Du",      "tinh_trang":"Đã bàn giao"},
    {"stt":9,  "chuong_trinh":"CB Mỹ Tho – RCV tại trường",               "yeu_cau":"CB Mỹ Tho (Mr.Hoàng)","dia_diem":"Trường THPT Trần Hưng Đạo",    "thu":"Thứ 5",    "ngay":"21/05","gio":"Sáng 7:00",   "phu_trach":"Vinh",    "tinh_trang":"Đã bàn giao"},
    {"stt":10, "chuong_trinh":"Chụp ảnh lớp học CB Mekong",               "yeu_cau":"Media (Mr.Hòa)",    "dia_diem":"CB Mekong",                       "thu":"Thứ 7",    "ngay":"23/05","gio":"Sáng 9:15",   "phu_trach":"Vinh",    "tinh_trang":"Đã bàn giao"},
    {"stt":11, "chuong_trinh":"Chụp ảnh lớp học CB Mỹ Tho",              "yeu_cau":"Media (Mr.Hòa)",    "dia_diem":"CB Mỹ Tho",                       "thu":"Thứ 7",    "ngay":"23/05","gio":"Chiều 3:00",  "phu_trach":"Du",      "tinh_trang":"Đã bàn giao"},
    {"stt":12, "chuong_trinh":"Quay source giới thiệu Futura",             "yeu_cau":"SBU FTR (Mr.Duy)",  "dia_diem":"CB Mekong",                       "thu":"Thứ 6",    "ngay":"29/05","gio":"Chiều 2:00",  "phu_trach":"Du",      "tinh_trang":""},
    {"stt":13, "chuong_trinh":"Hội thi Hùng biện tiếng Anh x Đồng Tháp", "yeu_cau":"SBU FTR (Mr.Duy)",  "dia_diem":"CB Mekong",                       "thu":"Thứ 7",    "ngay":"30/05","gio":"TBU",         "phu_trach":"Vinh, Du","tinh_trang":""},
    {"stt":14, "chuong_trinh":"Store – Set up quay/chụp SP",              "yeu_cau":"SBU Store (Ms.Trang)","dia_diem":"Studio",                         "thu":"Thứ 3",    "ngay":"26/05","gio":"Chiều 3:30",  "phu_trach":"",        "tinh_trang":""},
]

# ============================================================
# PART 3: Video Checklist (hardcoded from image, May 2026)
# ============================================================
video_checklist = [
    {"stt":36, "du_an":"CB Adventures",            "video":"Giới thiệu Chương trình hè",     "kb":"Kịch bản video – chương trình hè", "nhan_order":"05/05/2026","ngay_quay":"",          "deadline":"16/05/2026","pic":"Du",       "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":"16/05/2026"},
    {"stt":37, "du_an":"CB Wonders",               "video":"Giới thiệu Chương trình học",    "kb":"Kịch bản video – chương trình hè", "nhan_order":"05/05/2026","ngay_quay":"",          "deadline":"16/05/2026","pic":"Du",       "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":"16/05/2026"},
    {"stt":38, "du_an":"CB Mỹ Tho",                "video":"Animation Ads PreS",             "kb":"Video ads mỹ tho",                 "nhan_order":"05/05/2026","ngay_quay":"",          "deadline":"11/05/2026","pic":"Vinh",     "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":"11/05/2026"},
    {"stt":39, "du_an":"CB Mỹ Tho",                "video":"Animation Ads PreS – Revised 1", "kb":"Video ads mỹ tho",                 "nhan_order":"11/05/2026","ngay_quay":"",          "deadline":"12/05/2026","pic":"Vinh",     "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":"12/05/2026"},
    {"stt":40, "du_an":"CB Mỹ Tho",                "video":"Animation Ads YLE",              "kb":"Video ads mỹ tho",                 "nhan_order":"05/05/2026","ngay_quay":"",          "deadline":"09/05/2026","pic":"Vinh",     "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":"09/05/2026"},
    {"stt":41, "du_an":"CB WorldReady",             "video":"Tour Sing",                      "kb":"CB Worldready – Singapore Tour",   "nhan_order":"05/05/2026","ngay_quay":"",          "deadline":"08/05/2026","pic":"Du",       "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":"08/05/2026"},
    {"stt":42, "du_an":"CB WorldReady",             "video":"Tour Emasi",                     "kb":"CB Worldready – Emasi Tour",       "nhan_order":"05/05/2026","ngay_quay":"",          "deadline":"09/05/2026","pic":"",         "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":"08/05/2026"},
    {"stt":43, "du_an":"CB Adventures",            "video":"Animation (hè)",                 "kb":"KB Video animation hè",            "nhan_order":"05/05/2026","ngay_quay":"",          "deadline":"",          "pic":"Vinh",     "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":""},
    {"stt":44, "du_an":"Hội thảo",                  "video":"Bộ quà tổng Hộp bút – Revised", "kb":"Cắt video phỏng vấn Hội thảo",    "nhan_order":"29/05/2026","ngay_quay":"",          "deadline":"29/05/2026","pic":"Vinh",     "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":"29/05/2026"},
    {"stt":45, "du_an":"Hội thảo",                  "video":"Cut raw",                        "kb":"Cắt video phỏng vấn Hội thảo",    "nhan_order":"05/05/2026","ngay_quay":"",          "deadline":"13/05/2026","pic":"Du",       "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":"12/05/2026"},
    {"stt":46, "du_an":"Hội thảo",                  "video":"Short PV",                       "kb":"Cắt video phỏng vấn Hội thảo",    "nhan_order":"05/05/2026","ngay_quay":"",          "deadline":"15/05/2026","pic":"Du",       "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":"13/05/2026"},
    {"stt":47, "du_an":"CX-Seri Khi Cici | CB talk","video":"Ep1 – Char",                    "kb":"",                                 "nhan_order":"",          "ngay_quay":"",          "deadline":"",          "pic":"Du",       "tien_do":0,  "tinh_trang":"",         "hoan_thanh":""},
    {"stt":48, "du_an":"CX-Seri Khi Cici | CB talk","video":"Ep1 – Storyboard",              "kb":"",                                 "nhan_order":"",          "ngay_quay":"",          "deadline":"",          "pic":"Du",       "tien_do":0,  "tinh_trang":"",         "hoan_thanh":""},
    {"stt":49, "du_an":"CX-Seri Khi Cici | CB talk","video":"Ep1 – Demo",                    "kb":"",                                 "nhan_order":"",          "ngay_quay":"",          "deadline":"",          "pic":"Du",       "tien_do":0,  "tinh_trang":"",         "hoan_thanh":""},
    {"stt":50, "du_an":"CB NEWS",                   "video":"#143",                           "kb":"CB News 143",                      "nhan_order":"12/05/2026","ngay_quay":"",          "deadline":"16/05/2026","pic":"Vinh",     "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":"16/05/2026"},
    {"stt":51, "du_an":"CB NEWS",                   "video":"Voice thông báo CB Summer",      "kb":"",                                 "nhan_order":"",          "ngay_quay":"",          "deadline":"",          "pic":"Du",       "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":"16/05/2026"},
    {"stt":52, "du_an":"Họp Phụ Huynh 2026",        "video":"Recap sự kiện họp PH TTC",       "kb":"",                                 "nhan_order":"",          "ngay_quay":"",          "deadline":"",          "pic":"Du",       "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":"15/05/2026"},
    {"stt":53, "du_an":"CB WorldReady",             "video":"Voice thông báo Tour Sing",      "kb":"CB Worldready – Phát thanh",       "nhan_order":"19/05/2026","ngay_quay":"19/05/2026","deadline":"",          "pic":"Du",       "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":"19/05/2026"},
    {"stt":54, "du_an":"CB WorldReady",             "video":"Voice thông báo Tour Emasi",     "kb":"CB Worldready – Phát thanh",       "nhan_order":"19/05/2026","ngay_quay":"19/05/2026","deadline":"",          "pic":"",         "tien_do":0,  "tinh_trang":"",         "hoan_thanh":""},
    {"stt":55, "du_an":"CB Chợ Gao",               "video":"Ads PreS",                       "kb":"Đổi địa chỉ 2 CS Chợ Gao",        "nhan_order":"19/05/2026","ngay_quay":"",          "deadline":"20/05/2026","pic":"Vinh",     "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":"20/05/2026"},
    {"stt":56, "du_an":"CB Chợ Gao",               "video":"Ads YLE",                        "kb":"Đổi địa chỉ 2 CS Chợ Gao",        "nhan_order":"19/05/2026","ngay_quay":"",          "deadline":"20/05/2026","pic":"",         "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":"20/05/2026"},
    {"stt":57, "du_an":"Về Xanh yêu Thương",        "video":"Voice thông báo",                "kb":"Phát thanh 2026",                  "nhan_order":"19/05/2026","ngay_quay":"",          "deadline":"19/05/2026","pic":"Du",       "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":"19/05/2026"},
    {"stt":58, "du_an":"CB Thừa Đức",              "video":"Ads YLE",                        "kb":"CB Khai trường TH Thừa Đức",       "nhan_order":"19/05/2026","ngay_quay":"",          "deadline":"31/05/2026","pic":"Vinh",     "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":"21/05/2026"},
    {"stt":59, "du_an":"Series Phụ huynh nói gì?", "video":"Phụ huynh: Diễm Thúy",          "kb":"Series Phụ huynh CB nói gì?",      "nhan_order":"21/05/2026","ngay_quay":"",          "deadline":"26/05/2026","pic":"Du",       "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":"01/06/2026"},
    {"stt":60, "du_an":"Series Phụ huynh nói gì?", "video":"Phụ huynh: Khánh Linh",          "kb":"Series Phụ huynh CB nói gì?",      "nhan_order":"21/05/2026","ngay_quay":"",          "deadline":"26/05/2026","pic":"Du",       "tien_do":100,"tinh_trang":"",         "hoan_thanh":""},
    {"stt":61, "du_an":"Series HV nói gì?",        "video":"Hải Phong",                      "kb":"",                                 "nhan_order":"21/05/2026","ngay_quay":"",          "deadline":"26/05/2026","pic":"Du",       "tien_do":100,"tinh_trang":"",         "hoan_thanh":""},
    {"stt":62, "du_an":"Series HV nói gì?",        "video":"Khánh Vy",                       "kb":"Series Học viên CB nói gì?",       "nhan_order":"21/05/2026","ngay_quay":"",          "deadline":"28/05/2026","pic":"Du",       "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":""},
    {"stt":63, "du_an":"Series HV nói gì?",        "video":"Nhú Uyên",                       "kb":"Series Học viên CB nói gì?",       "nhan_order":"21/05/2026","ngay_quay":"",          "deadline":"28/05/2026","pic":"Du",       "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":""},
    {"stt":64, "du_an":"Series HV nói gì?",        "video":"Bảo Lam",                        "kb":"Series Học viên CB nói gì?",       "nhan_order":"21/05/2026","ngay_quay":"",          "deadline":"28/05/2026","pic":"Du",       "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":""},
    {"stt":65, "du_an":"Series HV nói gì?",        "video":"Ngọc Hân",                       "kb":"Series Học viên CB nói gì?",       "nhan_order":"21/05/2026","ngay_quay":"",          "deadline":"28/05/2026","pic":"Du",       "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":""},
    {"stt":66, "du_an":"Series HV nói gì?",        "video":"Hữu Vĩnh",                       "kb":"Series Học viên CB nói gì?",       "nhan_order":"21/05/2026","ngay_quay":"",          "deadline":"28/05/2026","pic":"Du",       "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":""},
    {"stt":67, "du_an":"Series HV nói gì?",        "video":"Ngọc Thiện",                     "kb":"Series Học viên CB nói gì?",       "nhan_order":"21/05/2026","ngay_quay":"",          "deadline":"28/05/2026","pic":"Du",       "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":""},
    {"stt":68, "du_an":"CB Gò Công",               "video":"Ads PreS",                       "kb":"Đổi địa chỉ Gò Công",              "nhan_order":"26/05/2026","ngay_quay":"",          "deadline":"27/05/2026","pic":"Vinh",     "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":"27/05/2026"},
    {"stt":69, "du_an":"CB Gò Công",               "video":"Ads YLE",                        "kb":"Đổi địa chỉ Gò Công",              "nhan_order":"26/05/2026","ngay_quay":"",          "deadline":"27/05/2026","pic":"Du",       "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":"27/05/2026"},
    {"stt":70, "du_an":"Hệ thống",                  "video":"Ads PreS",                       "kb":"Hotline + web hệ thống",           "nhan_order":"26/05/2026","ngay_quay":"",          "deadline":"26/05/2026","pic":"Vinh",     "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":"27/05/2026"},
    {"stt":71, "du_an":"Hệ thống",                  "video":"Ads YLE",                        "kb":"Hotline + web hệ thống",           "nhan_order":"26/05/2026","ngay_quay":"",          "deadline":"26/05/2026","pic":"Vinh",     "tien_do":100,"tinh_trang":"",         "hoan_thanh":""},
    {"stt":72, "du_an":"Hội thi HBTA tỉnh Đồng Tháp","video":"Recap",                        "kb":"Recap",                            "nhan_order":"26/05/2026","ngay_quay":"",          "deadline":"05/06/2026","pic":"Du",       "tien_do":10, "tinh_trang":"Đang làm",  "hoan_thanh":""},
    {"stt":73, "du_an":"CB Futura Ads",             "video":"Futura Q2",                      "kb":"Ads Futura Quý 2",                 "nhan_order":"26/05/2026","ngay_quay":"29/05/2026","deadline":"",          "pic":"Du",       "tien_do":100,"tinh_trang":"Hoàn thành","hoan_thanh":"01/06/2026"},
]

# JSON
projects_json  = json.dumps(projects,          ensure_ascii=False)
shooting_json  = json.dumps(shooting_schedule, ensure_ascii=False)
video_json     = json.dumps(video_checklist,   ensure_ascii=False)

now_str = datetime.now().strftime('%H:%M – %d/%m/%Y')

# Stats
total_tasks = sum(len(p['tasks']) for p in projects)
tasks_done  = sum(1 for p in projects for t in p['tasks'] if (t.get('state') or '').find('Hoàn thành') >= 0)
tasks_doing = total_tasks - tasks_done

shoot_done  = sum(1 for s in shooting_schedule if s['tinh_trang'] == 'Đã bàn giao')
vid_done    = sum(1 for v in video_checklist   if v['tinh_trang'] == 'Hoàn thành')
vid_doing   = sum(1 for v in video_checklist   if v['tinh_trang'] == 'Đang làm')

print(f"Tasks T5: {total_tasks} | Done: {tasks_done} | Doing: {tasks_doing}")
print(f"Shoot: {len(shooting_schedule)} | Delivered: {shoot_done}")
print(f"Videos: {len(video_checklist)} | Done: {vid_done} | Doing: {vid_doing}")

html = f'''<!DOCTYPE html>
<html lang="vi">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Báo Cáo Tháng 5/2026 – Team Media CB</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700;900&display=swap" rel="stylesheet">
<style>
:root {{
  --brand-red:    #BA110F;
  --brand-navy:   #191970;
  --c-bg:         #f8fafc;
  --c-surface:    #ffffff;
  --c-card:       #ffffff;
  --c-card2:      #f1f5f9;
  --c-border:     #e2e8f0;
  --c-text:       #0f172a;
  --c-text2:      #475569;
  --c-text3:      #94a3b8;
  --c-purple:     var(--brand-navy);
  --c-purple-l:   var(--brand-navy);
  --c-cyan:       var(--brand-red);
  --c-teal:       #059669;
  --c-green:      #10b981;
  --c-amber:      #d97706;
  --c-orange:     #ea580c;
  --c-red:        #dc2626;
  --c-pink:       #db2777;
  --r-sm:  6px;
  --r-md:  10px;
  --r-lg:  12px;
}}

*,*::before,*::after {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family: 'Inter', sans-serif; background: var(--c-bg); color: var(--c-text); font-size: 14px; line-height: 1.5; }}
.wrap {{ max-width:1440px; margin:0 auto; padding:40px 28px; }}

.site-header {{ margin-bottom: 40px; border-bottom: 2px solid var(--brand-navy); padding-bottom: 20px; }}
.kicker {{ color: var(--brand-red); font-weight: 800; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 8px; }}
.hdr-title {{ font-size: 42px; font-weight: 900; color: var(--brand-navy); line-height: 1.1; }}

.kpi-row {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 20px; margin-bottom: 40px; }}
.kpi-block {{ background: var(--c-surface); border: 1px solid var(--c-border); padding: 20px; border-radius: var(--r-md); box-shadow: 0 2px 4px rgba(0,0,0,0.05); }}
.kpi-num {{ font-size: 32px; font-weight: 900; color: var(--brand-red); }}

.tab-nav {{ display: flex; gap: 10px; margin-bottom: 20px; }}
.tab-btn {{ padding: 10px 20px; border: 2px solid var(--brand-navy); background: transparent; color: var(--brand-navy); font-weight: 700; cursor: pointer; border-radius: var(--r-sm); }}
.tab-btn.active {{ background: var(--brand-navy); color: #fff; }}

.tab-panel {{ display: none; }}
.tab-panel.active {{ display: block; }}

.proj-card {{ background: var(--c-surface); border: 1px solid var(--c-border); border-radius: var(--r-md); margin-bottom: 10px; overflow: hidden; }}
.proj-head {{ padding: 15px 20px; font-weight: 700; background: #f8fafc; border-bottom: 1px solid var(--c-border); display: flex; justify-content: space-between; }}

.vc-table {{ width: 100%; border-collapse: collapse; background: var(--c-surface); }}
.vc-table th {{ background: var(--brand-navy); color: #fff; padding: 12px; text-align: left; }}
.vc-table td {{ padding: 12px; border-bottom: 1px solid var(--c-border); }}

footer {{ margin-top: 60px; padding: 20px; text-align: center; border-top: 1px solid var(--c-border); color: var(--c-text3); }}
</style>
</head>
<body>
<div class="wrap">
<header class="site-header">
        <span class="hdr-sep">·</span>
        <span>📍 HIAS Master · CB</span>
      </div>
    </div>
    <div class="hdr-right">
      <div class="period-badge">
        <div class="pb-icon">📅</div>
        <div>
          <div class="pb-label">Kỳ báo cáo</div>
          <div class="pb-val">01/05/2026 → 31/05/2026</div>
        </div>
      </div>
      <div class="gen-time">🕐 Xuất: {now_str}</div>
    </div>
  </div>
</header>

<!-- ============ KPI ROW ============ -->
<div class="kpi-row">
  <div class="kpi-block kb-tasks">
    <div class="kpi-icon">🎨</div>
    <div class="kpi-info">
      <div class="kpi-num">{total_tasks}</div>
      <div class="kpi-label">Task thiết kế tháng 5</div>
      <div class="kpi-pills">
        <span class="kpill done">✅ {tasks_done} Hoàn thành</span>
        <span class="kpill doing">🔵 {tasks_doing} Đang làm</span>
      </div>
    </div>
  </div>
  <div class="kpi-block kb-shoot">
    <div class="kpi-icon">📷</div>
    <div class="kpi-info">
      <div class="kpi-num">{len(shooting_schedule)}</div>
      <div class="kpi-label">Lịch quay chụp tháng 5</div>
      <div class="kpi-pills">
        <span class="kpill done">✅ {shoot_done} Đã bàn giao</span>
        <span class="kpill pend">⏳ {len(shooting_schedule)-shoot_done} Chờ/TBU</span>
      </div>
    </div>
  </div>
  <div class="kpi-block kb-video">
    <div class="kpi-icon">🎬</div>
    <div class="kpi-info">
      <div class="kpi-num">{len(video_checklist)}</div>
      <div class="kpi-label">Video Checklist tháng 5</div>
      <div class="kpi-pills">
        <span class="kpill done">✅ {vid_done} Hoàn thành</span>
        <span class="kpill doing">🔵 {vid_doing} Đang làm</span>
        <span class="kpill pend">⏳ {len(video_checklist)-vid_done-vid_doing} Chờ</span>
      </div>
    </div>
  </div>
</div>

<!-- ============ TABS ============ -->
<div class="tab-nav" role="tablist">
  <button class="tab-btn active" onclick="switchTab('tasks')" id="tab-tasks" role="tab">
    🎨 <span class="tb-txt">Công việc chính</span> <span class="tbadge">{total_tasks}</span>
  </button>
  <button class="tab-btn" onclick="switchTab('shoot')" id="tab-shoot" role="tab">
    📷 <span class="tb-txt">Lịch quay chụp</span> <span class="tbadge">{len(shooting_schedule)}</span>
  </button>
  <button class="tab-btn" onclick="switchTab('video')" id="tab-video" role="tab">
    🎬 <span class="tb-txt">Video Checklist</span> <span class="tbadge">{len(video_checklist)}</span>
  </button>
</div>

<!-- ============ TAB: TASKS ============ -->
<div class="tab-panel active" id="panel-tasks">

  <div class="sstat-grid" id="tasksStats"></div>

  <div class="controls">
    <div class="s-wrap">
      <span class="s-ico">🔍</span>
      <input type="text" id="taskSearch" placeholder="Tìm dự án, công việc, nhân sự..." autocomplete="off">
    </div>
    <div class="f-group" id="taskFilters">
      <button class="fbtn active" data-f="all">Tất cả</button>
      <button class="fbtn" data-f="Hoàn thành">✅ Hoàn thành</button>
      <button class="fbtn" data-f="Đang làm">🔵 Đang làm</button>
      <button class="fbtn" data-f="Hoãn">⏸ Hoãn</button>
    </div>
  </div>

  <div class="rbar">
    <div class="rinfo" id="taskInfo"></div>
    <button class="expand-btn" id="expandBtn" onclick="toggleAll('tasks')">⤢ Mở rộng tất cả</button>
  </div>

  <div class="proj-list" id="projList"></div>
  <div class="pagination" id="taskPager"></div>
</div>

<!-- ============ TAB: SHOOTING ============ -->
<div class="tab-panel" id="panel-shoot">

  <div class="sstat-grid" id="shootStats"></div>

  <div class="shoot-note">
    <span class="note-icon">⚠️</span>
    <div><strong style="color:var(--c-amber)">Lưu ý:</strong> Source hình ảnh bàn giao trong <strong>1 ngày</strong> kể từ khi kết thúc sự kiện.</div>
  </div>

  <div class="controls">
    <div class="s-wrap">
      <span class="s-ico">🔍</span>
      <input type="text" id="shootSearch" placeholder="Tìm sự kiện, địa điểm, phụ trách..." autocomplete="off">
    </div>
    <div class="f-group" id="shootFilters">
      <button class="fbtn active" data-f="all">Tất cả</button>
      <button class="fbtn" data-f="delivered">✅ Đã bàn giao</button>
      <button class="fbtn" data-f="pending">⏳ Chưa / TBU</button>
    </div>
  </div>

  <div class="rbar">
    <div class="rinfo" id="shootInfo"></div>
  </div>

  <div class="shoot-grid" id="shootGrid"></div>
</div>

<!-- ============ TAB: VIDEO ============ -->
<div class="tab-panel" id="panel-video">

  <div class="sstat-grid" id="videoStats"></div>

  <div class="controls">
    <div class="s-wrap">
      <span class="s-ico">🔍</span>
      <input type="text" id="videoSearch" placeholder="Tìm dự án, video, kịch bản, PIC..." autocomplete="off">
    </div>
    <div class="f-group" id="videoFilters">
      <button class="fbtn active" data-f="all">Tất cả</button>
      <button class="fbtn" data-f="Hoàn thành">✅ Hoàn thành</button>
      <button class="fbtn" data-f="Đang làm">🔵 Đang làm</button>
      <button class="fbtn" data-f="pending">⏳ Chờ xử lý</button>
    </div>
  </div>

  <div class="rbar">
    <div class="rinfo" id="videoInfo"></div>
  </div>

  <div class="vc-table-wrap">
    <table class="vc-table">
      <thead>
        <tr>
          <th style="text-align:center">#</th>
          <th>Dự án / CN</th>
          <th>Video</th>
          <th>Kịch bản</th>
          <th>Nhận order</th>
          <th>Ngày quay</th>
          <th>Deadline</th>
          <th>PIC</th>
          <th>Tiến độ</th>
          <th>Tình trạng</th>
          <th>Hoàn thành</th>
        </tr>
      </thead>
      <tbody id="vcBody"></tbody>
    </table>
  </div>
</div>

<!-- ============ FOOTER ============ -->
<footer>
  <p>📊 Báo cáo tháng 5/2026 · Team Media · CB Media Production Team · HIAS Master · Xuất: {now_str}</p>
</footer>

</div><!-- /wrap -->

<script>
// ==============================
// DATA
// ==============================
const PROJECTS   = {projects_json};
const SHOOTING   = {shooting_json};
const VIDEOS     = {video_json};

// ==============================
// TABS
// ==============================
function switchTab(id) {{
  document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
  document.querySelectorAll('.tab-panel').forEach(p => p.classList.remove('active'));
  document.getElementById('tab-' + id).classList.add('active');
  document.getElementById('panel-' + id).classList.add('active');
}}

// ==============================
// HELPERS
// ==============================
function sClass(state, cancel) {{
  if (cancel) return 'canceled';
  if (!state) return 'doing';
  if (state.includes('Hoàn thành')) return 'done';
  if (state.includes('Hoãn'))       return 'paused';
  if (state.includes('Quá hạn'))    return 'late';
  return 'doing';
}}

function sLabel(state, cancel) {{
  if (cancel) return '🚫 ' + cancel;
  if (!state) return '🔵 Đang làm';
  if (state.includes('Hoàn thành')) return '✅ Hoàn thành';
  if (state.includes('Hoãn'))       return '⏸ Hoãn';
  if (state.includes('Quá hạn'))    return '🔴 Quá hạn';
  return '🔵 Đang làm';
}}

function projPct(p) {{
  if (!p.tasks.length) return 0;
  const done = p.tasks.filter(t => sClass(t.state, t.cancel) === 'done').length;
  return Math.round(done / p.tasks.length * 100);
}}

function fmtD(d) {{
  if (!d) return '–';
  const pts = d.split('-');
  return pts.length === 3 ? `${{pts[2]}}/${{pts[1]}}` : d;
}}

// ==============================
// TASK SECTION
// ==============================
const PAGE = 20;
let tFilter = 'all', tSearch = '', tPage = 1, tFiltered = [];
let tasksAllOpen = false;

function renderTaskStats() {{
  let total=0, done=0, doing=0, paused=0;
  PROJECTS.forEach(p => p.tasks.forEach(t => {{
    total++;
    const sc = sClass(t.state, t.cancel);
    if (sc==='done') done++;
    else if (sc==='doing') doing++;
    else if (sc==='paused'||sc==='canceled') paused++;
  }}));
  const pct = total ? Math.round(done/total*100) : 0;

  document.getElementById('tasksStats').innerHTML = [
    {{v:PROJECTS.length, l:'Dự án',         c:'purple', i:'📁'}},
    {{v:total,           l:'Công việc',      c:'blue',   i:'📋'}},
    {{v:done,            l:'Hoàn thành',     c:'green',  i:'✅'}},
    {{v:doing,           l:'Đang thực hiện', c:'teal',   i:'🔵'}},
    {{v:paused,          l:'Hoãn / Huỷ',    c:'orange', i:'⏸'}},
    {{v:pct+'%',         l:'Tỉ lệ HT',      c:'pink',   i:'📈'}},
  ].map(s => `<div class="sstat ${{s.c}}"><div class="sstat-icon">${{s.i}}</div><div class="sstat-val">${{s.v}}</div><div class="sstat-lbl">${{s.l}}</div></div>`).join('');
}}

function getTaskFiltered() {{
  return PROJECTS.filter(p => {{
    const mF = tFilter==='all' || p.tasks.some(t=>(t.state||'').includes(tFilter));
    const q  = tSearch.toLowerCase();
    const mS = !q || p.name.toLowerCase().includes(q) ||
               p.tasks.some(t=>(t.name||'').toLowerCase().includes(q)||(t.pic||'').toLowerCase().includes(q));
    return mF && mS;
  }});
}}

function renderProjects() {{
  tFiltered = getTaskFiltered();
  const totalT = tFiltered.reduce((s,p) => s + p.tasks.length, 0);
  document.getElementById('taskInfo').innerHTML =
    `Hiển thị <strong>${{tFiltered.length}}</strong> dự án · <strong>${{totalT}}</strong> công việc`;

  const start = (tPage-1)*PAGE;
  const page  = tFiltered.slice(start, start+PAGE);

  document.getElementById('projList').innerHTML = page.length
    ? page.map((p,i) => renderProjCard(p, start+i)).join('')
    : `<div class="no-res"><span class="nr-icon">🔍</span><p>Không tìm thấy kết quả</p></div>`;

  renderTaskPager(tFiltered.length);
}}

function renderProjCard(p, idx) {{
  const pct  = projPct(p);
  const done = p.tasks.filter(t => sClass(t.state,t.cancel)==='done').length;
  return `<div class="proj-card" id="pc-${{idx}}">
    <div class="proj-head" onclick="toggleProj(${{idx}})">
      <div class="ph-left">
        <span class="ph-num">#${{p.id||idx+1}}</span>
        <span class="ph-name" title="${{p.name}}">${{p.name}}</span>
      </div>
      <div class="ph-right">
        <span class="tc-badge">${{done}}/${{p.tasks.length}} việc</span>
        <div class="ph-prog">
          <div class="ph-bar"><div class="ph-fill" style="width:${{pct}}%"></div></div>
          <span class="ph-pct">${{pct}}%</span>
        </div>
        <span class="chevron">▾</span>
      </div>
    </div>
    <div class="task-wrap">
      <table class="tt">
        <thead><tr><th>Công việc</th><th>PIC</th><th>Bắt đầu</th><th>Kết thúc</th><th>Trạng thái</th><th>% HT</th></tr></thead>
        <tbody>
          ${{p.tasks.map(t => `<tr>
            <td class="t-name">${{t.name||''}}</td>
            <td class="t-pic">${{t.pic||'–'}}</td>
            <td class="t-date">${{fmtD(t.start)}}</td>
            <td class="t-date${{t.extended?' ext':''}}">${{fmtD(t.extended||t.end)}}${{t.extended?' ⤴':''}}</td>
            <td><span class="sb ${{sClass(t.state,t.cancel)}}">${{sLabel(t.state,t.cancel)}}</span></td>
            <td><div class="c-wrap"><div class="c-bar"><div class="c-fill" style="width:${{Math.round((t.completion||0)*100)}}%"></div></div><span class="c-txt">${{Math.round((t.completion||0)*100)}}%</span></div></td>
          </tr>`).join('')}}
        </tbody>
      </table>
    </div>
  </div>`;
}}

function toggleProj(idx) {{ document.getElementById('pc-'+idx).classList.toggle('open'); }}

function toggleAll(section) {{
  tasksAllOpen = !tasksAllOpen;
  document.querySelectorAll('#projList .proj-card').forEach(c => c.classList.toggle('open', tasksAllOpen));
  document.getElementById('expandBtn').textContent = tasksAllOpen ? '⤡ Thu gọn tất cả' : '⤢ Mở rộng tất cả';
}}

function renderTaskPager(total) {{
  const pages = Math.ceil(total/PAGE);
  if (pages<=1) {{ document.getElementById('taskPager').innerHTML=''; return; }}
  let h = '';
  if (tPage>1) h += `<button class="pg" onclick="tGoPage(${{tPage-1}})">←</button>`;
  for (let p=1; p<=pages; p++) {{
    if (p===1||p===pages||Math.abs(p-tPage)<=1)
      h += `<button class="pg${{p===tPage?' active':''}}" onclick="tGoPage(${{p}})">${{p}}</button>`;
    else if (h.slice(-3) !== '…</') h += `<button class="pg" disabled>…</button>`;
  }}
  if (tPage<pages) h += `<button class="pg" onclick="tGoPage(${{tPage+1}})">→</button>`;
  document.getElementById('taskPager').innerHTML = h;
}}

function tGoPage(p) {{ tPage=p; renderProjects(); window.scrollTo({{top:0,behavior:'smooth'}}); }}

// ==============================
// SHOOTING SECTION
// ==============================
let shootFilter = 'all', shootSearch = '';

function renderShootStats() {{
  const total     = SHOOTING.length;
  const delivered = SHOOTING.filter(s => s.tinh_trang === 'Đã bàn giao').length;
  const pending   = total - delivered;
  const pics = {{}};
  SHOOTING.forEach(s => (s.phu_trach||'').split(',').forEach(n => {{ n=n.trim(); if(n) pics[n]=(pics[n]||0)+1; }}));

  document.getElementById('shootStats').innerHTML = [
    {{v:total,     l:'Lịch quay chụp',  c:'blue',  i:'📷'}},
    {{v:delivered, l:'Đã bàn giao',     c:'green', i:'✅'}},
    {{v:pending,   l:'Chờ / TBU',       c:'amber', i:'⏳', cls:'orange'}},
    {{v:Object.keys(pics).length, l:'Nhân sự tham gia', c:'purple', i:'👥'}},
  ].map(s => `<div class="sstat ${{s.cls||s.c}}"><div class="sstat-icon">${{s.i}}</div><div class="sstat-val">${{s.v}}</div><div class="sstat-lbl">${{s.l}}</div></div>`).join('');
}}

function getShootFiltered() {{
  return SHOOTING.filter(s => {{
    const mF = shootFilter === 'all'
      || (shootFilter === 'delivered' && s.tinh_trang === 'Đã bàn giao')
      || (shootFilter === 'pending'   && s.tinh_trang !== 'Đã bàn giao');
    const q  = shootSearch.toLowerCase();
    const mS = !q
      || s.chuong_trinh.toLowerCase().includes(q)
      || s.dia_diem.toLowerCase().includes(q)
      || s.phu_trach.toLowerCase().includes(q)
      || s.yeu_cau.toLowerCase().includes(q);
    return mF && mS;
  }});
}}

function renderShooting() {{
  const list = getShootFiltered();
  document.getElementById('shootInfo').innerHTML =
    `Hiển thị <strong>${{list.length}}</strong> / ${{SHOOTING.length}} lịch quay chụp`;

  if (!list.length) {{
    document.getElementById('shootGrid').innerHTML =
      `<div class="no-res" style="grid-column:1/-1"><span class="nr-icon">📷</span><p>Không tìm thấy kết quả</p></div>`;
    return;
  }}

  document.getElementById('shootGrid').innerHTML = list.map((s, i) => {{
    const statusHtml = s.tinh_trang === 'Đã bàn giao'
      ? `<span class="status-delivered">✅ Đã bàn giao</span>`
      : s.ngay === 'TBU'
      ? `<span class="status-tbu">📌 TBU</span>`
      : `<span class="status-pending">⏳ Chờ thực hiện</span>`;

    const picHtml = s.phu_trach
      ? s.phu_trach.split(',').map(n => `<span class="pic-chip">👤 ${{n.trim()}}</span>`).join('')
      : '<span style="color:var(--c-text3);font-size:12px">Chưa phân công</span>';

    const datetime = `${{s.thu !== 'TBU' ? s.thu : ''}} ${{s.ngay}} · ${{s.gio}}`.trim();

    return `<div class="shoot-card" style="animation-delay:${{i*.04}}s">
      <div class="shoot-top">
        <div class="shoot-stt">${{s.stt}}</div>
        <div class="shoot-title">${{s.chuong_trinh}}</div>
        <div class="shoot-status">${{statusHtml}}</div>
      </div>
      <div class="shoot-body">
        <div class="shoot-row">
          <span class="shoot-row-icon">📌</span>
          <div>
            <div class="shoot-row-label">Yêu cầu</div>
            <div class="shoot-row-val">${{s.yeu_cau}}</div>
          </div>
        </div>
        <div class="shoot-row">
          <span class="shoot-row-icon">📍</span>
          <div>
            <div class="shoot-row-label">Địa điểm</div>
            <div class="shoot-row-val">${{s.dia_diem}}</div>
          </div>
        </div>
        <div class="shoot-row">
          <span class="shoot-row-icon">🗓</span>
          <div>
            <div class="shoot-row-label">Thời gian</div>
            <div class="shoot-row-val">
              <span class="datetime-chip">🗓 ${{s.ngay}}/2026 &nbsp;·&nbsp; ${{s.gio}} &nbsp;(${{s.thu}})</span>
            </div>
          </div>
        </div>
        <div class="shoot-row">
          <span class="shoot-row-icon">👤</span>
          <div>
            <div class="shoot-row-label">Phụ trách</div>
            <div class="pic-chips">${{picHtml}}</div>
          </div>
        </div>
      </div>
    </div>`;
  }}).join('');
}}

// ==============================
// VIDEO SECTION
// ==============================
let vidFilter = 'all', vidSearch = '';

function renderVideoStats() {{
  const total  = VIDEOS.length;
  const done   = VIDEOS.filter(v => v.tinh_trang === 'Hoàn thành').length;
  const doing  = VIDEOS.filter(v => v.tinh_trang === 'Đang làm').length;
  const pend   = total - done - doing;
  const pct    = total ? Math.round(done/total*100) : 0;

  document.getElementById('videoStats').innerHTML = [
    {{v:total, l:'Tổng video',     c:'orange', i:'🎬'}},
    {{v:done,  l:'Hoàn thành',     c:'green',  i:'✅'}},
    {{v:doing, l:'Đang xử lý',     c:'blue',   i:'🔵'}},
    {{v:pend,  l:'Chờ xử lý',      c:'orange', i:'⏳'}},
    {{v:pct+'%',l:'Tỉ lệ HT',     c:'pink',   i:'📈'}},
  ].map(s => `<div class="sstat ${{s.c}}"><div class="sstat-icon">${{s.i}}</div><div class="sstat-val">${{s.v}}</div><div class="sstat-lbl">${{s.l}}</div></div>`).join('');
}}

function getVidFiltered() {{
  return VIDEOS.filter(v => {{
    const mF = vidFilter === 'all'
      || (vidFilter === 'Hoàn thành' && v.tinh_trang === 'Hoàn thành')
      || (vidFilter === 'Đang làm'   && v.tinh_trang === 'Đang làm')
      || (vidFilter === 'pending'    && v.tinh_trang !== 'Hoàn thành' && v.tinh_trang !== 'Đang làm');
    const q  = vidSearch.toLowerCase();
    const mS = !q
      || v.du_an.toLowerCase().includes(q)
      || v.video.toLowerCase().includes(q)
      || (v.kb||'').toLowerCase().includes(q)
      || (v.pic||'').toLowerCase().includes(q);
    return mF && mS;
  }});
}}

function renderVideos() {{
  const list = getVidFiltered();
  document.getElementById('videoInfo').innerHTML =
    `Hiển thị <strong>${{list.length}}</strong> / ${{VIDEOS.length}} video`;

  const progColor = pct => pct === 100 ? '#10b981' : pct > 0 ? '#3b82f6' : '#475569';

  document.getElementById('vcBody').innerHTML = list.map(v => {{
    const sc = v.tinh_trang === 'Hoàn thành' ? 'vs-done'
             : v.tinh_trang === 'Đang làm'   ? 'vs-doing'
             : 'vs-pend';
    const slabel = v.tinh_trang === 'Hoàn thành' ? '✅ Hoàn thành'
                 : v.tinh_trang === 'Đang làm'   ? '🔵 Đang làm'
                 : v.tinh_trang                  || '⏳ Chờ xử lý';

    return `<tr>
      <td class="vc-stt">${{v.stt}}</td>
      <td class="vc-da">${{v.du_an}}</td>
      <td class="vc-vid">${{v.video}}</td>
      <td style="color:var(--c-text2);font-size:11.5px">${{v.kb||'–'}}</td>
      <td class="vc-date">${{v.nhan_order||'–'}}</td>
      <td class="vc-date">${{v.ngay_quay||'–'}}</td>
      <td class="vc-deadline">${{v.deadline||'–'}}</td>
      <td class="vc-pic">${{v.pic||'–'}}</td>
      <td>
        <div class="prog-inline">
          <div class="p-bar"><div class="p-fill" style="width:${{v.tien_do}}%;background:${{progColor(v.tien_do)}}"></div></div>
          <span class="p-txt">${{v.tien_do}}%</span>
        </div>
      </td>
      <td><span class="sb ${{sc}}">${{slabel}}</span></td>
      <td class="vc-done">${{v.hoan_thanh||'–'}}</td>
    </tr>`;
  }}).join('');
}}

// ==============================
// INIT
// ==============================
document.addEventListener('DOMContentLoaded', () => {{

  renderTaskStats();
  renderShootStats();
  renderVideoStats();
  renderProjects();
  renderShooting();
  renderVideos();

  // Task filters
  document.querySelectorAll('#taskFilters .fbtn').forEach(b => {{
    b.addEventListener('click', () => {{
      document.querySelectorAll('#taskFilters .fbtn').forEach(x => x.classList.remove('active'));
      b.classList.add('active');
      tFilter = b.dataset.f; tPage = 1; renderProjects();
    }});
  }});

  let ts;
  document.getElementById('taskSearch').addEventListener('input', e => {{
    clearTimeout(ts);
    ts = setTimeout(() => {{ tSearch=e.target.value.trim(); tPage=1; renderProjects(); }}, 220);
  }});

  // Shoot filters
  document.querySelectorAll('#shootFilters .fbtn').forEach(b => {{
    b.addEventListener('click', () => {{
      document.querySelectorAll('#shootFilters .fbtn').forEach(x => x.classList.remove('active'));
      b.classList.add('active');
      shootFilter = b.dataset.f; renderShooting();
    }});
  }});

  let ss;
  document.getElementById('shootSearch').addEventListener('input', e => {{
    clearTimeout(ss);
    ss = setTimeout(() => {{ shootSearch=e.target.value.trim(); renderShooting(); }}, 220);
  }});

  // Video filters
  document.querySelectorAll('#videoFilters .fbtn').forEach(b => {{
    b.addEventListener('click', () => {{
      document.querySelectorAll('#videoFilters .fbtn').forEach(x => x.classList.remove('active'));
      b.classList.add('active');
      vidFilter = b.dataset.f; renderVideos();
    }});
  }});

  let vs;
  document.getElementById('videoSearch').addEventListener('input', e => {{
    clearTimeout(vs);
    vs = setTimeout(() => {{ vidSearch=e.target.value.trim(); renderVideos(); }}, 220);
  }});

}});
</script>
</body>
</html>
'''

with open('2026_T5_MEDIA_FULL_REPORT.html', 'w', encoding='utf-8') as f:
    f.write(html)

print("Done! Saved: 2026_T5_MEDIA_FULL_REPORT.html")
