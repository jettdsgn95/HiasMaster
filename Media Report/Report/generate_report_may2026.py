import sys
import io
import json
import re
from datetime import datetime, date

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl import load_workbook

wb = load_workbook('2026_[MEDIA - TASKS TRACKER].xlsx')
ws = wb['All Task']

def extract_computed(val):
    if val is None:
        return None
    val = str(val)
    if 'COMPUTED_VALUE' in val or 'IFERROR' in val:
        m = re.search(r'COMPUTED_VALUE[^,]*,\"(.*?)\"\)', val)
        if m:
            return m.group(1)
        m = re.search(r',\"([^\"]+)\"\)$', val)
        if m:
            return m.group(1)
        return None
    return val

HEADERS = {
    'tt': 0, 'level1': 1, 'level2': 2, 'priority': 3,
    'pic': 4, 'support': 6, 'start': 7, 'end': 8,
    'extended_date': 9, 'days_left': 10, 'extension_reason': 11,
    'content': 12, 'status': 13, 'completion': 14,
    'cancel': 15, 'state': 16,
}

# ---- Date filter: Tháng 5/2026 ----
FILTER_FROM = date(2026, 5, 1)
FILTER_TO   = date(2026, 5, 31)

def parse_date(s):
    if not s:
        return None
    try:
        return datetime.strptime(s, '%Y-%m-%d').date()
    except:
        return None

def task_in_may2026(task):
    """
    Task được tính vào tháng 5/2026 nếu:
    - start <= FILTER_TO AND end >= FILTER_FROM
    (khoảng [start, end] giao với [01/05, 31/05])
    Nếu không có ngày start/end thì bỏ qua.
    """
    start = parse_date(task.get('start'))
    # Dùng extended date nếu có, nếu không dùng end
    end_raw = task.get('extended') or task.get('end')
    end   = parse_date(end_raw)

    if start is None and end is None:
        return False

    # Nếu chỉ có start
    if end is None:
        return FILTER_FROM <= start <= FILTER_TO

    # Nếu chỉ có end
    if start is None:
        return FILTER_FROM <= end <= FILTER_TO

    # Có cả hai: khoảng giao nhau
    return start <= FILTER_TO and end >= FILTER_FROM

# ---- Extract data ----
all_projects_raw = []
current_project = None

for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
    if row_idx < 4:
        continue
    cells = []
    for cell in row:
        if isinstance(cell, datetime):
            cells.append(cell.strftime('%Y-%m-%d'))
        elif cell is None:
            cells.append(None)
        else:
            val = str(cell)
            if 'COMPUTED_VALUE' in val or 'IFERROR' in val:
                val = extract_computed(val)
            cells.append(val)

    if all(c is None for c in cells):
        continue

    def get(key):
        idx = HEADERS.get(key, -1)
        return cells[idx] if 0 <= idx < len(cells) else None

    level1 = get('level1')
    level2 = get('level2')
    tt = get('tt')

    if level1 and not level2:
        current_project = {'id': tt, 'name': level1, 'tasks': []}
        all_projects_raw.append(current_project)
    elif level2 and current_project is not None:
        try:
            comp_val = float(get('completion')) if get('completion') else 0
        except:
            comp_val = 0
        task = {
            'tt': tt, 'name': level2,
            'pic': get('pic'), 'start': get('start'), 'end': get('end'),
            'extended': get('extended_date'), 'completion': comp_val,
            'state': get('state'), 'cancel': get('cancel'),
            'content': get('content'), 'priority': get('priority'),
        }
        current_project['tasks'].append(task)

# ---- Filter: chỉ giữ task trong tháng 5/2026 ----
projects = []
for proj in all_projects_raw:
    filtered_tasks = [t for t in proj['tasks'] if task_in_may2026(t)]
    if filtered_tasks:
        projects.append({
            'id': proj['id'],
            'name': proj['name'],
            'tasks': filtered_tasks,
        })

print(f"Projects sau lọc: {len(projects)}")
print(f"Tasks sau lọc: {sum(len(p['tasks']) for p in projects)}")
states = {}
for p in projects:
    for t in p['tasks']:
        s = t.get('state') or 'Khác'
        states[s] = states.get(s, 0) + 1
print("States:", states)

# ---- JSON for JS ----
projects_json = json.dumps(projects, ensure_ascii=False)

# ---- Generate HTML ----
now_str = datetime.now().strftime('%d/%m/%Y %H:%M')

html = '''<!DOCTYPE html>
<html lang="vi">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Báo Cáo Tháng 5/2026 – Team Media</title>
<meta name="description" content="Báo cáo công việc tháng 5/2026 Team Media – Phòng Marketing. Tổng hợp tiến độ từ 01/05/2026 đến 31/05/2026.">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap" rel="stylesheet">
<style>
:root {
  --bg-primary: #080c18;
  --bg-secondary: #0d1226;
  --bg-card: #111827;
  --bg-card-hover: #182032;
  --accent-purple: #7c3aed;
  --accent-purple-light: #a855f7;
  --accent-blue: #3b82f6;
  --accent-cyan: #06b6d4;
  --accent-teal: #14b8a6;
  --accent-pink: #ec4899;
  --accent-green: #10b981;
  --accent-orange: #f59e0b;
  --accent-red: #ef4444;
  --text-primary: #f1f5f9;
  --text-secondary: #94a3b8;
  --text-muted: #475569;
  --border: rgba(148, 163, 184, 0.08);
  --border-bright: rgba(124, 58, 237, 0.35);
  --shadow-glow: 0 0 50px rgba(124, 58, 237, 0.12);
  --radius: 16px;
  --radius-sm: 10px;
}

*, *::before, *::after { margin: 0; padding: 0; box-sizing: border-box; }

body {
  font-family: 'Inter', sans-serif;
  background: var(--bg-primary);
  color: var(--text-primary);
  min-height: 100vh;
  overflow-x: hidden;
}

/* ========== BACKGROUND MESH ========== */
.bg-mesh {
  position: fixed; inset: 0; z-index: 0; pointer-events: none;
  background:
    radial-gradient(ellipse 70% 50% at 15% 0%,   rgba(124,58,237,0.18) 0%, transparent 55%),
    radial-gradient(ellipse 50% 40% at 90% 100%,  rgba(6,182,212,0.12)  0%, transparent 55%),
    radial-gradient(ellipse 35% 25% at 60% 40%,   rgba(59,130,246,0.06) 0%, transparent 55%);
}

/* Animated grid lines */
.bg-grid {
  position: fixed; inset: 0; z-index: 0; pointer-events: none;
  background-image:
    linear-gradient(rgba(148,163,184,0.03) 1px, transparent 1px),
    linear-gradient(90deg, rgba(148,163,184,0.03) 1px, transparent 1px);
  background-size: 60px 60px;
}

.container {
  position: relative; z-index: 1;
  max-width: 1380px; margin: 0 auto; padding: 0 24px;
}

/* ========== HEADER ========== */
header {
  padding: 44px 0 36px;
  border-bottom: 1px solid var(--border);
  margin-bottom: 44px;
}

.header-inner {
  display: flex; align-items: flex-start;
  justify-content: space-between; flex-wrap: wrap; gap: 24px;
}

.pill-badge {
  display: inline-flex; align-items: center; gap: 8px;
  padding: 5px 14px; border-radius: 50px;
  font-size: 11px; font-weight: 700; letter-spacing: 1.5px;
  text-transform: uppercase; margin-bottom: 14px;
}

.pill-badge.month {
  background: linear-gradient(135deg, rgba(124,58,237,0.25), rgba(6,182,212,0.15));
  border: 1px solid rgba(124,58,237,0.4);
  color: var(--accent-purple-light);
}

.pill-badge .dot {
  width: 6px; height: 6px; border-radius: 50%;
  background: var(--accent-cyan);
  animation: blink 1.8s ease-in-out infinite;
}

@keyframes blink {
  0%,100% { opacity:1; transform:scale(1); }
  50% { opacity:.3; transform:scale(.7); }
}

h1 {
  font-size: clamp(26px, 3.5vw, 46px);
  font-weight: 900; line-height: 1.1; margin-bottom: 10px;
  background: linear-gradient(135deg, #f1f5f9 0%, #a855f7 45%, #06b6d4 100%);
  -webkit-background-clip: text; -webkit-text-fill-color: transparent;
  background-clip: text;
}

.header-sub {
  font-size: 14px; color: var(--text-secondary); font-weight: 400;
  display: flex; align-items: center; gap: 8px; flex-wrap: wrap;
}

.header-sub .sep { color: var(--text-muted); }

.header-right {
  display: flex; flex-direction: column; align-items: flex-end; gap: 10px;
}

.date-range-badge {
  display: flex; align-items: center; gap: 10px;
  background: var(--bg-card); border: 1px solid var(--border-bright);
  border-radius: var(--radius-sm); padding: 12px 20px;
}

.date-range-badge .icon { font-size: 20px; }
.date-range-badge .range-label { font-size: 11px; color: var(--text-muted); font-weight: 600; letter-spacing: 1px; text-transform: uppercase; }
.date-range-badge .range-val { font-size: 14px; font-weight: 700; color: var(--text-primary); margin-top: 2px; }

.generated-at {
  font-size: 12px; color: var(--text-muted);
}

/* ========== STAT CARDS ========== */
.stats-grid {
  display: grid;
  grid-template-columns: repeat(6, 1fr);
  gap: 14px;
  margin-bottom: 36px;
}

@media(max-width:1100px) { .stats-grid { grid-template-columns: repeat(3,1fr); } }
@media(max-width:600px)  { .stats-grid { grid-template-columns: repeat(2,1fr); } }

.stat-card {
  background: var(--bg-card);
  border: 1px solid var(--border);
  border-radius: var(--radius);
  padding: 22px 20px;
  position: relative; overflow: hidden;
  cursor: default;
  transition: transform .22s, box-shadow .22s, border-color .22s;
  animation: fadeUp .5s ease both;
}

.stat-card:hover {
  transform: translateY(-5px);
  box-shadow: var(--shadow-glow);
  border-color: var(--border-bright);
}

@keyframes fadeUp {
  from { opacity:0; transform:translateY(22px); }
  to   { opacity:1; transform:translateY(0); }
}
.stat-card:nth-child(1){animation-delay:.04s}
.stat-card:nth-child(2){animation-delay:.09s}
.stat-card:nth-child(3){animation-delay:.14s}
.stat-card:nth-child(4){animation-delay:.19s}
.stat-card:nth-child(5){animation-delay:.24s}
.stat-card:nth-child(6){animation-delay:.29s}

.stat-card::before {
  content:''; position:absolute; top:0; left:0; right:0; height:3px;
  border-radius: var(--radius) var(--radius) 0 0;
}
.stat-card.c-purple::before { background: linear-gradient(90deg,#7c3aed,#a855f7); }
.stat-card.c-green::before  { background: linear-gradient(90deg,#059669,#10b981); }
.stat-card.c-blue::before   { background: linear-gradient(90deg,#2563eb,#06b6d4); }
.stat-card.c-orange::before { background: linear-gradient(90deg,#d97706,#f59e0b); }
.stat-card.c-red::before    { background: linear-gradient(90deg,#dc2626,#ef4444); }
.stat-card.c-pink::before   { background: linear-gradient(90deg,#db2777,#ec4899); }

.stat-icon { font-size: 26px; margin-bottom: 10px; }
.stat-value { font-size: 36px; font-weight: 800; line-height: 1; margin-bottom: 5px; }
.stat-label { font-size: 12px; color: var(--text-secondary); font-weight: 500; }

.stat-card.c-purple .stat-value { color: #a855f7; }
.stat-card.c-green  .stat-value { color: #10b981; }
.stat-card.c-blue   .stat-value { color: #06b6d4; }
.stat-card.c-orange .stat-value { color: #f59e0b; }
.stat-card.c-red    .stat-value { color: #ef4444; }
.stat-card.c-pink   .stat-value { color: #ec4899; }

/* ========== CHARTS ========== */
.charts-row {
  display: grid; grid-template-columns: 5fr 7fr;
  gap: 16px; margin-bottom: 36px;
}
@media(max-width:900px){ .charts-row { grid-template-columns: 1fr; } }

.chart-card {
  background: var(--bg-card);
  border: 1px solid var(--border);
  border-radius: var(--radius);
  padding: 26px;
}

.chart-title {
  font-size: 14px; font-weight: 700; color: var(--text-primary);
  margin-bottom: 22px;
  display: flex; align-items: center; gap: 10px;
}

.chart-title::before {
  content:''; display:inline-block;
  width:4px; height:18px; border-radius:2px;
  background: linear-gradient(180deg, #7c3aed, #06b6d4);
}

/* donut */
.donut-wrap { display:flex; align-items:center; gap:24px; flex-wrap:wrap; }
.donut-legend { display:flex; flex-direction:column; gap:12px; flex:1; min-width:140px; }
.legend-item { display:flex; align-items:center; gap:10px; font-size:13px; }
.legend-dot { width:10px; height:10px; border-radius:50%; flex-shrink:0; }
.legend-label { color:var(--text-secondary); flex:1; }
.legend-count { font-weight:700; font-size:14px; }

/* bar chart */
.bar-list { display:flex; flex-direction:column; gap:13px; }
.bar-row { }
.bar-hd { display:flex; justify-content:space-between; align-items:center; margin-bottom:5px; }
.bar-name { font-size:13px; font-weight:600; color:var(--text-primary); }
.bar-meta { display:flex; align-items:center; gap:8px; }
.bar-num  { font-size:12px; color:var(--text-secondary); }
.bar-pct-label { font-size:11px; color:var(--text-muted); min-width:30px; text-align:right; }
.bar-track { height:8px; background:rgba(255,255,255,0.05); border-radius:50px; overflow:hidden; }
.bar-fill  { height:100%; border-radius:50px; transition: width 1.3s cubic-bezier(.16,1,.3,1); }

/* ========== CONTROLS ========== */
.controls {
  display:flex; gap:12px; flex-wrap:wrap; align-items:center;
  margin-bottom:20px;
}

.search-wrap { flex:1; min-width:240px; position:relative; }
.search-wrap input {
  width:100%;
  background: var(--bg-card);
  border: 1px solid var(--border);
  border-radius: 50px;
  padding: 11px 20px 11px 44px;
  font-size:14px; font-family:'Inter',sans-serif;
  color: var(--text-primary); outline:none;
  transition: border-color .2s, box-shadow .2s;
}
.search-wrap input::placeholder { color: var(--text-muted); }
.search-wrap input:focus {
  border-color: var(--accent-purple);
  box-shadow: 0 0 0 3px rgba(124,58,237,.18);
}
.search-icon {
  position:absolute; left:16px; top:50%; transform:translateY(-50%);
  color:var(--text-muted); font-size:15px; pointer-events:none;
}

.filter-btns { display:flex; gap:8px; flex-wrap:wrap; }

.fbtn {
  padding: 9px 16px; border-radius:50px;
  border: 1px solid var(--border);
  background: var(--bg-card);
  color: var(--text-secondary);
  font-size:13px; font-weight:500;
  cursor:pointer; font-family:'Inter',sans-serif;
  transition: all .2s; white-space:nowrap;
}
.fbtn:hover { border-color: var(--accent-purple); color:var(--text-primary); }
.fbtn.active {
  background: linear-gradient(135deg,#7c3aed,#3b82f6);
  border-color:transparent; color:#fff;
  box-shadow: 0 4px 14px rgba(124,58,237,.35);
}

/* ========== SUMMARY BAR ========== */
.results-bar {
  display:flex; align-items:center; justify-content:space-between;
  margin-bottom:18px; flex-wrap:wrap; gap:8px;
}
.results-info { font-size:13px; color:var(--text-secondary); }
.results-info strong { color:var(--text-primary); }

/* expand all */
.expand-all-btn {
  padding:7px 14px; border-radius:8px;
  border:1px solid var(--border);
  background:var(--bg-card); color:var(--text-secondary);
  font-size:12px; font-weight:600; cursor:pointer;
  font-family:'Inter',sans-serif;
  transition:all .2s;
}
.expand-all-btn:hover { border-color:var(--accent-purple); color:var(--text-primary); }

/* ========== PROJECT CARD ========== */
.proj-list { display:flex; flex-direction:column; gap:12px; }

.proj-card {
  background: var(--bg-card);
  border: 1px solid var(--border);
  border-radius: var(--radius);
  overflow:hidden;
  transition: border-color .2s;
}
.proj-card:hover { border-color: rgba(124,58,237,.25); }

.proj-head {
  display:flex; align-items:center;
  justify-content:space-between;
  padding:18px 22px; cursor:pointer;
  gap:14px; user-select:none;
  transition: background .15s;
}
.proj-head:hover { background: var(--bg-card-hover); }

.proj-left { display:flex; align-items:center; gap:14px; flex:1; min-width:0; }

.proj-num {
  font-size:10px; font-weight:800; letter-spacing:.5px;
  color: var(--accent-purple-light);
  background: rgba(124,58,237,.15);
  border: 1px solid rgba(124,58,237,.28);
  padding:3px 10px; border-radius:50px; flex-shrink:0;
}

.proj-name {
  font-size:14px; font-weight:700; color:var(--text-primary);
  overflow:hidden; text-overflow:ellipsis; white-space:nowrap;
}

.proj-right { display:flex; align-items:center; gap:12px; flex-shrink:0; }

.task-badge {
  font-size:11px; color:var(--text-secondary);
  background:rgba(255,255,255,.05);
  padding:3px 10px; border-radius:50px; white-space:nowrap;
}

.prog-wrap { display:flex; align-items:center; gap:8px; min-width:110px; }
.prog-bar { flex:1; height:5px; background:rgba(255,255,255,.07); border-radius:50px; overflow:hidden; }
.prog-fill { height:100%; border-radius:50px; background:linear-gradient(90deg,#7c3aed,#06b6d4); }
.prog-pct { font-size:11px; font-weight:700; color:#06b6d4; min-width:34px; text-align:right; }

.chevron { color:var(--text-muted); font-size:16px; transition:transform .28s; flex-shrink:0; }
.proj-card.open .chevron { transform:rotate(180deg); }

/* ========== TASK TABLE ========== */
.task-wrap {
  max-height:0; overflow:hidden; transition:max-height .38s ease;
  border-top:1px solid transparent;
}
.proj-card.open .task-wrap {
  max-height:4000px;
  border-top-color: var(--border);
}

.task-table { width:100%; border-collapse:collapse; font-size:12.5px; }

.task-table th {
  padding:10px 14px;
  text-align:left;
  font-size:10px; font-weight:700; letter-spacing:1.2px;
  text-transform:uppercase; color:var(--text-muted);
  background:rgba(0,0,0,.25); white-space:nowrap;
}

.task-table td {
  padding:11px 14px;
  border-top:1px solid var(--border);
  vertical-align:middle;
}

.task-table tr:hover td { background:rgba(255,255,255,.02); }

.t-name { font-weight:500; color:var(--text-primary); }
.t-pic  { color:var(--accent-cyan); font-weight:600; font-size:12px; }
.t-date { color:var(--text-secondary); white-space:nowrap; font-size:11.5px; }
.t-date.ext { color:var(--accent-orange); }

/* status badge */
.sbadge {
  display:inline-flex; align-items:center; gap:4px;
  padding:2px 9px; border-radius:50px;
  font-size:10.5px; font-weight:700; white-space:nowrap;
}
.sbadge.done     { background:rgba(16,185,129,.13); color:#10b981; border:1px solid rgba(16,185,129,.28); }
.sbadge.doing    { background:rgba(59,130,246,.13);  color:#3b82f6; border:1px solid rgba(59,130,246,.28); }
.sbadge.late     { background:rgba(239,68,68,.13);   color:#ef4444; border:1px solid rgba(239,68,68,.28); }
.sbadge.paused   { background:rgba(245,158,11,.13);  color:#f59e0b; border:1px solid rgba(245,158,11,.28); }
.sbadge.canceled { background:rgba(148,163,184,.08); color:var(--text-secondary); border:1px solid var(--border); }

/* completion mini */
.comp-wrap { display:inline-flex; align-items:center; gap:6px; }
.comp-bar { width:52px; height:4px; background:rgba(255,255,255,.07); border-radius:50px; overflow:hidden; }
.comp-fill { height:100%; border-radius:50px; background:linear-gradient(90deg,#7c3aed,#06b6d4); }
.comp-txt { font-size:11px; color:var(--text-secondary); }

/* ========== PAGINATION ========== */
.pagination {
  display:flex; justify-content:center; align-items:center;
  gap:6px; margin-top:28px; flex-wrap:wrap;
}

.pg-btn {
  min-width:36px; height:36px; padding:0 10px;
  border-radius:8px; border:1px solid var(--border);
  background:var(--bg-card); color:var(--text-secondary);
  font-size:13px; cursor:pointer;
  display:flex; align-items:center; justify-content:center;
  font-family:'Inter',sans-serif; transition:all .18s;
}
.pg-btn:hover { border-color:var(--accent-purple); color:var(--text-primary); }
.pg-btn.active {
  background:linear-gradient(135deg,#7c3aed,#3b82f6);
  border-color:transparent; color:#fff;
  box-shadow:0 4px 12px rgba(124,58,237,.35);
}
.pg-btn:disabled { opacity:.35; cursor:default; }

/* ========== FOOTER ========== */
footer {
  margin-top:56px; padding:28px 0;
  border-top:1px solid var(--border);
  text-align:center; color:var(--text-muted); font-size:12.5px;
}

/* ========== NO RESULTS ========== */
.no-results {
  text-align:center; padding:64px 20px;
  color:var(--text-muted); font-size:15px;
}
.no-results .icon { font-size:52px; margin-bottom:12px; display:block; }

/* ========== RESPONSIVE ========== */
@media(max-width:900px) {
  .proj-right .prog-wrap { display:none; }
  .task-table th:nth-child(n+5), .task-table td:nth-child(n+5) { display:none; }
}

/* scrollbar */
::-webkit-scrollbar { width:5px; height:5px; }
::-webkit-scrollbar-track { background:var(--bg-primary); }
::-webkit-scrollbar-thumb { background:rgba(124,58,237,.35); border-radius:3px; }
</style>
</head>
<body>
<div class="bg-mesh"></div>
<div class="bg-grid"></div>

<div class="container">

  <!-- ======== HEADER ======== -->
  <header>
    <div class="header-inner">
      <div>
        <div class="pill-badge month"><span class="dot"></span> Tháng 5 · 2026</div>
        <h1>BÁO CÁO CÔNG VIỆC<br>THÁNG 5/2026</h1>
        <div class="header-sub">
          <span>📂 Team Media</span>
          <span class="sep">·</span>
          <span>🏢 Phòng Marketing</span>
          <span class="sep">·</span>
          <span>📊 HIAS Master</span>
        </div>
      </div>
      <div class="header-right">
        <div class="date-range-badge">
          <span class="icon">📅</span>
          <div>
            <div class="range-label">Kỳ báo cáo</div>
            <div class="range-val">01/05/2026 → 31/05/2026</div>
          </div>
        </div>
        <div class="generated-at">🕐 Xuất báo cáo: ''' + now_str + '''</div>
      </div>
    </div>
  </header>

  <!-- ======== STATS ======== -->
  <div class="stats-grid" id="statsGrid"></div>

  <!-- ======== CHARTS ======== -->
  <div class="charts-row">
    <div class="chart-card">
      <div class="chart-title">Trạng thái công việc</div>
      <div class="donut-wrap">
        <svg id="donutSvg" width="136" height="136" viewBox="0 0 136 136" style="flex-shrink:0"></svg>
        <div class="donut-legend" id="donutLegend"></div>
      </div>
    </div>
    <div class="chart-card">
      <div class="chart-title">Phân công nhân sự – Tháng 5</div>
      <div class="bar-list" id="picBars"></div>
    </div>
  </div>

  <!-- ======== CONTROLS ======== -->
  <div class="controls">
    <div class="search-wrap">
      <span class="search-icon">🔍</span>
      <input type="text" id="searchInput" placeholder="Tìm dự án, công việc, tên người phụ trách..." autocomplete="off">
    </div>
    <div class="filter-btns" id="filterBtns">
      <button class="fbtn active" data-f="all">Tất cả</button>
      <button class="fbtn" data-f="Hoàn thành">✅ Hoàn thành</button>
      <button class="fbtn" data-f="Đang làm">🔵 Đang làm</button>
      <button class="fbtn" data-f="Hoãn">⏸ Hoãn</button>
      <button class="fbtn" data-f="Quá hạn">🔴 Quá hạn</button>
    </div>
  </div>

  <div class="results-bar">
    <div class="results-info" id="resultsInfo"></div>
    <button class="expand-all-btn" id="expandAllBtn" onclick="toggleAll()">⤢ Mở rộng tất cả</button>
  </div>

  <!-- ======== LIST ======== -->
  <div class="proj-list" id="projList"></div>

  <!-- ======== PAGINATION ======== -->
  <div class="pagination" id="pagination"></div>

  <!-- ======== FOOTER ======== -->
  <footer>
    <p>📊 Báo cáo tháng 5/2026 · Kỳ: 01/05/2026 – 31/05/2026 · Team Media · HIAS Master</p>
  </footer>
</div>

<script>
// ============= DATA =============
const ALL_PROJECTS = ''' + projects_json + ''';

// ============= HELPERS =============
function sClass(state, cancel) {
  if (cancel) return 'canceled';
  if (!state)  return 'doing';
  if (state.includes('Hoàn thành')) return 'done';
  if (state.includes('Hoãn'))       return 'paused';
  if (state.includes('Quá hạn'))    return 'late';
  if (state.includes('Đang làm'))   return 'doing';
  return 'doing';
}

function sLabel(state, cancel) {
  if (cancel) return '🚫 ' + cancel;
  if (!state)  return '🔵 Đang làm';
  if (state.includes('Hoàn thành')) return '✅ Hoàn thành';
  if (state.includes('Hoãn'))       return '⏸ Hoãn';
  if (state.includes('Quá hạn'))    return '🔴 Quá hạn';
  if (state.includes('Đang làm'))   return '🔵 Đang làm';
  return state;
}

function projPct(proj) {
  if (!proj.tasks.length) return 0;
  const done = proj.tasks.filter(t => sClass(t.state, t.cancel) === 'done').length;
  return Math.round(done / proj.tasks.length * 100);
}

function fmtDate(d) {
  if (!d) return '–';
  const parts = d.split('-');
  if (parts.length === 3) return `${parts[2]}/${parts[1]}`;
  return d;
}

// ============= STATS =============
function renderStats() {
  let total = 0, done = 0, doing = 0, paused = 0, late = 0;
  ALL_PROJECTS.forEach(p => p.tasks.forEach(t => {
    total++;
    const sc = sClass(t.state, t.cancel);
    if (sc === 'done')     done++;
    else if (sc === 'doing')   doing++;
    else if (sc === 'paused' || sc === 'canceled') paused++;
    else if (sc === 'late')    late++;
  }));
  const pct = total ? Math.round(done / total * 100) : 0;

  const cards = [
    { icon:'📁', val: ALL_PROJECTS.length, lbl:'Dự án có task', cls:'c-purple' },
    { icon:'📋', val: total,  lbl:'Công việc',     cls:'c-blue'   },
    { icon:'✅', val: done,   lbl:'Hoàn thành',    cls:'c-green'  },
    { icon:'🔵', val: doing,  lbl:'Đang thực hiện',cls:'c-blue'   },
    { icon:'⏸',  val: paused, lbl:'Hoãn / Huỷ',   cls:'c-orange' },
    { icon:'📈', val: pct+'%',lbl:'Tỉ lệ hoàn thành',cls:'c-pink'},
  ];

  document.getElementById('statsGrid').innerHTML = cards.map(c => `
    <div class="stat-card ${c.cls}">
      <div class="stat-icon">${c.icon}</div>
      <div class="stat-value">${c.val}</div>
      <div class="stat-label">${c.lbl}</div>
    </div>`).join('');
}

// ============= DONUT =============
function renderDonut() {
  let done=0,doing=0,paused=0,late=0;
  ALL_PROJECTS.forEach(p=>p.tasks.forEach(t=>{
    const sc=sClass(t.state,t.cancel);
    if(sc==='done') done++;
    else if(sc==='doing') doing++;
    else if(sc==='paused'||sc==='canceled') paused++;
    else if(sc==='late') late++;
  }));
  const total=done+doing+paused+late;
  if(!total) return;

  const segs=[
    {lbl:'Hoàn thành',val:done,  col:'#10b981'},
    {lbl:'Đang làm',  val:doing, col:'#3b82f6'},
    {lbl:'Hoãn/Huỷ', val:paused,col:'#f59e0b'},
    {lbl:'Quá hạn',  val:late,  col:'#ef4444'},
  ].filter(s=>s.val>0);

  const cx=68,cy=68,r=50,sw=20,circ=2*Math.PI*r;
  let off=0, paths='';
  segs.forEach(s=>{
    const d=circ*(s.val/total), g=circ-d;
    paths+=`<circle cx="${cx}" cy="${cy}" r="${r}" fill="none" stroke="${s.col}" stroke-width="${sw}"
      stroke-dasharray="${d.toFixed(2)} ${g.toFixed(2)}"
      stroke-dashoffset="${(-off).toFixed(2)}"
      transform="rotate(-90 ${cx} ${cy})"/>`;
    off+=d;
  });
  const pct=Math.round(done/total*100);
  paths+=`<text x="${cx}" y="${cy-5}" text-anchor="middle" fill="#f1f5f9" font-size="20" font-weight="800" font-family="Inter">${pct}%</text>`;
  paths+=`<text x="${cx}" y="${cy+13}" text-anchor="middle" fill="#94a3b8" font-size="9.5" font-family="Inter">hoàn thành</text>`;
  document.getElementById('donutSvg').innerHTML=paths;

  document.getElementById('donutLegend').innerHTML=segs.map(s=>`
    <div class="legend-item">
      <span class="legend-dot" style="background:${s.col}"></span>
      <span class="legend-label">${s.lbl}</span>
      <span class="legend-count" style="color:${s.col}">${s.val}</span>
    </div>`).join('');
}

// ============= PIC BARS =============
function renderPicBars() {
  const pics={};
  ALL_PROJECTS.forEach(p=>p.tasks.forEach(t=>{
    (t.pic||'').split(',').forEach(name=>{
      name=name.trim();
      if(name) pics[name]=(pics[name]||0)+1;
    });
  }));
  const sorted=Object.entries(pics).sort((a,b)=>b[1]-a[1]).slice(0,8);
  if(!sorted.length){document.getElementById('picBars').innerHTML='<p style="color:var(--text-muted)">Không có dữ liệu</p>';return;}
  const max=sorted[0][1];
  const total=Object.values(pics).reduce((a,b)=>a+b,0);
  const colors=['#7c3aed','#3b82f6','#06b6d4','#10b981','#f59e0b','#ec4899','#a855f7','#14b8a6'];

  document.getElementById('picBars').innerHTML=sorted.map(([name,cnt],i)=>`
    <div class="bar-row">
      <div class="bar-hd">
        <span class="bar-name">👤 ${name}</span>
        <div class="bar-meta">
          <span class="bar-num">${cnt} task</span>
          <span class="bar-pct-label">${Math.round(cnt/total*100)}%</span>
        </div>
      </div>
      <div class="bar-track">
        <div class="bar-fill" style="width:${Math.round(cnt/max*100)}%;background:${colors[i]}"></div>
      </div>
    </div>`).join('');
}

// ============= PROJECT RENDER =============
const PAGE = 20;
let curFilter='all', curSearch='', curPage=1, filtered=[];

function getFiltered() {
  return ALL_PROJECTS.filter(p => {
    const matchF = curFilter==='all' || p.tasks.some(t=>(t.state||'').includes(curFilter));
    const q=curSearch.toLowerCase();
    const matchS = !q ||
      p.name.toLowerCase().includes(q) ||
      p.tasks.some(t=>(t.name||'').toLowerCase().includes(q)||(t.pic||'').toLowerCase().includes(q));
    return matchF && matchS;
  });
}

function renderProjCard(p, gIdx) {
  const pct = projPct(p);
  const done = p.tasks.filter(t=>sClass(t.state,t.cancel)==='done').length;
  return `
  <div class="proj-card" id="pc-${gIdx}">
    <div class="proj-head" onclick="toggle(${gIdx})">
      <div class="proj-left">
        <span class="proj-num">#${p.id||gIdx+1}</span>
        <span class="proj-name" title="${p.name}">${p.name}</span>
      </div>
      <div class="proj-right">
        <span class="task-badge">${done}/${p.tasks.length} việc</span>
        <div class="prog-wrap">
          <div class="prog-bar"><div class="prog-fill" style="width:${pct}%"></div></div>
          <span class="prog-pct">${pct}%</span>
        </div>
        <span class="chevron">▾</span>
      </div>
    </div>
    <div class="task-wrap">
      <table class="task-table">
        <thead><tr>
          <th>Công việc</th>
          <th>PIC</th>
          <th>Bắt đầu</th>
          <th>Kết thúc</th>
          <th>Trạng thái</th>
          <th>% HT</th>
        </tr></thead>
        <tbody>
          ${p.tasks.map(t=>`
          <tr>
            <td class="t-name">${t.name||''}</td>
            <td class="t-pic">${t.pic||'–'}</td>
            <td class="t-date">${fmtDate(t.start)}</td>
            <td class="t-date${t.extended?' ext':''}">${fmtDate(t.extended||t.end)}${t.extended?' ⤴':''}</td>
            <td><span class="sbadge ${sClass(t.state,t.cancel)}">${sLabel(t.state,t.cancel)}</span></td>
            <td>
              <div class="comp-wrap">
                <div class="comp-bar"><div class="comp-fill" style="width:${Math.round((t.completion||0)*100)}%"></div></div>
                <span class="comp-txt">${Math.round((t.completion||0)*100)}%</span>
              </div>
            </td>
          </tr>`).join('')}
        </tbody>
      </table>
    </div>
  </div>`;
}

function renderProjects() {
  filtered = getFiltered();
  const totalT = filtered.reduce((s,p)=>s+p.tasks.length,0);
  document.getElementById('resultsInfo').innerHTML =
    `Hiển thị <strong>${filtered.length}</strong> dự án &nbsp;·&nbsp; <strong>${totalT}</strong> công việc trong tháng 5/2026`;

  const start = (curPage-1)*PAGE;
  const page  = filtered.slice(start, start+PAGE);

  document.getElementById('projList').innerHTML = page.length
    ? page.map((p,i)=>renderProjCard(p,start+i)).join('')
    : `<div class="no-results"><span class="icon">🔍</span><p>Không tìm thấy kết quả phù hợp</p></div>`;

  renderPagination(filtered.length);
}

function renderPagination(total) {
  const pages = Math.ceil(total/PAGE);
  if(pages<=1){document.getElementById('pagination').innerHTML='';return;}

  const range=[];
  for(let p=1;p<=pages;p++){
    if(p===1||p===pages||Math.abs(p-curPage)<=2) range.push(p);
    else if(range[range.length-1]!=='…') range.push('…');
  }

  let h='';
  if(curPage>1) h+=`<button class="pg-btn" onclick="goPage(${curPage-1})">← Trước</button>`;
  range.forEach(p=>{
    if(p==='…') h+=`<button class="pg-btn" disabled>…</button>`;
    else h+=`<button class="pg-btn${p===curPage?' active':''}" onclick="goPage(${p})">${p}</button>`;
  });
  if(curPage<pages) h+=`<button class="pg-btn" onclick="goPage(${curPage+1})">Sau →</button>`;

  document.getElementById('pagination').innerHTML=h;
}

function goPage(p){curPage=p;renderProjects();window.scrollTo({top:0,behavior:'smooth'});}

function toggle(idx) {
  document.getElementById('pc-'+idx).classList.toggle('open');
}

let allOpen=false;
function toggleAll(){
  allOpen=!allOpen;
  document.querySelectorAll('.proj-card').forEach(c=>c.classList.toggle('open',allOpen));
  document.getElementById('expandAllBtn').textContent=allOpen?'⤡ Thu gọn tất cả':'⤢ Mở rộng tất cả';
}

// ============= INIT =============
document.addEventListener('DOMContentLoaded',()=>{
  renderStats();
  renderDonut();
  renderPicBars();
  renderProjects();

  document.querySelectorAll('.fbtn').forEach(btn=>{
    btn.addEventListener('click',()=>{
      document.querySelectorAll('.fbtn').forEach(b=>b.classList.remove('active'));
      btn.classList.add('active');
      curFilter=btn.dataset.f;
      curPage=1; renderProjects();
    });
  });

  let timer;
  document.getElementById('searchInput').addEventListener('input',e=>{
    clearTimeout(timer);
    timer=setTimeout(()=>{curSearch=e.target.value.trim();curPage=1;renderProjects();},220);
  });
});
</script>
</body>
</html>
'''

with open('2026_T5_MEDIA_REPORT.html', 'w', encoding='utf-8') as f:
    f.write(html)

print("Done! Saved: 2026_T5_MEDIA_REPORT.html")
