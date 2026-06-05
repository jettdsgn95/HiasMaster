import sys, io, json, re
from datetime import datetime, date
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook

# ── Excel extraction ──────────────────────────────────────
wb = load_workbook('2026_[MEDIA - TASKS TRACKER].xlsx')
ws = wb['All Task']

def ext(val):
    if val is None: return None
    val = str(val)
    if 'COMPUTED_VALUE' in val or 'IFERROR' in val:
        m = re.search(r'COMPUTED_VALUE[^,]*,\"(.*?)\"\)', val)
        if m: return m.group(1)
        m = re.search(r',\"([^\"]+)\"\)$', val)
        if m: return m.group(1)
        return None
    return val

HDR = {'tt':0,'level1':1,'level2':2,'priority':3,'pic':4,'support':6,
       'start':7,'end':8,'extended_date':9,'completion':14,'state':16,'cancel':15,'content':12}

D_FROM, D_TO = date(2026,5,1), date(2026,5,31)

def pd(s):
    try: return datetime.strptime(s,'%Y-%m-%d').date() if s else None
    except: return None

def in_may(t):
    s, e = pd(t.get('start')), pd(t.get('extended') or t.get('end'))
    if not s and not e: return False
    if not e: return D_FROM <= s <= D_TO
    if not s: return D_FROM <= e <= D_TO
    return s <= D_TO and e >= D_FROM

raw_projs, cur = [], None
for ri, row in enumerate(ws.iter_rows(values_only=True)):
    if ri < 4: continue
    cells = []
    for c in row:
        if isinstance(c, datetime): cells.append(c.strftime('%Y-%m-%d'))
        elif c is None: cells.append(None)
        else:
            v = str(c)
            cells.append(ext(v) if ('COMPUTED_VALUE' in v or 'IFERROR' in v) else v)
    if all(x is None for x in cells): continue
    def g(k): i=HDR.get(k,-1); return cells[i] if 0<=i<len(cells) else None
    l1, l2 = g('level1'), g('level2')
    if l1 and not l2:
        cur = {'id':g('tt'),'name':l1,'tasks':[]}; raw_projs.append(cur)
    elif l2 and cur:
        try: cv = float(g('completion')) if g('completion') else 0
        except: cv = 0
        cur['tasks'].append({'tt':g('tt'),'name':l2,'pic':g('pic'),'start':g('start'),
            'end':g('end'),'extended':g('extended_date'),'completion':cv,
            'state':g('state'),'cancel':g('cancel'),'content':g('content')})

projects = [{'id':p['id'],'name':p['name'],'tasks':[t for t in p['tasks'] if in_may(t)]}
            for p in raw_projs if any(in_may(t) for t in p['tasks'])]

# ── Static data ───────────────────────────────────────────
shooting = [
    {"stt":1, "ten":"Hội nghị Đối tác 2026",                    "yc":"Event (Ms.Quyên)",      "dd":"Thiên Hộ Plaza",               "thu":"Chủ Nhật","ngay":"03/05","gio":"TBU",        "pt":"Du, Vinh","tt":"Đã bàn giao"},
    {"stt":2, "ten":"Họp Phụ huynh KV Sóc Trăng",               "yc":"Event (Ms.Quyên)",      "dd":"Sóc Trăng",                    "thu":"Thứ 7",   "ngay":"09/05","gio":"TBU",        "pt":"",        "tt":"Đã bàn giao"},
    {"stt":3, "ten":"Họp Phụ huynh KV Vĩnh Long",               "yc":"Event (Ms.Quyên)",      "dd":"TTC Palace",                   "thu":"Chủ Nhật","ngay":"10/05","gio":"TBU",        "pt":"Vinh, Du","tt":"Đã bàn giao"},
    {"stt":4, "ten":"Khai trương Gò Công Đông",                  "yc":"Event (Ms.Quyên)",      "dd":"CB Gò Công Đông",              "thu":"TBU",     "ngay":"TBU",  "gio":"TBU",        "pt":"",        "tt":""},
    {"stt":5, "ten":"Hội thảo 'Chương trình làm quen tiếng Anh'","yc":"MKT (Mr.Phong)",        "dd":"TBU",                          "thu":"Thứ 4",   "ngay":"06/05","gio":"Sáng - TBU", "pt":"Vinh",    "tt":"Đã bàn giao"},
    {"stt":6, "ten":"SaiGon x Emasi Tour",                       "yc":"SBU WR (Ms.Thu)",       "dd":"Emasi – SG",                   "thu":"Thứ 5",   "ngay":"14/05","gio":"TBU",        "pt":"Vinh",    "tt":""},
    {"stt":7, "ten":"Store – Set up đèn quay phim",              "yc":"SBU Store (Ms.Trang)",  "dd":"Studio",                       "thu":"Thứ 6",   "ngay":"15/05","gio":"Chiều 4:00", "pt":"Du",      "tt":"Đã bàn giao"},
    {"stt":8, "ten":"CB Mỹ Tho – Ngoại khóa tại trường",        "yc":"CB Mỹ Tho (Mr.Hoàng)", "dd":"Trường TH Thiên Hộ Dương CS1","thu":"Thứ 2",   "ngay":"18/05","gio":"Sáng 7:00",  "pt":"Du",      "tt":"Đã bàn giao"},
    {"stt":9, "ten":"CB Mỹ Tho – RCV tại trường",               "yc":"CB Mỹ Tho (Mr.Hoàng)", "dd":"Trường THPT Trần Hưng Đạo",   "thu":"Thứ 5",   "ngay":"21/05","gio":"Sáng 7:00",  "pt":"Vinh",    "tt":"Đã bàn giao"},
    {"stt":10,"ten":"Chụp ảnh lớp học CB Mekong",               "yc":"Media (Mr.Hòa)",        "dd":"CB Mekong",                    "thu":"Thứ 7",   "ngay":"23/05","gio":"Sáng 9:15",  "pt":"Vinh",    "tt":"Đã bàn giao"},
    {"stt":11,"ten":"Chụp ảnh lớp học CB Mỹ Tho",              "yc":"Media (Mr.Hòa)",        "dd":"CB Mỹ Tho",                    "thu":"Thứ 7",   "ngay":"23/05","gio":"Chiều 3:00", "pt":"Du",      "tt":"Đã bàn giao"},
    {"stt":12,"ten":"Quay source giới thiệu Futura",             "yc":"SBU FTR (Mr.Duy)",      "dd":"CB Mekong",                    "thu":"Thứ 6",   "ngay":"29/05","gio":"Chiều 2:00", "pt":"Du",      "tt":""},
    {"stt":13,"ten":"Hội thi Hùng biện tiếng Anh x Đồng Tháp", "yc":"SBU FTR (Mr.Duy)",      "dd":"CB Mekong",                    "thu":"Thứ 7",   "ngay":"30/05","gio":"TBU",        "pt":"Vinh, Du","tt":""},
    {"stt":14,"ten":"Store – Set up quay/chụp SP",              "yc":"SBU Store (Ms.Trang)",  "dd":"Studio",                       "thu":"Thứ 3",   "ngay":"26/05","gio":"Chiều 3:30", "pt":"",        "tt":""},
]

videos = [
    {"stt":36,"da":"CB Adventures",             "vid":"Giới thiệu Chương trình hè",     "kb":"Kịch bản video – CT hè",  "ord":"05/05","quay":"",     "dl":"16/05","pic":"Du",    "pct":100,"tt":"Hoàn thành","done":"16/05"},
    {"stt":37,"da":"CB Wonders",                "vid":"Giới thiệu Chương trình học",    "kb":"Kịch bản video – CT hè",  "ord":"05/05","quay":"",     "dl":"16/05","pic":"Du",    "pct":100,"tt":"Hoàn thành","done":"16/05"},
    {"stt":38,"da":"CB Mỹ Tho",                 "vid":"Animation Ads PreS",             "kb":"Video ads mỹ tho",        "ord":"05/05","quay":"",     "dl":"11/05","pic":"Vinh",  "pct":100,"tt":"Hoàn thành","done":"11/05"},
    {"stt":39,"da":"CB Mỹ Tho",                 "vid":"Animation Ads PreS – Revised 1", "kb":"Video ads mỹ tho",        "ord":"11/05","quay":"",     "dl":"12/05","pic":"Vinh",  "pct":100,"tt":"Hoàn thành","done":"12/05"},
    {"stt":40,"da":"CB Mỹ Tho",                 "vid":"Animation Ads YLE",              "kb":"Video ads mỹ tho",        "ord":"05/05","quay":"",     "dl":"09/05","pic":"Vinh",  "pct":100,"tt":"Hoàn thành","done":"09/05"},
    {"stt":41,"da":"CB WorldReady",              "vid":"Tour Sing",                      "kb":"CB Worldready – SG Tour", "ord":"05/05","quay":"",     "dl":"08/05","pic":"Du",    "pct":100,"tt":"Hoàn thành","done":"08/05"},
    {"stt":42,"da":"CB WorldReady",              "vid":"Tour Emasi",                     "kb":"CB Worldready – Emasi Tour","ord":"05/05","quay":"",   "dl":"09/05","pic":"",      "pct":100,"tt":"Hoàn thành","done":"08/05"},
    {"stt":43,"da":"CB Adventures",             "vid":"Animation (hè)",                 "kb":"KB Video animation hè",   "ord":"05/05","quay":"",     "dl":"",      "pic":"Vinh",  "pct":100,"tt":"Hoàn thành","done":""},
    {"stt":44,"da":"Hội thảo",                   "vid":"Bộ quà tổng Hộp bút – Revised", "kb":"Cắt video phỏng vấn",     "ord":"29/05","quay":"",     "dl":"29/05","pic":"Vinh",  "pct":100,"tt":"Hoàn thành","done":"29/05"},
    {"stt":45,"da":"Hội thảo",                   "vid":"Cut raw",                        "kb":"Cắt video phỏng vấn",     "ord":"05/05","quay":"",     "dl":"13/05","pic":"Du",    "pct":100,"tt":"Hoàn thành","done":"12/05"},
    {"stt":46,"da":"Hội thảo",                   "vid":"Short PV",                       "kb":"Cắt video phỏng vấn",     "ord":"05/05","quay":"",     "dl":"15/05","pic":"Du",    "pct":100,"tt":"Hoàn thành","done":"13/05"},
    {"stt":47,"da":"CX – Khi Cici | CB talk",   "vid":"Ep1 – Char",                    "kb":"",                        "ord":"",     "quay":"",     "dl":"",      "pic":"Du",    "pct":0,  "tt":"",         "done":""},
    {"stt":48,"da":"CX – Khi Cici | CB talk",   "vid":"Ep1 – Storyboard",              "kb":"",                        "ord":"",     "quay":"",     "dl":"",      "pic":"Du",    "pct":0,  "tt":"",         "done":""},
    {"stt":49,"da":"CX – Khi Cici | CB talk",   "vid":"Ep1 – Demo",                    "kb":"",                        "ord":"",     "quay":"",     "dl":"",      "pic":"Du",    "pct":0,  "tt":"",         "done":""},
    {"stt":50,"da":"CB NEWS",                    "vid":"#143",                           "kb":"CB News 143",             "ord":"12/05","quay":"",     "dl":"16/05","pic":"Vinh",  "pct":100,"tt":"Hoàn thành","done":"16/05"},
    {"stt":51,"da":"CB NEWS",                    "vid":"Voice thông báo CB Summer",      "kb":"",                        "ord":"",     "quay":"",     "dl":"",      "pic":"Du",    "pct":100,"tt":"Hoàn thành","done":"16/05"},
    {"stt":52,"da":"Họp Phụ Huynh 2026",         "vid":"Recap sự kiện họp PH TTC",       "kb":"",                        "ord":"",     "quay":"",     "dl":"",      "pic":"Du",    "pct":100,"tt":"Hoàn thành","done":"15/05"},
    {"stt":53,"da":"CB WorldReady",              "vid":"Voice thông báo Tour Sing",      "kb":"CB Worldready – Phát thanh","ord":"19/05","quay":"19/05","dl":"",     "pic":"Du",    "pct":100,"tt":"Hoàn thành","done":"19/05"},
    {"stt":54,"da":"CB WorldReady",              "vid":"Voice thông báo Tour Emasi",     "kb":"CB Worldready – Phát thanh","ord":"19/05","quay":"19/05","dl":"",     "pic":"",      "pct":0,  "tt":"",         "done":""},
    {"stt":55,"da":"CB Chợ Gao",                "vid":"Ads PreS",                       "kb":"Đổi địa chỉ 2 CS",        "ord":"19/05","quay":"",     "dl":"20/05","pic":"Vinh",  "pct":100,"tt":"Hoàn thành","done":"20/05"},
    {"stt":56,"da":"CB Chợ Gao",                "vid":"Ads YLE",                        "kb":"Đổi địa chỉ 2 CS",        "ord":"19/05","quay":"",     "dl":"20/05","pic":"",      "pct":100,"tt":"Hoàn thành","done":"20/05"},
    {"stt":57,"da":"Về Xanh yêu Thương",         "vid":"Voice thông báo",                "kb":"Phát thanh 2026",          "ord":"19/05","quay":"",     "dl":"19/05","pic":"Du",    "pct":100,"tt":"Hoàn thành","done":"19/05"},
    {"stt":58,"da":"CB Thừa Đức",               "vid":"Ads YLE",                        "kb":"CB Khai trường TH Thừa Đức","ord":"19/05","quay":"",    "dl":"31/05","pic":"Vinh",  "pct":100,"tt":"Hoàn thành","done":"21/05"},
    {"stt":59,"da":"Series Phụ huynh nói gì?",  "vid":"Phụ huynh: Diễm Thúy",          "kb":"Series PH CB nói gì?",    "ord":"21/05","quay":"",     "dl":"26/05","pic":"Du",    "pct":100,"tt":"Hoàn thành","done":"01/06"},
    {"stt":60,"da":"Series Phụ huynh nói gì?",  "vid":"Phụ huynh: Khánh Linh",          "kb":"Series PH CB nói gì?",    "ord":"21/05","quay":"",     "dl":"26/05","pic":"Du",    "pct":100,"tt":"",         "done":""},
    {"stt":61,"da":"Series HV nói gì?",         "vid":"Hải Phong",                      "kb":"",                        "ord":"21/05","quay":"",     "dl":"26/05","pic":"Du",    "pct":100,"tt":"",         "done":""},
    {"stt":62,"da":"Series HV nói gì?",         "vid":"Khánh Vy",                       "kb":"Series HV CB nói gì?",    "ord":"21/05","quay":"",     "dl":"28/05","pic":"Du",    "pct":100,"tt":"Hoàn thành","done":""},
    {"stt":63,"da":"Series HV nói gì?",         "vid":"Nhú Uyên",                       "kb":"Series HV CB nói gì?",    "ord":"21/05","quay":"",     "dl":"28/05","pic":"Du",    "pct":100,"tt":"Hoàn thành","done":""},
    {"stt":64,"da":"Series HV nói gì?",         "vid":"Bảo Lam",                        "kb":"Series HV CB nói gì?",    "ord":"21/05","quay":"",     "dl":"28/05","pic":"Du",    "pct":100,"tt":"Hoàn thành","done":""},
    {"stt":65,"da":"Series HV nói gì?",         "vid":"Ngọc Hân",                       "kb":"Series HV CB nói gì?",    "ord":"21/05","quay":"",     "dl":"28/05","pic":"Du",    "pct":100,"tt":"Hoàn thành","done":""},
    {"stt":66,"da":"Series HV nói gì?",         "vid":"Hữu Vĩnh",                       "kb":"Series HV CB nói gì?",    "ord":"21/05","quay":"",     "dl":"28/05","pic":"Du",    "pct":100,"tt":"Hoàn thành","done":""},
    {"stt":67,"da":"Series HV nói gì?",         "vid":"Ngọc Thiện",                     "kb":"Series HV CB nói gì?",    "ord":"21/05","quay":"",     "dl":"28/05","pic":"Du",    "pct":100,"tt":"Hoàn thành","done":""},
    {"stt":68,"da":"CB Gò Công",                "vid":"Ads PreS",                       "kb":"Đổi địa chỉ Gò Công",     "ord":"26/05","quay":"",     "dl":"27/05","pic":"Vinh",  "pct":100,"tt":"Hoàn thành","done":"27/05"},
    {"stt":69,"da":"CB Gò Công",                "vid":"Ads YLE",                        "kb":"Đổi địa chỉ Gò Công",     "ord":"26/05","quay":"",     "dl":"27/05","pic":"Du",    "pct":100,"tt":"Hoàn thành","done":"27/05"},
    {"stt":70,"da":"Hệ thống",                   "vid":"Ads PreS",                       "kb":"Hotline + web hệ thống",  "ord":"26/05","quay":"",     "dl":"26/05","pic":"Vinh",  "pct":100,"tt":"Hoàn thành","done":"27/05"},
    {"stt":71,"da":"Hệ thống",                   "vid":"Ads YLE",                        "kb":"Hotline + web hệ thống",  "ord":"26/05","quay":"",     "dl":"26/05","pic":"Vinh",  "pct":100,"tt":"",         "done":""},
    {"stt":72,"da":"Hội thi HBTA tỉnh Đồng Tháp","vid":"Recap",                         "kb":"Recap",                   "ord":"26/05","quay":"",     "dl":"05/06","pic":"Du",    "pct":10, "tt":"Đang làm", "done":""},
    {"stt":73,"da":"CB Futura Ads",              "vid":"Futura Q2",                      "kb":"Ads Futura Quý 2",         "ord":"26/05","quay":"29/05","dl":"",      "pic":"Du",    "pct":100,"tt":"Hoàn thành","done":"01/06"},
]

# ── Stats ─────────────────────────────────────────────────
ttasks = sum(len(p['tasks']) for p in projects)
tdone  = sum(1 for p in projects for t in p['tasks'] if 'Hoàn thành' in (t.get('state') or ''))
sdone  = sum(1 for s in shooting if s['tt'] == 'Đã bàn giao')
vdone  = sum(1 for v in videos if v['tt'] == 'Hoàn thành')
vdoing = sum(1 for v in videos if v['tt'] == 'Đang làm')

pj  = json.dumps(projects,  ensure_ascii=False)
sj  = json.dumps(shooting,  ensure_ascii=False)
vj  = json.dumps(videos,    ensure_ascii=False)
now = datetime.now().strftime('%H:%M – %d/%m/%Y')

HTML = f'''<!DOCTYPE html>
<html lang="vi">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Báo Cáo Tháng 5/2026 – Team Media CB</title>
<meta name="description" content="Báo cáo tổng hợp tháng 5/2026: Task thiết kế, Lịch quay chụp và Video Checklist – Team Media CB.">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap" rel="stylesheet">
<style>
/* ════════════════════════════════════════════
   BRAND TOKENS  #BA110F · #191970
════════════════════════════════════════════ */
:root {{
  --red:       #BA110F;
  --red-dk:    #8C0D0B;
  --red-md:    #D4140F;
  --red-lt:    #E63836;
  --red-pale:  #FDE8E8;
  --red-tint:  #FFF4F4;
  --red-alpha: rgba(186,17,15,.1);

  --navy:      #191970;
  --navy-dk:   #0F0F50;
  --navy-md:   #222290;
  --navy-lt:   #3333BB;
  --navy-pale: #EAEAF8;
  --navy-tint: #F2F2FC;
  --navy-alpha:rgba(25,25,112,.1);

  /* Surfaces */
  --bg:        #F5F6FA;
  --surf:      #FFFFFF;
  --surf2:     #F8F9FD;
  --border:    #E2E5F0;
  --divider:   rgba(25,25,112,.07);

  /* Text */
  --t1:  #0D0D30;
  --t2:  #424275;
  --t3:  #8888BB;

  /* Status */
  --green:  #0A7A52;
  --amber:  #B07600;
  --orange: #C05000;

  /* Shadows */
  --sh-xs:  0 1px 3px rgba(25,25,112,.06);
  --sh-sm:  0 2px 8px rgba(25,25,112,.08);
  --sh-md:  0 4px 20px rgba(25,25,112,.12);
  --sh-red: 0 4px 16px rgba(186,17,15,.18);
  --sh-nav: 0 4px 16px rgba(25,25,112,.22);

  /* Radii */
  --r0: 6px;
  --r1: 10px;
  --r2: 16px;
  --r3: 24px;
}}

*, *::before, *::after {{ margin:0; padding:0; box-sizing:border-box; }}
html {{ scroll-behavior:smooth; }}
body {{
  font-family:'Inter',system-ui,sans-serif;
  background:var(--bg); color:var(--t1);
  font-size:14px; line-height:1.55;
  overflow-x:hidden;
}}

/* ── Scrollbar ── */
::-webkit-scrollbar {{ width:5px; height:5px; }}
::-webkit-scrollbar-track {{ background:var(--bg); }}
::-webkit-scrollbar-thumb {{ background:rgba(25,25,112,.25); border-radius:3px; }}
::-webkit-scrollbar-thumb:hover {{ background:var(--navy); }}

/* ════════════════════════════════════════════
   LAYOUT
════════════════════════════════════════════ */
.page {{ max-width:1440px; margin:0 auto; padding:0; }}

/* ════════════════════════════════════════════
   HERO HEADER
════════════════════════════════════════════ */
.hero {{
  background: linear-gradient(135deg, var(--navy-dk) 0%, var(--navy) 55%, var(--navy-md) 100%);
  padding: 44px 40px 38px;
  position: relative; overflow:hidden;
}}

/* Decorative circles */
.hero::before {{
  content:'';
  position:absolute; top:-80px; right:-80px;
  width:340px; height:340px; border-radius:50%;
  background:radial-gradient(circle, rgba(186,17,15,.3) 0%, transparent 70%);
  pointer-events:none;
}}
.hero::after {{
  content:'';
  position:absolute; bottom:-60px; left:20%;
  width:220px; height:220px; border-radius:50%;
  background:radial-gradient(circle, rgba(255,255,255,.05) 0%, transparent 70%);
  pointer-events:none;
}}

.hero-inner {{
  position:relative; z-index:1;
  display:flex; align-items:flex-end;
  justify-content:space-between; flex-wrap:wrap; gap:24px;
}}

.hero-left {{ flex:1; min-width:280px; }}

.hero-kicker {{
  display:inline-flex; align-items:center; gap:8px;
  background:rgba(255,255,255,.1); border:1px solid rgba(255,255,255,.18);
  border-radius:50px; padding:5px 16px; margin-bottom:14px;
  font-size:11px; font-weight:700; letter-spacing:1.8px;
  text-transform:uppercase; color:rgba(255,255,255,.9);
  backdrop-filter:blur(8px);
}}

.hero-dot {{
  width:6px; height:6px; border-radius:50%;
  background:#FF6B69; box-shadow:0 0 8px #FF6B69;
  animation:blink 2s ease-in-out infinite;
}}

@keyframes blink {{ 0%,100%{{opacity:1;transform:scale(1)}} 50%{{opacity:.25;transform:scale(.6)}} }}

.hero-title {{
  font-size:clamp(26px,3.5vw,48px);
  font-weight:900; line-height:1.08; margin-bottom:10px;
  color:#fff; letter-spacing:-.5px;
}}

.hero-title .hl {{ color:#FF8A88; }}

.hero-sub {{
  display:flex; align-items:center; gap:10px; flex-wrap:wrap;
  font-size:13px; color:rgba(255,255,255,.6);
}}
.hero-sub .sep {{ color:rgba(255,255,255,.25); }}

.hero-right {{
  display:flex; flex-direction:column; align-items:flex-end; gap:10px;
  flex-shrink:0;
}}

.period-card {{
  background:rgba(255,255,255,.1); border:1px solid rgba(255,255,255,.18);
  border-radius:var(--r1); padding:14px 20px;
  display:flex; align-items:center; gap:12px;
  backdrop-filter:blur(10px);
}}

.pc-icon {{ font-size:24px; }}
.pc-lbl {{ font-size:10px; font-weight:700; letter-spacing:1.2px; text-transform:uppercase; color:rgba(255,255,255,.45); }}
.pc-val {{ font-size:14px; font-weight:800; color:#fff; margin-top:2px; }}
.hero-time {{ font-size:11.5px; color:rgba(255,255,255,.4); }}

/* Red-to-Navy gradient accent stripe */
.accent-stripe {{
  height:4px;
  background:linear-gradient(90deg, var(--red) 0%, var(--red-lt) 30%, var(--navy-lt) 70%, var(--navy) 100%);
}}

/* ════════════════════════════════════════════
   KPI OVERVIEW ROW
════════════════════════════════════════════ */
.kpi-row {{
  display:grid; grid-template-columns:repeat(3,1fr);
  gap:16px; padding:28px 40px 24px;
  border-bottom:1px solid var(--divider);
}}

@media(max-width:800px) {{ .kpi-row {{ grid-template-columns:1fr; padding:20px; }} }}

.kpi-card {{
  background:var(--surf); border:1px solid var(--border);
  border-radius:var(--r2); padding:22px;
  display:flex; align-items:center; gap:18px;
  box-shadow:var(--sh-sm);
  position:relative; overflow:hidden;
  transition:transform .2s, box-shadow .2s;
}}

.kpi-card:hover {{ transform:translateY(-4px); box-shadow:var(--sh-md); }}

.kpi-card::before {{
  content:''; position:absolute;
  top:0; left:0; bottom:0; width:4px;
  border-radius:var(--r2) 0 0 var(--r2);
}}
.kpi-card.k-task::before {{ background:linear-gradient(180deg,var(--navy-dk),var(--navy-lt)); }}
.kpi-card.k-shoot::before {{ background:linear-gradient(180deg,var(--red-dk),var(--red-lt)); }}
.kpi-card.k-video::before {{ background:linear-gradient(180deg,#6B21A8,#A855F7); }}

.kpi-icon {{ font-size:40px; flex-shrink:0; }}
.kpi-info {{ flex:1; }}
.kpi-num {{ font-size:40px; font-weight:900; line-height:1; margin-bottom:4px; }}
.k-task  .kpi-num {{ color:var(--navy); }}
.k-shoot .kpi-num {{ color:var(--red); }}
.k-video .kpi-num {{ color:#7B21A8; }}
.kpi-lbl {{ font-size:13px; font-weight:600; color:var(--t2); }}
.kpi-chips {{ display:flex; gap:6px; margin-top:10px; flex-wrap:wrap; }}

.chip {{
  font-size:11px; font-weight:700; padding:3px 10px; border-radius:50px;
  white-space:nowrap;
}}
.chip-g {{ background:rgba(10,122,82,.1); color:var(--green); border:1px solid rgba(10,122,82,.2); }}
.chip-n {{ background:var(--navy-pale); color:var(--navy); border:1px solid rgba(25,25,112,.18); }}
.chip-a {{ background:rgba(176,118,0,.1); color:var(--amber); border:1px solid rgba(176,118,0,.18); }}

/* ════════════════════════════════════════════
   TABS
════════════════════════════════════════════ */
.tabs-wrap {{
  padding:24px 40px 0;
  position:sticky; top:0; z-index:200;
  background:rgba(245,246,250,.92);
  backdrop-filter:blur(14px);
  -webkit-backdrop-filter:blur(14px);
  border-bottom:1px solid var(--border);
}}

.tab-nav {{
  display:flex; gap:0;
  background:var(--surf); border:1px solid var(--border);
  border-radius:var(--r2); padding:5px;
  box-shadow:var(--sh-sm);
  width:fit-content;
}}

.tab-btn {{
  display:flex; align-items:center; gap:8px;
  padding:10px 20px; border-radius:var(--r1);
  background:transparent; border:none;
  color:var(--t2); font-size:13.5px; font-weight:600;
  cursor:pointer; font-family:inherit;
  transition:all .2s; white-space:nowrap;
}}
.tab-btn:hover {{ color:var(--navy); background:var(--navy-tint); }}
.tab-btn.active {{
  background:linear-gradient(135deg,var(--navy),var(--navy-lt));
  color:#fff; box-shadow:var(--sh-nav);
}}
.tab-btn.active.t-red {{
  background:linear-gradient(135deg,var(--red-dk),var(--red-md));
  box-shadow:var(--sh-red);
}}
.tab-btn.active.t-purple {{
  background:linear-gradient(135deg,#6B21A8,#9333EA);
  box-shadow:0 4px 16px rgba(107,33,168,.25);
}}

.tbadge {{
  font-size:10px; font-weight:800;
  padding:2px 8px; border-radius:50px;
  background:var(--navy-pale); color:var(--navy);
  min-width:22px; text-align:center;
}}
.tab-btn.active .tbadge {{ background:rgba(255,255,255,.25); color:#fff; }}

.tab-panel {{ display:none; }}
.tab-panel.active {{ display:block; animation:fadeUp .28s ease; }}
@keyframes fadeUp {{ from{{opacity:0;transform:translateY(6px)}} to{{opacity:1;transform:translateY(0)}} }}

/* ════════════════════════════════════════════
   CONTENT AREA
════════════════════════════════════════════ */
.content {{ padding:28px 40px 40px; }}

/* ── Section stat cards ── */
.scard-grid {{
  display:grid; grid-template-columns:repeat(auto-fit,minmax(150px,1fr));
  gap:12px; margin-bottom:22px;
}}

.scard {{
  background:var(--surf); border:1px solid var(--border);
  border-radius:var(--r1); padding:16px;
  box-shadow:var(--sh-xs); position:relative; overflow:hidden;
  transition:transform .18s, box-shadow .18s;
  animation:slideUp .38s ease both;
}}
.scard:hover {{ transform:translateY(-3px); box-shadow:var(--sh-sm); }}
@keyframes slideUp {{ from{{opacity:0;transform:translateY(12px)}} to{{opacity:1;transform:translateY(0)}} }}
.scard:nth-child(1){{animation-delay:.04s}} .scard:nth-child(2){{animation-delay:.08s}}
.scard:nth-child(3){{animation-delay:.12s}} .scard:nth-child(4){{animation-delay:.16s}}
.scard:nth-child(5){{animation-delay:.20s}} .scard:nth-child(6){{animation-delay:.24s}}

.scard::before {{
  content:''; position:absolute; top:0; left:0; right:0; height:3px;
  border-radius:var(--r1) var(--r1) 0 0;
}}
.scard.cn::before {{ background:linear-gradient(90deg,var(--navy-dk),var(--navy-lt)); }}
.scard.cr::before {{ background:linear-gradient(90deg,var(--red-dk),var(--red-lt)); }}
.scard.cg::before {{ background:linear-gradient(90deg,#0A5C3E,#0A7A52); }}
.scard.cb::before {{ background:linear-gradient(90deg,#1D4ED8,#60A5FA); }}
.scard.ca::before {{ background:linear-gradient(90deg,#92400E,#D97706); }}
.scard.cp::before {{ background:linear-gradient(90deg,#6B21A8,#A855F7); }}

.scard-icon {{ font-size:20px; margin-bottom:8px; }}
.scard-val  {{ font-size:30px; font-weight:900; line-height:1; margin-bottom:3px; }}
.scard-lbl  {{ font-size:11.5px; color:var(--t2); font-weight:500; }}
.scard.cn .scard-val {{ color:var(--navy); }}
.scard.cr .scard-val {{ color:var(--red); }}
.scard.cg .scard-val {{ color:var(--green); }}
.scard.cb .scard-val {{ color:#1D4ED8; }}
.scard.ca .scard-val {{ color:var(--amber); }}
.scard.cp .scard-val {{ color:#7B21A8; }}

/* ── Controls ── */
.controls {{
  display:flex; gap:10px; flex-wrap:wrap; align-items:center; margin-bottom:16px;
}}

.s-wrap {{ flex:1; min-width:220px; position:relative; }}
.s-wrap input {{
  width:100%; background:var(--surf); border:1px solid var(--border);
  border-radius:50px; padding:10px 20px 10px 42px;
  font-size:13.5px; font-family:inherit; color:var(--t1); outline:none;
  box-shadow:var(--sh-xs); transition:border-color .2s, box-shadow .2s;
}}
.s-wrap input::placeholder {{ color:var(--t3); }}
.s-wrap input:focus {{ border-color:var(--navy); box-shadow:0 0 0 3px var(--navy-alpha); }}
.s-ico {{ position:absolute; left:14px; top:50%; transform:translateY(-50%); color:var(--t3); font-size:14px; }}

.f-group {{ display:flex; gap:6px; flex-wrap:wrap; }}
.fbtn {{
  padding:8px 15px; border-radius:50px;
  border:1px solid var(--border); background:var(--surf);
  color:var(--t2); font-size:12.5px; font-weight:600;
  cursor:pointer; font-family:inherit; box-shadow:var(--sh-xs);
  transition:all .18s; white-space:nowrap;
}}
.fbtn:hover {{ border-color:var(--navy); color:var(--navy); background:var(--navy-tint); }}
.fbtn.active {{
  background:linear-gradient(135deg,var(--navy),var(--navy-lt));
  border-color:transparent; color:#fff; box-shadow:var(--sh-nav);
}}

/* ── Results bar ── */
.rbar {{
  display:flex; align-items:center; justify-content:space-between;
  margin-bottom:14px; flex-wrap:wrap; gap:8px;
}}
.rinfo {{ font-size:12.5px; color:var(--t2); }}
.rinfo strong {{ color:var(--t1); }}

.expand-btn {{
  padding:6px 14px; border-radius:var(--r0);
  border:1px solid var(--border); background:var(--surf);
  color:var(--t2); font-size:12px; font-weight:600;
  cursor:pointer; font-family:inherit; box-shadow:var(--sh-xs);
  transition:all .18s;
}}
.expand-btn:hover {{ border-color:var(--navy); color:var(--navy); background:var(--navy-tint); }}

/* ════════════════════════════════════════════
   PROJECT ACCORDION
════════════════════════════════════════════ */
.proj-list {{ display:flex; flex-direction:column; gap:10px; }}

.proj-card {{
  background:var(--surf); border:1px solid var(--border);
  border-radius:var(--r1); overflow:hidden;
  box-shadow:var(--sh-xs);
  transition:border-color .18s, box-shadow .18s;
}}
.proj-card:hover {{ border-color:rgba(25,25,112,.28); box-shadow:var(--sh-sm); }}

.proj-head {{
  display:flex; align-items:center; justify-content:space-between;
  padding:14px 18px; cursor:pointer; gap:12px; user-select:none;
  transition:background .14s;
}}
.proj-head:hover {{ background:var(--navy-tint); }}

.ph-left {{ display:flex; align-items:center; gap:12px; flex:1; min-width:0; }}

.ph-num {{
  font-size:9.5px; font-weight:800; letter-spacing:.5px;
  color:var(--navy); background:var(--navy-pale);
  border:1px solid rgba(25,25,112,.18);
  padding:3px 9px; border-radius:50px; flex-shrink:0;
}}
.ph-name {{
  font-size:13.5px; font-weight:700; color:var(--t1);
  overflow:hidden; text-overflow:ellipsis; white-space:nowrap;
}}

.ph-right {{ display:flex; align-items:center; gap:10px; flex-shrink:0; }}

.tc-badge {{
  font-size:11px; color:var(--t2);
  background:var(--surf2); border:1px solid var(--border);
  padding:3px 10px; border-radius:50px; white-space:nowrap;
}}

.ph-prog {{ display:flex; align-items:center; gap:7px; min-width:105px; }}
.ph-bar  {{ flex:1; height:5px; background:var(--border); border-radius:50px; overflow:hidden; }}
.ph-fill {{ height:100%; border-radius:50px; background:linear-gradient(90deg,var(--navy),var(--navy-lt)); }}
.ph-pct  {{ font-size:11px; font-weight:700; color:var(--navy); min-width:32px; text-align:right; }}

.chevron {{ color:var(--t3); font-size:14px; transition:transform .25s; flex-shrink:0; }}
.proj-card.open .chevron {{ transform:rotate(180deg); }}

.task-wrap {{
  max-height:0; overflow:hidden; transition:max-height .35s ease;
  border-top:1px solid transparent;
}}
.proj-card.open .task-wrap {{
  max-height:4000px; border-top-color:var(--divider);
}}

/* task table */
.tt {{ width:100%; border-collapse:collapse; font-size:12.5px; }}
.tt th {{
  padding:8px 14px; text-align:left;
  font-size:10px; font-weight:700; letter-spacing:1px;
  text-transform:uppercase; color:var(--t3);
  background:var(--surf2); white-space:nowrap;
  border-bottom:1px solid var(--border);
}}
.tt td {{ padding:10px 14px; border-top:1px solid var(--divider); vertical-align:middle; }}
.tt tr:hover td {{ background:var(--navy-tint); }}

.t-name {{ font-weight:600; color:var(--t1); }}
.t-pic  {{ color:var(--navy); font-weight:700; font-size:12px; }}
.t-date {{ color:var(--t2); white-space:nowrap; font-size:11.5px; }}
.t-date.ext {{ color:var(--amber); font-weight:600; }}

/* status badge */
.sb {{
  display:inline-flex; align-items:center; gap:4px;
  padding:2px 9px; border-radius:50px;
  font-size:10.5px; font-weight:700; white-space:nowrap;
}}
.sb.done     {{ background:rgba(10,122,82,.1); color:#0A5C3E; border:1px solid rgba(10,122,82,.18); }}
.sb.doing    {{ background:var(--navy-pale);    color:var(--navy); border:1px solid rgba(25,25,112,.18); }}
.sb.late     {{ background:var(--red-pale);     color:var(--red-dk); border:1px solid rgba(186,17,15,.18); }}
.sb.paused   {{ background:rgba(176,118,0,.1); color:var(--amber); border:1px solid rgba(176,118,0,.18); }}
.sb.canceled {{ background:var(--surf2); color:var(--t3); border:1px solid var(--border); }}

/* mini progress */
.c-wrap {{ display:inline-flex; align-items:center; gap:5px; }}
.c-bar  {{ width:50px; height:4px; background:var(--border); border-radius:50px; overflow:hidden; }}
.c-fill {{ height:100%; border-radius:50px; background:linear-gradient(90deg,var(--navy),var(--navy-lt)); }}
.c-txt  {{ font-size:11px; color:var(--t2); }}

/* pagination */
.pagination {{ display:flex; justify-content:center; align-items:center; gap:5px; margin-top:22px; flex-wrap:wrap; }}
.pg {{
  min-width:34px; height:34px; padding:0 10px;
  border-radius:var(--r0); border:1px solid var(--border);
  background:var(--surf); color:var(--t2);
  font-size:13px; cursor:pointer;
  display:flex; align-items:center; justify-content:center;
  font-family:inherit; box-shadow:var(--sh-xs); transition:all .18s;
}}
.pg:hover {{ border-color:var(--navy); color:var(--navy); background:var(--navy-tint); }}
.pg.active {{ background:linear-gradient(135deg,var(--navy),var(--navy-lt)); border-color:transparent; color:#fff; box-shadow:var(--sh-nav); }}
.pg:disabled {{ opacity:.35; cursor:default; }}

/* ════════════════════════════════════════════
   SHOOTING SCHEDULE
════════════════════════════════════════════ */
.shoot-note {{
  display:flex; align-items:flex-start; gap:10px;
  background:rgba(186,17,15,.05); border:1px solid rgba(186,17,15,.18);
  border-left:3px solid var(--red); border-radius:var(--r0);
  padding:12px 16px; font-size:12.5px; color:var(--t2); margin-bottom:20px;
}}

.shoot-grid {{
  display:grid; grid-template-columns:repeat(auto-fill,minmax(300px,1fr)); gap:14px;
}}

.scard-item {{
  background:var(--surf); border:1px solid var(--border);
  border-radius:var(--r1); overflow:hidden; box-shadow:var(--sh-xs);
  transition:transform .2s, box-shadow .2s, border-color .2s;
  animation:slideUp .38s ease both;
}}
.scard-item:hover {{ transform:translateY(-4px); box-shadow:var(--sh-md); border-color:rgba(186,17,15,.2); }}

.sc-head {{
  display:flex; align-items:center; gap:12px;
  padding:14px 16px;
  background:linear-gradient(135deg,var(--navy-tint),var(--red-tint));
  border-bottom:1px solid var(--border);
}}

.sc-num {{
  width:32px; height:32px; border-radius:var(--r0);
  background:linear-gradient(135deg,var(--navy),var(--navy-lt));
  display:flex; align-items:center; justify-content:center;
  font-size:12px; font-weight:800; color:#fff; flex-shrink:0;
  box-shadow:var(--sh-nav);
}}
.sc-title {{ font-size:13px; font-weight:700; color:var(--t1); flex:1; line-height:1.3; }}
.sc-status {{ flex-shrink:0; }}

.sc-body {{ padding:14px 16px; display:flex; flex-direction:column; gap:10px; }}
.sc-row  {{ display:flex; align-items:flex-start; gap:10px; font-size:12.5px; }}
.sc-icon {{ font-size:13px; flex-shrink:0; margin-top:1px; }}
.sc-lbl  {{ font-size:10.5px; font-weight:700; letter-spacing:.7px; text-transform:uppercase; color:var(--t3); min-width:58px; }}
.sc-val  {{ color:var(--t2); flex:1; }}

.dt-chip {{
  display:inline-flex; align-items:center; gap:6px;
  background:var(--navy-pale); border:1px solid rgba(25,25,112,.18);
  padding:4px 12px; border-radius:50px;
  font-size:12px; font-weight:700; color:var(--navy);
}}
.pic-chips {{ display:flex; flex-wrap:wrap; gap:6px; }}
.pic-chip {{
  font-size:11.5px; font-weight:700; padding:3px 10px; border-radius:50px;
  background:var(--red-pale); border:1px solid rgba(186,17,15,.18); color:var(--red-dk);
}}

/* status inline badges */
.st-del {{ background:rgba(10,122,82,.1); color:#0A5C3E; border:1px solid rgba(10,122,82,.18); }}
.st-pnd {{ background:rgba(176,118,0,.1); color:var(--amber); border:1px solid rgba(176,118,0,.18); }}
.st-tbu {{ background:var(--surf2); color:var(--t3); border:1px solid var(--border); }}

.st-del, .st-pnd, .st-tbu {{
  display:inline-flex; align-items:center; gap:4px;
  padding:3px 10px; border-radius:50px;
  font-size:10.5px; font-weight:700; white-space:nowrap;
}}

/* ════════════════════════════════════════════
   VIDEO CHECKLIST
════════════════════════════════════════════ */
.vc-wrap {{ border:1px solid var(--border); border-radius:var(--r1); overflow:hidden; overflow-x:auto; box-shadow:var(--sh-sm); }}

.vc-table {{ width:100%; border-collapse:collapse; font-size:12.5px; min-width:900px; }}

.vc-table thead tr {{ background:linear-gradient(135deg,var(--navy-dk),var(--navy)); }}
.vc-table th {{
  padding:12px 14px; text-align:left;
  font-size:10px; font-weight:800; letter-spacing:1px;
  text-transform:uppercase; color:rgba(255,255,255,.7);
  white-space:nowrap;
}}

.vc-table td {{
  padding:10px 14px; border-top:1px solid var(--divider);
  vertical-align:middle; background:var(--surf);
}}
.vc-table tr:hover td {{ background:var(--navy-tint); }}

.vc-stt  {{ color:var(--t3); font-size:11px; text-align:center; }}
.vc-da   {{ font-weight:700; color:var(--navy); }}
.vc-vid  {{ font-weight:500; color:var(--t1); }}
.vc-pic  {{ color:var(--red); font-weight:700; }}
.vc-date {{ color:var(--t2); font-size:11.5px; white-space:nowrap; }}
.vc-dl   {{ color:var(--red-dk); font-size:11.5px; white-space:nowrap; font-weight:600; }}
.vc-done {{ color:var(--green); font-size:11.5px; white-space:nowrap; }}

.prog-line {{ display:flex; align-items:center; gap:6px; min-width:88px; }}
.prog-line .pb {{ flex:1; height:5px; background:var(--border); border-radius:50px; overflow:hidden; }}
.prog-line .pf {{ height:100%; border-radius:50px; }}
.prog-line .pt {{ font-size:11px; color:var(--t2); min-width:28px; text-align:right; font-weight:600; }}

.vs-done {{ background:rgba(10,122,82,.1); color:#0A5C3E; border:1px solid rgba(10,122,82,.18); }}
.vs-doing {{ background:var(--navy-pale); color:var(--navy); border:1px solid rgba(25,25,112,.18); }}
.vs-pend  {{ background:var(--surf2); color:var(--t3); border:1px solid var(--border); }}

.vs-done,.vs-doing,.vs-pend {{
  display:inline-flex; align-items:center; gap:4px;
  padding:2px 9px; border-radius:50px;
  font-size:10.5px; font-weight:700; white-space:nowrap;
}}

/* ════════════════════════════════════════════
   FOOTER
════════════════════════════════════════════ */
footer {{
  background:linear-gradient(135deg,var(--navy-dk),var(--navy));
  padding:28px 40px;
  text-align:center; color:rgba(255,255,255,.5); font-size:12.5px;
  margin-top:48px;
}}
footer strong {{ color:rgba(255,255,255,.8); }}

/* ── No results ── */
.no-res {{ text-align:center; padding:56px 20px; color:var(--t3); font-size:15px; }}
.no-res .nr-icon {{ font-size:48px; display:block; margin-bottom:12px; }}

/* ── Responsive ── */
@media(max-width:900px) {{
  .ph-prog {{ display:none; }}
  .tt th:nth-child(n+4),.tt td:nth-child(n+4) {{ display:none; }}
  .tab-btn .tb-txt {{ display:none; }}
  .hero {{ padding:30px 20px; }}
  .kpi-row,.content,.tabs-wrap {{ padding-left:20px; padding-right:20px; }}
}}
@media(max-width:640px) {{
  .hero-right {{ display:none; }}
  .shoot-grid {{ grid-template-columns:1fr; }}
}}
</style>
</head>
<body>
<div class="page">

<!-- ═══════ HERO HEADER ═══════ -->
<div class="hero">
  <div class="hero-inner">
    <div class="hero-left">
      <div class="hero-kicker"><span class="hero-dot"></span>Tháng 5 · 2026 · Team Media</div>
      <h1 class="hero-title">BÁO CÁO CÔNG VIỆC<br><span class="hl">THÁNG 5 / 2026</span></h1>
      <div class="hero-sub">
        <span>📂 Team Media</span><span class="sep">·</span>
        <span>🏢 Phòng Marketing</span><span class="sep">·</span>
        <span>📍 HIAS Master · CB</span>
      </div>
    </div>
    <div class="hero-right">
      <div class="period-card">
        <div class="pc-icon">📅</div>
        <div>
          <div class="pc-lbl">Kỳ báo cáo</div>
          <div class="pc-val">01/05/2026 → 31/05/2026</div>
        </div>
      </div>
      <div class="hero-time">🕐 Xuất: {now}</div>
    </div>
  </div>
</div>
<div class="accent-stripe"></div>

<!-- ═══════ KPI ROW ═══════ -->
<div class="kpi-row">
  <div class="kpi-card k-task">
    <div class="kpi-icon">🎨</div>
    <div class="kpi-info">
      <div class="kpi-num">{ttasks}</div>
      <div class="kpi-lbl">Task thiết kế tháng 5</div>
      <div class="kpi-chips">
        <span class="chip chip-g">✅ {tdone} Hoàn thành</span>
        <span class="chip chip-n">🔵 {ttasks-tdone} Đang làm</span>
      </div>
    </div>
  </div>
  <div class="kpi-card k-shoot">
    <div class="kpi-icon">📷</div>
    <div class="kpi-info">
      <div class="kpi-num">{len(shooting)}</div>
      <div class="kpi-lbl">Lịch quay chụp tháng 5</div>
      <div class="kpi-chips">
        <span class="chip chip-g">✅ {sdone} Đã bàn giao</span>
        <span class="chip chip-a">⏳ {len(shooting)-sdone} Chờ/TBU</span>
      </div>
    </div>
  </div>
  <div class="kpi-card k-video">
    <div class="kpi-icon">🎬</div>
    <div class="kpi-info">
      <div class="kpi-num">{len(videos)}</div>
      <div class="kpi-lbl">Video Checklist tháng 5</div>
      <div class="kpi-chips">
        <span class="chip chip-g">✅ {vdone} Hoàn thành</span>
        <span class="chip chip-n">🔵 {vdoing} Đang làm</span>
        <span class="chip chip-a">⏳ {len(videos)-vdone-vdoing} Chờ</span>
      </div>
    </div>
  </div>
</div>

<!-- ═══════ TABS ═══════ -->
<div class="tabs-wrap">
  <div class="tab-nav">
    <button class="tab-btn active" id="tab-tasks" onclick="sw('tasks')">
      🎨 <span class="tb-txt">Công việc chính</span> <span class="tbadge">{ttasks}</span>
    </button>
    <button class="tab-btn t-red" id="tab-shoot" onclick="sw('shoot')">
      📷 <span class="tb-txt">Lịch quay chụp</span> <span class="tbadge">{len(shooting)}</span>
    </button>
    <button class="tab-btn t-purple" id="tab-video" onclick="sw('video')">
      🎬 <span class="tb-txt">Video Checklist</span> <span class="tbadge">{len(videos)}</span>
    </button>
  </div>
</div>

<!-- ═══════ TAB: TASKS ═══════ -->
<div class="tab-panel active content" id="panel-tasks">
  <div class="scard-grid" id="tStats"></div>
  <div class="controls">
    <div class="s-wrap"><span class="s-ico">🔍</span><input type="text" id="tSearch" placeholder="Tìm dự án, công việc, nhân sự..."></div>
    <div class="f-group" id="tFilters">
      <button class="fbtn active" data-f="all">Tất cả</button>
      <button class="fbtn" data-f="Hoàn thành">✅ Hoàn thành</button>
      <button class="fbtn" data-f="Đang làm">🔵 Đang làm</button>
      <button class="fbtn" data-f="Hoãn">⏸ Hoãn</button>
    </div>
  </div>
  <div class="rbar">
    <div class="rinfo" id="tInfo"></div>
    <button class="expand-btn" id="expandBtn" onclick="toggleAll()">⤢ Mở rộng tất cả</button>
  </div>
  <div class="proj-list" id="projList"></div>
  <div class="pagination" id="tPager"></div>
</div>

<!-- ═══════ TAB: SHOOT ═══════ -->
<div class="tab-panel content" id="panel-shoot">
  <div class="scard-grid" id="sStats"></div>
  <div class="shoot-note">
    <span style="font-size:16px">⚠️</span>
    <div><strong style="color:var(--red)">Lưu ý:</strong> Source hình ảnh bàn giao trong <strong>1 ngày</strong> kể từ khi kết thúc sự kiện.</div>
  </div>
  <div class="controls">
    <div class="s-wrap"><span class="s-ico">🔍</span><input type="text" id="sSearch" placeholder="Tìm sự kiện, địa điểm, phụ trách..."></div>
    <div class="f-group" id="sFilters">
      <button class="fbtn active" data-f="all">Tất cả</button>
      <button class="fbtn" data-f="del">✅ Đã bàn giao</button>
      <button class="fbtn" data-f="pnd">⏳ Chưa / TBU</button>
    </div>
  </div>
  <div class="rbar"><div class="rinfo" id="sInfo"></div></div>
  <div class="shoot-grid" id="shootGrid"></div>
</div>

<!-- ═══════ TAB: VIDEO ═══════ -->
<div class="tab-panel content" id="panel-video">
  <div class="scard-grid" id="vStats"></div>
  <div class="controls">
    <div class="s-wrap"><span class="s-ico">🔍</span><input type="text" id="vSearch" placeholder="Tìm dự án, video, kịch bản, PIC..."></div>
    <div class="f-group" id="vFilters">
      <button class="fbtn active" data-f="all">Tất cả</button>
      <button class="fbtn" data-f="Hoàn thành">✅ Hoàn thành</button>
      <button class="fbtn" data-f="Đang làm">🔵 Đang làm</button>
      <button class="fbtn" data-f="pend">⏳ Chờ xử lý</button>
    </div>
  </div>
  <div class="rbar"><div class="rinfo" id="vInfo"></div></div>
  <div class="vc-wrap">
    <table class="vc-table">
      <thead><tr>
        <th style="text-align:center">#</th>
        <th>Dự án / CN</th><th>Video</th><th>Kịch bản</th>
        <th>Nhận order</th><th>Ngày quay</th><th>Deadline</th>
        <th>PIC</th><th>Tiến độ</th><th>Tình trạng</th><th>Hoàn thành</th>
      </tr></thead>
      <tbody id="vBody"></tbody>
    </table>
  </div>
</div>

<!-- ═══════ FOOTER ═══════ -->
<footer>
  <p>📊 Báo cáo tháng 5/2026 &nbsp;·&nbsp; <strong>01/05 – 31/05/2026</strong> &nbsp;·&nbsp; Team Media · CB Media Production Team · HIAS Master</p>
  <p style="margin-top:6px">Xuất lúc: {now}</p>
</footer>

</div><!-- /page -->

<script>
// ══════════════════════════════
// DATA
// ══════════════════════════════
const PROJ  = {pj};
const SHOOT = {sj};
const VIDS  = {vj};

// ══════════════════════════════
// TABS
// ══════════════════════════════
function sw(id) {{
  ['tasks','shoot','video'].forEach(t => {{
    document.getElementById('tab-'+t).classList.remove('active');
    document.getElementById('panel-'+t).classList.remove('active');
  }});
  document.getElementById('tab-'+id).classList.add('active');
  document.getElementById('panel-'+id).classList.add('active');
}}

// ══════════════════════════════
// HELPERS
// ══════════════════════════════
function sc(state,cancel) {{
  if (cancel) return 'canceled';
  if (!state) return 'doing';
  if (state.includes('Hoàn thành')) return 'done';
  if (state.includes('Hoãn'))       return 'paused';
  if (state.includes('Quá hạn'))    return 'late';
  return 'doing';
}}
function sl(state,cancel) {{
  if (cancel) return '🚫 '+cancel;
  if (!state) return '🔵 Đang làm';
  if (state.includes('Hoàn thành')) return '✅ Hoàn thành';
  if (state.includes('Hoãn'))       return '⏸ Hoãn';
  if (state.includes('Quá hạn'))    return '🔴 Quá hạn';
  return '🔵 Đang làm';
}}
function pct(p) {{
  if (!p.tasks.length) return 0;
  return Math.round(p.tasks.filter(t=>sc(t.state,t.cancel)==='done').length/p.tasks.length*100);
}}
function fd(d) {{
  if (!d) return '–';
  const pts=d.split('-');
  return pts.length===3?`${{pts[2]}}/${{pts[1]}}`:d;
}}

// ══════════════════════════════
// TASK SECTION
// ══════════════════════════════
const PAGE=20; let tFil='all',tSrch='',tPage=1,tFiltd=[],allOpen=false;

function renderTaskStats() {{
  let tot=0,done=0,doing=0,paused=0;
  PROJ.forEach(p=>p.tasks.forEach(t=>{{
    tot++;
    const s=sc(t.state,t.cancel);
    if(s==='done')done++;
    else if(s==='doing')doing++;
    else if(s==='paused'||s==='canceled')paused++;
  }}));
  const p=tot?Math.round(done/tot*100):0;
  document.getElementById('tStats').innerHTML=[
    {{v:PROJ.length,l:'Dự án',        c:'cn',i:'📁'}},
    {{v:tot,        l:'Công việc',     c:'cb',i:'📋'}},
    {{v:done,       l:'Hoàn thành',    c:'cg',i:'✅'}},
    {{v:doing,      l:'Đang thực hiện',c:'cn',i:'🔵'}},
    {{v:paused,     l:'Hoãn / Huỷ',   c:'ca',i:'⏸'}},
    {{v:p+'%',      l:'Tỉ lệ HT',     c:'cp',i:'📈'}},
  ].map(s=>`<div class="scard ${{s.c}}"><div class="scard-icon">${{s.i}}</div><div class="scard-val">${{s.v}}</div><div class="scard-lbl">${{s.l}}</div></div>`).join('');
}}

function getFiltd() {{
  return PROJ.filter(p=>{{
    const mF=tFil==='all'||p.tasks.some(t=>(t.state||'').includes(tFil));
    const q=tSrch.toLowerCase();
    const mS=!q||p.name.toLowerCase().includes(q)||p.tasks.some(t=>(t.name||'').toLowerCase().includes(q)||(t.pic||'').toLowerCase().includes(q));
    return mF&&mS;
  }});
}}

function renderProjs() {{
  tFiltd=getFiltd();
  const totT=tFiltd.reduce((s,p)=>s+p.tasks.length,0);
  document.getElementById('tInfo').innerHTML=`Hiển thị <strong>${{tFiltd.length}}</strong> dự án · <strong>${{totT}}</strong> công việc`;
  const start=(tPage-1)*PAGE, page=tFiltd.slice(start,start+PAGE);
  document.getElementById('projList').innerHTML=page.length
    ?page.map((p,i)=>renderPC(p,start+i)).join('')
    :'<div class="no-res"><span class="nr-icon">🔍</span><p>Không tìm thấy kết quả</p></div>';
  renderTPager(tFiltd.length);
}}

function renderPC(p,idx) {{
  const pct2=pct(p),done=p.tasks.filter(t=>sc(t.state,t.cancel)==='done').length;
  return `<div class="proj-card" id="pc-${{idx}}">
    <div class="proj-head" onclick="toggleP(${{idx}})">
      <div class="ph-left">
        <span class="ph-num">#${{p.id||idx+1}}</span>
        <span class="ph-name" title="${{p.name}}">${{p.name}}</span>
      </div>
      <div class="ph-right">
        <span class="tc-badge">${{done}}/${{p.tasks.length}} việc</span>
        <div class="ph-prog"><div class="ph-bar"><div class="ph-fill" style="width:${{pct2}}%"></div></div><span class="ph-pct">${{pct2}}%</span></div>
        <span class="chevron">▾</span>
      </div>
    </div>
    <div class="task-wrap">
      <table class="tt">
        <thead><tr><th>Công việc</th><th>PIC</th><th>Bắt đầu</th><th>Kết thúc</th><th>Trạng thái</th><th>% HT</th></tr></thead>
        <tbody>
          ${{p.tasks.map(t=>`<tr>
            <td class="t-name">${{t.name||''}}</td>
            <td class="t-pic">${{t.pic||'–'}}</td>
            <td class="t-date">${{fd(t.start)}}</td>
            <td class="t-date${{t.extended?' ext':''}}">${{fd(t.extended||t.end)}}${{t.extended?' ⤴':''}}</td>
            <td><span class="sb ${{sc(t.state,t.cancel)}}">${{sl(t.state,t.cancel)}}</span></td>
            <td><div class="c-wrap"><div class="c-bar"><div class="c-fill" style="width:${{Math.round((t.completion||0)*100)}}%"></div></div><span class="c-txt">${{Math.round((t.completion||0)*100)}}%</span></div></td>
          </tr>`).join('')}}
        </tbody>
      </table>
    </div>
  </div>`;
}}

function toggleP(idx) {{ document.getElementById('pc-'+idx).classList.toggle('open'); }}
function toggleAll() {{
  allOpen=!allOpen;
  document.querySelectorAll('#projList .proj-card').forEach(c=>c.classList.toggle('open',allOpen));
  document.getElementById('expandBtn').textContent=allOpen?'⤡ Thu gọn tất cả':'⤢ Mở rộng tất cả';
}}

function renderTPager(tot) {{
  const pages=Math.ceil(tot/PAGE);
  if(pages<=1){{document.getElementById('tPager').innerHTML='';return;}}
  let h='';
  if(tPage>1)h+=`<button class="pg" onclick="goTP(${{tPage-1}})">←</button>`;
  for(let p=1;p<=pages;p++){{
    if(p===1||p===pages||Math.abs(p-tPage)<=1)
      h+=`<button class="pg${{p===tPage?' active':''}}" onclick="goTP(${{p}})">${{p}}</button>`;
    else if(h.slice(-10)!=='>…</button>')h+=`<button class="pg" disabled>…</button>`;
  }}
  if(tPage<pages)h+=`<button class="pg" onclick="goTP(${{tPage+1}})">→</button>`;
  document.getElementById('tPager').innerHTML=h;
}}
function goTP(p){{tPage=p;renderProjs();window.scrollTo({{top:0,behavior:'smooth'}});}}

// ══════════════════════════════
// SHOOT SECTION
// ══════════════════════════════
let sFil='all', sSrch='';

function renderShootStats() {{
  const tot=SHOOT.length, del=SHOOT.filter(s=>s.tt==='Đã bàn giao').length;
  document.getElementById('sStats').innerHTML=[
    {{v:tot,          l:'Lịch quay chụp',   c:'cr',i:'📷'}},
    {{v:del,          l:'Đã bàn giao',       c:'cg',i:'✅'}},
    {{v:tot-del,      l:'Chờ / TBU',         c:'ca',i:'⏳'}},
  ].map(s=>`<div class="scard ${{s.c}}"><div class="scard-icon">${{s.i}}</div><div class="scard-val">${{s.v}}</div><div class="scard-lbl">${{s.l}}</div></div>`).join('');
}}

function getSFiltd() {{
  return SHOOT.filter(s=>{{
    const mF=sFil==='all'||(sFil==='del'&&s.tt==='Đã bàn giao')||(sFil==='pnd'&&s.tt!=='Đã bàn giao');
    const q=sSrch.toLowerCase();
    const mS=!q||s.ten.toLowerCase().includes(q)||s.dd.toLowerCase().includes(q)||s.pt.toLowerCase().includes(q)||s.yc.toLowerCase().includes(q);
    return mF&&mS;
  }});
}}

function renderShoot() {{
  const list=getSFiltd();
  document.getElementById('sInfo').innerHTML=`Hiển thị <strong>${{list.length}}</strong> / ${{SHOOT.length}} lịch`;
  if(!list.length){{document.getElementById('shootGrid').innerHTML='<div class="no-res" style="grid-column:1/-1"><span class="nr-icon">📷</span><p>Không tìm thấy</p></div>';return;}}
  document.getElementById('shootGrid').innerHTML=list.map((s,i)=>{{
    const stBadge=s.tt==='Đã bàn giao'
      ?'<span class="st-del">✅ Đã bàn giao</span>'
      :s.ngay==='TBU'
      ?'<span class="st-tbu">📌 TBU</span>'
      :'<span class="st-pnd">⏳ Chờ thực hiện</span>';
    const picH=s.pt?s.pt.split(',').map(n=>`<span class="pic-chip">👤 ${{n.trim()}}</span>`).join('')
                   :'<span style="color:var(--t3);font-size:12px">Chưa phân công</span>';
    return `<div class="scard-item" style="animation-delay:${{i*.04}}s">
      <div class="sc-head">
        <div class="sc-num">${{s.stt}}</div>
        <div class="sc-title">${{s.ten}}</div>
        <div class="sc-status">${{stBadge}}</div>
      </div>
      <div class="sc-body">
        <div class="sc-row"><span class="sc-icon">📌</span><div><div class="sc-lbl">Yêu cầu</div><div class="sc-val">${{s.yc}}</div></div></div>
        <div class="sc-row"><span class="sc-icon">📍</span><div><div class="sc-lbl">Địa điểm</div><div class="sc-val">${{s.dd}}</div></div></div>
        <div class="sc-row"><span class="sc-icon">🗓</span><div><div class="sc-lbl">Thời gian</div><span class="dt-chip">${{s.ngay}}/2026 · ${{s.gio}} (${{s.thu}})</span></div></div>
        <div class="sc-row"><span class="sc-icon">👤</span><div><div class="sc-lbl">Phụ trách</div><div class="pic-chips">${{picH}}</div></div></div>
      </div>
    </div>`;
  }}).join('');
}}

// ══════════════════════════════
// VIDEO SECTION
// ══════════════════════════════
let vFil='all', vSrch='';

function renderVideoStats() {{
  const tot=VIDS.length,done=VIDS.filter(v=>v.tt==='Hoàn thành').length,doing=VIDS.filter(v=>v.tt==='Đang làm').length;
  const p=tot?Math.round(done/tot*100):0;
  document.getElementById('vStats').innerHTML=[
    {{v:tot,   l:'Tổng video',   c:'cp',i:'🎬'}},
    {{v:done,  l:'Hoàn thành',   c:'cg',i:'✅'}},
    {{v:doing, l:'Đang xử lý',   c:'cn',i:'🔵'}},
    {{v:tot-done-doing,l:'Chờ xử lý',c:'ca',i:'⏳'}},
    {{v:p+'%', l:'Tỉ lệ HT',     c:'cr',i:'📈'}},
  ].map(s=>`<div class="scard ${{s.c}}"><div class="scard-icon">${{s.i}}</div><div class="scard-val">${{s.v}}</div><div class="scard-lbl">${{s.l}}</div></div>`).join('');
}}

function getVFiltd() {{
  return VIDS.filter(v=>{{
    const mF=vFil==='all'
      ||(vFil==='Hoàn thành'&&v.tt==='Hoàn thành')
      ||(vFil==='Đang làm'  &&v.tt==='Đang làm')
      ||(vFil==='pend'      &&v.tt!=='Hoàn thành'&&v.tt!=='Đang làm');
    const q=vSrch.toLowerCase();
    const mS=!q||v.da.toLowerCase().includes(q)||v.vid.toLowerCase().includes(q)||(v.kb||'').toLowerCase().includes(q)||(v.pic||'').toLowerCase().includes(q);
    return mF&&mS;
  }});
}}

function renderVideos() {{
  const list=getVFiltd();
  document.getElementById('vInfo').innerHTML=`Hiển thị <strong>${{list.length}}</strong> / ${{VIDS.length}} video`;
  const pc=p=>p===100?'linear-gradient(90deg,#0A7A52,#10B981)':p>0?'linear-gradient(90deg,var(--navy),var(--navy-lt))':'var(--border)';
  document.getElementById('vBody').innerHTML=list.map(v=>{{
    const cls=v.tt==='Hoàn thành'?'vs-done':v.tt==='Đang làm'?'vs-doing':'vs-pend';
    const lbl=v.tt==='Hoàn thành'?'✅ Hoàn thành':v.tt==='Đang làm'?'🔵 Đang làm':v.tt||'⏳ Chờ xử lý';
    return `<tr>
      <td class="vc-stt">${{v.stt}}</td>
      <td class="vc-da">${{v.da}}</td>
      <td class="vc-vid">${{v.vid}}</td>
      <td style="color:var(--t2);font-size:11.5px">${{v.kb||'–'}}</td>
      <td class="vc-date">${{v.ord||'–'}}</td>
      <td class="vc-date">${{v.quay||'–'}}</td>
      <td class="vc-dl">${{v.dl||'–'}}</td>
      <td class="vc-pic">${{v.pic||'–'}}</td>
      <td><div class="prog-line"><div class="pb"><div class="pf" style="width:${{v.pct}}%;background:${{pc(v.pct)}}"></div></div><span class="pt">${{v.pct}}%</span></div></td>
      <td><span class="${{cls}}">${{lbl}}</span></td>
      <td class="vc-done">${{v.done||'–'}}</td>
    </tr>`;
  }}).join('');
}}

// ══════════════════════════════
// INIT
// ══════════════════════════════
document.addEventListener('DOMContentLoaded',()=>{{
  renderTaskStats(); renderShootStats(); renderVideoStats();
  renderProjs(); renderShoot(); renderVideos();

  // task filters
  document.querySelectorAll('#tFilters .fbtn').forEach(b=>b.addEventListener('click',()=>{{
    document.querySelectorAll('#tFilters .fbtn').forEach(x=>x.classList.remove('active'));
    b.classList.add('active'); tFil=b.dataset.f; tPage=1; renderProjs();
  }}));
  let tt; document.getElementById('tSearch').addEventListener('input',e=>{{
    clearTimeout(tt); tt=setTimeout(()=>{{tSrch=e.target.value.trim();tPage=1;renderProjs();}},220);
  }});

  // shoot filters
  document.querySelectorAll('#sFilters .fbtn').forEach(b=>b.addEventListener('click',()=>{{
    document.querySelectorAll('#sFilters .fbtn').forEach(x=>x.classList.remove('active'));
    b.classList.add('active'); sFil=b.dataset.f; renderShoot();
  }}));
  let st; document.getElementById('sSearch').addEventListener('input',e=>{{
    clearTimeout(st); st=setTimeout(()=>{{sSrch=e.target.value.trim();renderShoot();}},220);
  }});

  // video filters
  document.querySelectorAll('#vFilters .fbtn').forEach(b=>b.addEventListener('click',()=>{{
    document.querySelectorAll('#vFilters .fbtn').forEach(x=>x.classList.remove('active'));
    b.classList.add('active'); vFil=b.dataset.f; renderVideos();
  }}));
  let vt; document.getElementById('vSearch').addEventListener('input',e=>{{
    clearTimeout(vt); vt=setTimeout(()=>{{vSrch=e.target.value.trim();renderVideos();}},220);
  }});
}});
</script>
</body>
</html>'''

with open('2026_T5_MEDIA_LIGHT.html', 'w', encoding='utf-8') as f:
    f.write(HTML)

print(f"Done! Projects: {len(projects)} | Tasks: {ttasks} | Done: {tdone}")
print(f"Shooting: {len(shooting)} | Delivered: {sdone}")
print(f"Videos: {len(videos)} | Done: {vdone} | Doing: {vdoing}")
print("Saved: 2026_T5_MEDIA_LIGHT.html")
